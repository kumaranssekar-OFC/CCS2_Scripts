#################################################################################################################
#
# FILE:         create_prod_dir_ccs2.py
# DESCRIPTION:  This script can be used to trigger the production folder preparation which involves the copy of the artifacts and the document creation for CCS2 project
# USAGE:        see help_text
# PREREQUISITE: The files CCS2_SPL_TMPL_V01.00.xlsx, SOS_pdf_ccs2.xsl, xml2odxe.pl has to be in the same directory as this script
# HISTORY:
# Date         | Author          		| Modification
# 29.08.2024   | Nisharani C  			| Initial version
# 18.09.2024   | Nisharani C  			| Adaptation for new SAP request template - row number changes
# 24.10.2024   | Nisharani C			| Adaptation to update TM diag version to SOS based on version	
# 13.11.2024   | Nisharani C            | Adaptation for new DB Keys FLW4LINUX and DTV FW file
# 02.01.2024   | Nisharani C            | DTV changes 
# 10.01.2024   | Nisharani C            | Avoid DTV files download for Non DTV PNs
# 20.02.2025   | Nisharani C			| Temporary change to fix the ssh certificate verification error during download	
# 27.02.2025   | Nisharani C			| set ublox diag verion as balnk for EU 
# 24.04.2025   | Nisharani C            | Cosmetic changes
# 28.05.2025   | Nisharani C            | updated the url formation for non-gas appsw based on non-gas appsw link itself 
# 24.06.2025   | Abinaya M				| Sync with change in set definition tmpl (Col name ublox to gnss)	
# 24.06.2025   | Nisharani C            | change to include Supplier feed section related to "oemcrypto" in SOS 
# 21.08.2025   | Nisharani C  			| Change to have the SPL include plant name based on info from VM
# 08.09.2025   | Nisharani C            | Added verify_md5sum to verify the md5sum of the artifacts after copy
# 25.11.2025   | Nisharani C            | bug fixes
# 4.12.2025    | Nisharani C            | Changes to include new DBKey ASS_App_Android_Version into the SOS
# 9.12.2025    | Nisharani C            | Bug fixes
# 02.02.2026   | Kumaran Sekar          | Added a function "download from sharepoint" to download the files from sharepoint.
#################################################################################################################

import os
import sys
import re
import optparse
import requests
import subprocess
import pythoncom
import win32com.client
import pandas as pd # for dataframes
import xml.etree.ElementTree as ET
import xml.etree.ElementTree as ET1
import xml.etree.ElementTree as ETS
import xml.etree.ElementTree as ETOS
 
import openpyxl 
import numpy as np
#import hashlib
import shutil
#from datetime import date 
#import datetime
from datetime import datetime as dt
from requests_kerberos import HTTPKerberosAuth, OPTIONAL
from shutil import copy
from artifactory import ArtifactoryPath
from openpyxl import Workbook
from win32com import client         #for excel to pdf generation during spl call
#from xlsx2pdf import xlsx2pdf
import requests.packages.urllib3 as urllib3
urllib3.disable_warnings()

global _md5_dict
global _md5_dest_dict


fop_path = r'%fop_Sh%'
# fop_path = r'C:\ccstg\amb5cob_DI_TOOLS_FI_TOOLS_13.0V01_2.vws\di_tools\java\fop-0.20.5\fop.bat'
_aurix_header_flag = "notset"
_PD_Tooling_flag = "notset"
_CTS_Parameter_flag = "notset"
_App_Parameter_flag = "notset"
_Sec_Parameter_flag = "notset"
_VIP_Parameter_flag = "notset"
_md5_dict = {}
_md5_dest_dict = {}


def read_xml(_master_xml):
	pass
    
#def read_sap_sheet(_xls_File, _PN):
    # _xls_File = _xls_File
    # _PN = _PN
 
    
def read_sap_sheet():  
   # _PN_count = _PN_count
    global _PN_list
    global _ecn
    _sheet1 = "SAP Partnumbers"
    _sheet2 = "HW Mapping"
    
    _ecn_row =  3
    _ecr_row = 4 
    _cont_row = 6
    _setnum_row = 7
    _docnum_row = 8
    _Pnum_row = 9
    _UFS_blank_row = 12
    _scc_prog_row = 13
    _scc_blank_row = 14
    _ublox_prog_row = 15
    _ublox_blank_row = 16
    
    # _cts_pn_row = 21
    # _appsw_pn_row = 22
    # _aurix_srec_pn_row = 26
    # _aurix_dnl_pn_row = 27
    # _ublox_pn_row = 29
    
    _cts_pn_row = 19
    _appsw_pn_row = 20
    _aurix_srec_pn_row = 22
    _aurix_dnl_pn_row = 23
    _ublox_pn_row = 25
    _sxm_pn_row = 27
    _dtv_pn_row = 29
    _lont_pn_row = 31
    
    _elec_pn_row = 4
    _elec_pn_col = 6
    
    #may not require the below   
    _space_btn_sets = 3 
    _space_to_minor = 2  #may not need now as we do not have minor PNs concept now
   
    print ("\nat read_sap_sheet .." )
    
    wb =  openpyxl.load_workbook(_xls_File, data_only=True)     #load the work book , data_only flag helps to get the value instead of the formaula from a cell in the excel
    ws =  wb['SAP Partnumbers']
    ws2 =  wb['HW Mapping']    
    # print ("col count :",ws.columns)
    _row_start = 6
    _col_start = 3
    
    
    _r_max = ws.max_row + 1
    _c_max = ws.max_column + 1
    _i = 0
    # print("max rows: ",ws.max_row)
    # print("max cols: ",ws.max_column)  # max_column might not work if there was any entry and later on deleted, in that case pls delete the col and save
    
    _PNcount_check = _c_max - _PN_count
    if _PNcount_check != _PN_count:
        print ("\n please check PN count differes in SAP sheet " ) # throw error or log
    _PN_list = []
    #be careful to delete the formulas in the greyed out cells of the sap sheet
    #if there are formaulas in geryed cells, it would resolve to 0 and the list generated would have that value
    #take care of this for sap sheet creation
    for c in range (_col_start,_c_max):
        #print ("\n at column no: ", c)
        _i += 1
        _ecn = ws.cell(row=_ecn_row, column=_col_start)
        #print ("\n ecn :", _ecn.value )
        _ecr_s = ws.cell(row=_ecr_row, column=c)    #ecr from sap kept for reference, will be taken from xml
        for r in range(_row_start,_r_max):
            #val = ws.cell(row=r, column=_col_start)
            #print(val.value)
            
            _cont = ws.cell(row=_cont_row, column=c)
            _setnum = ws.cell(row=_setnum_row, column=c)    
            _docnum = ws.cell(row=_docnum_row, column=c)
            _Pnum = ws.cell(row=_Pnum_row, column=c)
            _UFS_b = ws.cell(row=_UFS_blank_row, column=c)  # ufs blank device number
            _scc_p = ws.cell(row=_scc_prog_row, column=c)   # scc programmed device number
            _scc_b = ws.cell(row=_scc_blank_row, column=c)
            _ublox_p = ws.cell(row=_ublox_prog_row, column=c)
            _ublox_b = ws.cell(row=_ublox_blank_row, column=c)
            _cts_i = ws.cell(row=_cts_pn_row, column=c)         # cts image file number
            _appsw_i = ws.cell(row=_appsw_pn_row, column=c)
            _a_srec_i = ws.cell(row=_aurix_srec_pn_row, column=c)
            _a_dnl_i = ws.cell(row=_aurix_dnl_pn_row, column=c)
            _ublox_i = ws.cell(row=_ublox_pn_row, column=c)
            _sxm_i = ws.cell(row=_sxm_pn_row, column=c)
            _dtv_i = ws.cell(row=_dtv_pn_row, column=c)
            _lont_i = ws.cell(row=_lont_pn_row, column=c)
            
            _elecPN = ws2.cell (row=_elec_pn_row,column=_elec_pn_col)
            
                
        if _ecn.value == None:
            print ("\n ecn is None, ecr value be set from xml ")
        
        _PN_list.append ([_ecn.value,_cont.value,_setnum.value,_docnum.value,_Pnum.value,_UFS_b.value,_scc_p.value,_scc_b.value,_ublox_p.value,_ublox_b.value,_cts_i.value,_appsw_i.value,_a_srec_i.value,_a_dnl_i.value,_ublox_i.value,_sxm_i.value, _dtv_i.value, _lont_i.value, _elecPN.value])    
        #print (" list diplay :", _PN_list )
        _elec_pn_row+=5   
    wb.close()
    
def create_prod_folders(_new_xml):
    global _container_names_list
    _container_list = []
    print ("creating production folder structure .. ")
    # tree = ET.ElementTree()
    # tree.parse(_new_xml)
    # root = tree.getroot()
    
    tree = ET1.ElementTree()
    tree.parse(_new_xml)
    root = tree.getroot()
    # parent=ET1.Element(root.tag)
    # child = ET1.SubElement(parent, 'SET_Infos/product')
    
    for _set_id,_set_info in _set_dict.items():
        for _key in _set_info:
            if _key == "product":
                _cont_name = _set_info[_key]
                _cont_name = "SW_Container_"+_cont_name
                _container_list.append(_cont_name)
    #print ("Container list : ", _container_list ) 
    _container_names_list =  _container_list.copy()  

    #print ("Container names : ", _container_names_list )     
    
    #target = r'\\bosch.com\dfsrb\DfsDE\DIV\CM\AI\SW_Production\Nissan\0060_CCS2_7515752366\00_SW\Test_folder'
    #_target = r'\\bosch.com\dfsrb\DfsDE\DIV\CM\AI\SW_Production\Nissan\0047_RN_AIVI_7513750800\Test_folder'  #hardcoded for now
    
    if not os.path.exists(_target):
        #print("\n creating production folder : ",_target )
        os.mkdir(_target)
                    
    
    if os.path.exists(_target):
        try:
        
            _dir_list = ["_Archive","BoardConfiguration","CD_Configuration","CTS_Configuration","PD_Configuration","QFIL_Download","Support_Files","UCB_Configuration","_Documentation"]
            _dir_list.extend(_container_list)
            _cont_sub_folders = ["_Archive","Data_to_plant","Product_SW"]
            _docu_sub_folders = ["_PCM_internal"]
            _PCMint_sub_folders = ["App_SW","Autosar-SW_Aurix","CTS","CTS_Config","Ublox","Reference_sheets"]
            _PCMint_sub_folders = ["Reference_sheets"]
            for _dirname in _dir_list:
                #print(_dirname)
                _new_dir = os.path.join(_target,_dirname)
                #os.mkdir(target)
                if not os.path.exists(_new_dir):
                    #print("\n creating folder : ",_new_dir )
                    os.mkdir(_new_dir)
                    if "SW_Container_" in _dirname:
                        for _subfolder in _cont_sub_folders:
                            _new_subfolder = os.path.join(_new_dir,_subfolder)
                            #print("\n creating sub folder : ",_new_subfolder )
                            os.mkdir(_new_subfolder)
                            
                            
                    if "_Documentation" in _dirname:
                        for _subfolder in _docu_sub_folders:
                            _new_docsubfolder = os.path.join(_new_dir,_subfolder)
                            #print("\n creating sub folder : ",_new_docsubfolder )
                            os.mkdir(_new_docsubfolder)
                            
                        for _subfolder in _PCMint_sub_folders:
                            _new_pcmsubfolder = os.path.join(_new_docsubfolder,_subfolder)
                            #print("\n creating sub folder : ",_new_pcmsubfolder )
                            os.mkdir(_new_pcmsubfolder)
                    
                else:
                    print ("dir :",_new_dir,"exists ")
            
            #get the container names from the _set_dict dictionar

            
                #exist_status = os.path.exists(target) 
            #    print(" Continuing after try .. ")
            
        except FileExistsError as e:    #this exception check is not needed as dir gets created only if path does not exist
            print('File already exists')
            #return False       #commented as to continue processing
        except OSError as e:
            print(f"An error has occurred: {e}")
            raise
        #print(" Continuing after except .. ")
    
    #get the container names from the xml and create the containers and sub-folders
    
    
        # _arc_path = os.path.join(target,"_Archive")
        # _arc_exist_status = os.path.exists(_arc_path)
        # if not os.path.exists(_arc_path):
            # print ("yes")
            # os.mkdir(target)
        
        
def check_artifacts():
    
    print("at check_artifacts :")
    global _cts_ver_file
    global _cts_cpver_file
    global _app_gver_file
    global _app_ngver_file
    global _aurix_srec_file
    global _aurix_dnl_file
    global _aurix_support_file
    global _PD_filename
    global _CFS_filename
    global _filenames_dict
    global _PD_file
    global _pd_ver_file
    global _prod_path
    global _SXM_fwver_file
    global _ublox_ver_file
    global _SXM_file
    global _DTV_SW_file
    global _DTV_P1_file
    global _DTV_P2_file
    global _DTV_swver_file
    global _DTV_P1ver_file
    global _DTV_P2ver_file
    global _lont_zip_file
    global _lont_ver_file
    
    
    _filenames_dict = {}    
    _prod_path = _target_prod
    #print ("\n _prod_path :", _prod_path)
    
    if "not" in _cts_file:
        print ("\n CTS file not available" )
    else:
        _bin_str = "CTS"
       
        _cts_path = os.path.join(_prod_path, _bin_str, _cts_ver)
        _cts_md5file = _cts_file + ".md5"     
        #print ("\n _cts_path :", _cts_path)
        if not os.path.exists(_cts_path):
            os.mkdir(os.path.join(_prod_path, _bin_str, _cts_ver))
        else:
            #print("cts version path exists")
            pass
        
        _cts_ver_file = _cts_path + "\\" + _cts_file
        _cts_ver_md5 = _cts_path + _cts_md5file
        
        _url = _cts_src + _cts_file
        start_download_from_artifactory(_url,_cts_ver_file ) 
        get_md5sum("cts_file_md5", _cts_ver_file)
       
        _filenames_dict["_cts_file"] = _cts_file 
       
    # CTS config starts 
    if "not" in _cts_cpfile:
        print ("\n CTS Config file not available" )
    else:   
        _bin_str = "CTS_Config"
        _cts_cppath = os.sep.join([_prod_path, _bin_str, _cts_cpver])
        #_cts_cppath = _prod_path + "\CTS_Config\\" + _cts_cpver
        _cts_cpmd5file = _cts_cpfile + ".md5" 
         
        if not os.path.exists(_cts_cppath):
            os.mkdir(_cts_cppath)
        else:
            pass
            #print("cts config version path exists")
            
        _cts_cpver_file = _cts_cppath + "\\" + _cts_cpfile
        _cts_cpver_md5 = _cts_cppath + _cts_cpmd5file
        # print("_cts_cpver_file ",_cts_cpver_file)
        # print("_cts_cpver_md5 ",_cts_cpver_md5)    
        
        _url = _cts_conf_src + _cts_cpfile
        #print("url : ", _url )
        start_download_from_artifactory(_url,_cts_cpver_file )      
        get_md5sum("cts_config_file_md5", _cts_cpver_file)      
        _filenames_dict["_cts_cpfile"] = _cts_cpfile 
    
    
    _rel_type = _tmpl_dir.split("\\")[-1]
    _sw_rel_type = _sw_ver.split("_")[-1]
    # print("_rel_type ",_rel_type )
    # print("_sw_rel_type ",_sw_rel_type )
       
    #_app_info starts
    if "Endunit" in _tmpl_dir and "EU" in _sw_rel_type:
        _bin_str = "App_SW"
        if "not" in _app_gfile:
            print ("\n App GAS file not available" )
        else:           
            _app_gmd5file = _app_gfile + ".md5" 
            # _app_ngmd5file= _app_ngfile + ".md5"
            
            _app_gpath = os.sep.join([_prod_path, _bin_str, _app_ver, _app_gver])
            
            
            if not os.path.exists(_app_gpath):
                os.makedirs(_app_gpath)
            else:
                #print("gas app version path exists")
                pass
           
            
            if "gas" in _device_type_dict: #download required only if gas PN is present
                _app_gver_file = _app_gpath + "\\" + _app_gfile
                _app_gver_md5 = _app_gpath + _app_gmd5file
               
                _url = _app_gsrc + _app_gfile
                    #print("url : ", _url )
                start_appsw_download_from_artifactory(_url,_app_gver_file )   
                get_md5sum("appsw_gas_file_md5",_app_gver_file)
                _filenames_dict["_app_gfile"] = _app_gfile 
            else:
                print ("No gas PNs available, gas binary file will not be downloaded")
            
        if "not" in _app_ngfile:
            print ("\n App NON GAS file not available")
        else:
            _app_ngmd5file= _app_ngfile + ".md5"
            _app_ngpath = os.sep.join([_prod_path, _bin_str, _app_ver, _app_ngver])  
            if not os.path.exists(_app_ngpath):
                os.makedirs(_app_ngpath)
            else:
                #print("non gas app version path exists")
                pass
                
            if "non-gas" in _device_type_dict:  #download required only if non-gas PN is present
                _app_ngver_file = _app_ngpath + "\\" + _app_ngfile
                _app_ngver_md5 = _app_ngpath + _app_ngmd5file
                
                _url = _app_ngsrc + _app_ngfile
                start_appsw_download_from_artifactory(_url,_app_ngver_file )  
                get_md5sum("appsw_nongas_file_md5", _app_ngver_file)
                _filenames_dict["_app_ngfile"] = _app_ngfile
            else:
                print ("No non-gas PNs available, non gas binary file will not be downloaded")
        
    else:
        print ("Not an end unit release, app sw binaries will not be downloaded ")
        appsw_gas_file_md5 = "not_applicable"
        appsw_nongas_file_md5 = "not_applicable"
    
    _sw_sample = _sw_ver.split("_")[0]
    #print("\n _sw_sample : ", _sw_sample )
    
    #_aurix_info start
    _aurix_srmd5file = _aurix_srec + ".md5"
    _aurix_dmd5file = _aurix_dnl + ".md5"
    _aurix_spmd5file = _aurix_support + ".md5"
    
    _bin_str = "Autosar-SW_Aurix"
    _aurix_path = os.sep.join([_prod_path, _bin_str, _aurix_ver])
    
    #if "Endunit" in _tmpl_dir and "EU" in _sw_rel_type: 
    if _aurixcryptosign == "kms":
        if "SOP" in _sw_sample or "MP" in _sw_sample or "PP" in _sw_sample:
            _aurix_path = os.sep.join([_aurix_path, "kms-signed"])
            #print("\n _aurix_path : ", _aurix_path )
            
    
    if not os.path.exists(_aurix_path):
        os.makedirs(_aurix_path)
    else:
        #print("aurix verion path exists")
        pass
        
    _aurix_srec_file = _aurix_path + "\\" + _aurix_srec
    #print("\n _aurix_srec_file for EU : ", _aurix_srec_file )
    _aurix_srec_md5 = _aurix_path + _aurix_srmd5file
    _aurix_dnl_file = _aurix_path + "\\" + _aurix_dnl
    _aurix_dnl_md5 = _aurix_path + _aurix_dmd5file
    _aurix_support_file = _aurix_path + "\\" + _aurix_support
    _aurix_support_md5 = _aurix_path + _aurix_spmd5file
    
     
    if "not" in _aurix_srec:
        print ("\n Aurix SREC file not available" )
    else:      
    #if not os.path.isfile(_aurix_srec_file):
        _url = _aurix_src + _aurix_srec
        
        start_download_from_artifactory(_url,_aurix_srec_file ) 
        get_md5sum("aurix_srec_file_md5", _aurix_srec_file) 
        _filenames_dict["_aurix_srec"] = _aurix_srec
    
    if "not" in _aurix_dnl:
        print ("\n Aurix DNL file not available" )
    else:
        _url = _aurix_src + _aurix_dnl
        start_download_from_artifactory(_url,_aurix_dnl_file ) 
        get_md5sum("aurix_dnl_file_md5", _aurix_dnl_file) 
        _filenames_dict["_aurix_dnl"] = _aurix_dnl
        
    if "not" in _aurix_support:
        print ("\n Aurix Disable Support file not available" )
    else:
        _url = _aurix_src + _aurix_support
        start_download_from_artifactory(_url,_aurix_support_file )
        get_md5sum("aurix_suport_file_md5",_aurix_support_file) 
        _filenames_dict["_aurix_support"] = _aurix_support


    #PD Config
    _PD_filename = _PDDel_info["file"]
    _PD_Version = _PDDel_info["version"]
    _url = _PDDel_info["src_file"] + "/"+_PD_filename
    #_pd_md5file = _PD_filename  + ".md5"  
    if "not" in _PD_filename:
        print ("\n PD Config file not available" )
    else:
        _bin_str = "PD_Config"
        _PD_path = os.sep.join([_prod_path, _bin_str, _PD_Version])
        #print("_PD_path : ", _PD_path )
        if not os.path.exists(_PD_path):
            os.makedirs(_PD_path)
        else:
            #print("PD path exists")
            pass 
            
        _pd_ver_file = _PD_path + "\\" + _PD_filename
        if (os.path.exists(_pd_ver_file)):
            print("PD file already exists: ", _pd_ver_file)
        else:     
            print("PD : The file doesn't exists, Hence it will get downloaded. \n")
            download_from_sharepoint(_url, _pd_ver_file)   
        #print("\n_pd_ver_file:", _pd_ver_file)
        #_pd_ver_md5 = _PD_path + _pd_md5file  
        #_url = _PD_path
        #download_from_sharepoint(_url, _pd_ver_file)
        get_md5sum("PD_config_file_md5", _pd_ver_file) 
        _filenames_dict["_PD_filename"] = _PD_filename
    
    #CFS file
    _CFS_filename =  _PTooling_info["file"]
    _CFS_Version = _PTooling_info["version"]
    if "not" in _CFS_filename:
        print ("\n Production Tooling CFS file not available" )
    else:
        _cfs_prod_path = _prod_path.split("00_SW")[0]
        _cfs_prod_path = _cfs_prod_path + "01_Tools\\production_tooling\\" + _CFS_Version.upper()
        #print ("\n _cfs_prod_path : ", _cfs_prod_path ) 
        _CFS_ver_file = _cfs_prod_path + "\\" + _CFS_filename
        if not os.path.exists(_CFS_ver_file):
            print (_CFS_ver_file, " does not exist, please check ")
            exit()
        else:
            print(f" \n {_CFS_filename} exists in the path {_cfs_prod_path} \n")
             
       
        get_md5sum("CFS_file_md5", _CFS_ver_file)
        _filenames_dict["_CFS_filename"] = _CFS_filename
    
    
    #uBlox file    
    if "Endunit" in _tmpl_dir and "EU" in _sw_rel_type:
        _ublox_file = _ublox_info["eu_file"]
    else:
        _ublox_file = _ublox_info["bu_file"]
   
    if "not" in _ublox_file:
        print ("\n uBlox file", _ublox_file, " not available" )
    else:
        _bin_str = "Ublox"
           
        _ublox_path = os.path.join(_prod_path, _bin_str,  _ublox_info["version"])
        #print ("\n _ublox_path :", _ublox_path)
        _ublox_md5file = _ublox_file + ".md5"     
          
        if not os.path.exists(_ublox_path):
            print ("\n _ublox_path here :", _ublox_path)
            os.mkdir(os.path.join(_prod_path, _bin_str, _ublox_info["version"]))
        else:
             #print("ublox version path exists")
             pass 
            
        _ublox_ver_file = _ublox_path + "\\" + _ublox_file
        _ublox_ver_md5 = _ublox_path + _ublox_md5file
        #ublox download in future we have to do via git-lfs , for 4.04 version, we proceed with the binaries we have
        #_url = _cts_src + _cts_file
        #start_download_from_artifactory(_url,_cts_ver_file ) 
        get_md5sum("ublox_file_md5", _ublox_ver_file)
       
        _filenames_dict["_ublox_filename"] = _ublox_file 
    
    _sxm_type_status = "no"
    for _i in _set_dict:	
        if _sxm_type_status != "yes":
            if _set_dict[_i]["sxm"].lower() == "yes":
                _sxm_type_status = "yes"
              
    
     #SXM file
    _SXM_file =  _sxm_info["file"]
    _SXM_fwVersion = _sxm_info["fw_version"]
    
    if "not" in _SXM_file:
        print ("\n SXM file not available" )
    else:
        _bin_str = "SXM"
        _SXM_path = os.path.join(_prod_path, _bin_str, _SXM_fwVersion) 
        #_SXM_ver_file = _SXM_path + "\\" + _SXM_file
        _SXM_md5file = _SXM_file + ".md5"
        if not os.path.exists(_SXM_path):
            print (_SXM_path, " does not exist, please check ")
            os.mkdir(os.path.join(_prod_path, _bin_str, _SXM_fwVersion))
        else:
            #print("SXM path exists")
            pass
        _SXM_fwver_file = _SXM_path + "\\" + _SXM_file
        _SXM_fwver_md5 = _SXM_path + _SXM_md5file
        _url = _sxm_info["src_file"] + _SXM_file
        # print ("\n display :", _sxm_info["src_file"] )
        # print ("\n _url :", _url )
        if _sxm_type_status == "yes":
            start_download_from_artifactory(_url,_SXM_fwver_file )
            get_md5sum("SXM_file_md5", _SXM_fwver_file)
            _filenames_dict["_sxm_filename"] = _SXM_file
    
    #DTV changes
    _dtv_type_status = "no"
    for _i in _set_dict:	
        if _dtv_type_status != "yes":
            if _set_dict[_i]["dtv"].lower() == "yes":
                _dtv_type_status = "yes"


    _DTV_SW_file =  _DTV_SW_info["file"]
    _DTV_SW_ver = _DTV_SW_info["version"]    
    if "not" in _DTV_SW_file:
        print ("\n _DTV_SW_file not available" )
    else:
        _bin_str = "DTV"
        _bin_sub_str = "DTV_SW"
        _DTV_swpath = os.path.join(_prod_path, _bin_str, _bin_sub_str, _DTV_SW_ver) 
        _DTV_swver_file = _DTV_swpath + "\\" + _DTV_SW_file
        _DTV_swmd5file = _DTV_SW_file + ".md5"
        if not os.path.exists(_DTV_swpath):
            print (_DTV_swpath, " does not exist, please check ")
            os.mkdir(os.path.join(_prod_path, _bin_str, _bin_sub_str, _DTV_SW_ver))
        else:
            #print("DTV SW path exists")
            pass
        #_DTV_swver_file = _DTV_swpath + "\\" + _DTV_SW_file
        _DTV_swver_md5 = _DTV_swpath + _DTV_swmd5file
        _url = _DTV_SW_info["src_file"] + _DTV_SW_file
        if _dtv_type_status == "yes":
            start_download_from_artifactory(_url,_DTV_swver_file )
            get_md5sum("DTV_swfile_md5", _DTV_swver_file)
            _filenames_dict["_dtv_swfilename"] = _DTV_SW_file
        
    _DTV_P1_file = _DTV_Par1_info["file"]
    _DTV_P1_ver = _DTV_Par1_info["version"]
    if "not" in _DTV_P1_file:
        print ("\n _DTV_P1_file not available" )
    else:
        _bin_str = "DTV"
        _bin_sub_str = "DTV_Parameter_1"
        _DTV_P1path = os.path.join(_prod_path, _bin_str, _bin_sub_str, _DTV_P1_ver) 
        _DTV_P1ver_file = _DTV_P1path + "\\" + _DTV_P1_file
        _DTV_P1md5file = _DTV_P1_file + ".md5"
        if not os.path.exists(_DTV_P1path):
            print (_DTV_P1path, " does not exist, please check ")
            os.mkdir(os.path.join(_prod_path, _bin_str, _bin_sub_str, _DTV_P1_ver))
        else:
            #print("DTV Par1 path exists")
            pass
        #_DTV_P1ver_file = _DTV_P1path + "\\" + _DTV_P1_file
        _DTV_P1ver_md5 = _DTV_P1path + _DTV_P1md5file
        #_url = _DTV_Par1_info["src_file"] + _DTV_P1_file
        
        if _dtv_type_status == "yes":
            #start_download_from_artifactory(_url,_DTV_P1ver_file ) # commented as DTV Parameter files are not there in artifactory now
            get_md5sum("DTV_P1file_md5", _DTV_P1ver_file)
            _filenames_dict["_dtv_p1filename"] = _DTV_P1_file
        
    _DTV_P2_file = _DTV_Par2_info["file"]
    _DTV_P2_ver = _DTV_Par2_info["version"]
    if "not" in _DTV_P2_file:
        print ("\n _DTV_21_file not available" )
    else:
        _bin_str = "DTV"
        _bin_sub_str = "DTV_Parameter_2"
        _DTV_P2path = os.path.join(_prod_path, _bin_str, _bin_sub_str, _DTV_P2_ver) 
        _DTV_P2ver_file = _DTV_P2path + "\\" + _DTV_P2_file
        _DTV_P2md5file = _DTV_P2_file + ".md5"
        if not os.path.exists(_DTV_P2path):
            print (_DTV_P2path, " does not exist, please check ")
            os.mkdir(os.path.join(_prod_path, _bin_str, _bin_sub_str, _DTV_P2_ver))
        else:
            #print("DTV Par2 path exists")
            pass
        #_DTV_P2ver_file = _DTV_P2path + "\\" + _DTV_P2_file
        _DTV_P2ver_md5 = _DTV_P2path + _DTV_P2md5file
        #_url = _DTV_Par2_info["src_file"] + _DTV_P2_file
        
        if _dtv_type_status == "yes":
            #start_download_from_artifactory(_url,_DTV_P2ver_file )
            get_md5sum("DTV_P2file_md5", _DTV_P2ver_file)
            _filenames_dict["_dtv_p2filename"] = _DTV_P2_file
            
    _lont_zip_file = _Lont_info["zip_file"]
    _lont_ver = _Lont_info["version"]
    if "not" in _lont_zip_file:
        print ("\n _lont_zip_file not available" )
    else:
        _bin_str = "Lontium"
        _lont_path = os.path.join(_prod_path, _bin_str, _lont_ver) 
        _lont_ver_file = _lont_path + "\\" + _lont_zip_file
        #_lont_md5file = _lont_ver_file + ".md5"
        if not os.path.exists(_lont_path):
            print (_lont_path, " does not exist, please check ")
            os.mkdir(os.path.join(_prod_path, _bin_str, _lont_ver))
        else:
            pass
        #lontium artifacts are not yet shared via artifactory/sharepoint, we have used the shared version, so we will not have .md5 file, also download from artifactory is not needed.
        if _dtv_type_status == "yes":
            #start_download_from_artifactory(_url,_DTV_P2ver_file ) #Lontium file not yet in artifactory
            get_md5sum("lont_file_md5", _lont_ver_file)            #md5sum not needed
            _filenames_dict["_lont_filename"] = _lont_zip_file
    
    
    
    update_sap_xml("md5_info", _new_xml)   #updates the  md5 to the sap xml
    #print ("\n _filenames_dict:", _filenames_dict)


def start_download_from_artifactory(_url,_filename):
    #Ralph suggested that the artifact need not be downloaded again , if it is already there, but md5 file to be downloaded and corsschecked against the md5 available
    
    #downloads the binary file if it is not available
    #print ("url: \t", _url )
    #timeout = 600
    
    if not os.path.isfile(_filename):
        print("\nstarting the download for file:",_filename ) 
        path = ArtifactoryPath(_url) 
             
        with path.open() as fd:
            with open(_filename, "wb") as out:
                out.write(fd.read())          
    else:
        pass
        #print( "\n",_filename, "exists, will not be downloaded again " ) 
    #downloads the md5 file in any case
    #_file = _filename
    if "dtv" not in _url:
        _url= _url + ".md5"
        path = ArtifactoryPath(_url)        
        #path = ArtifactoryPath(_url, verify=False)        
        _filename = _filename + ".md5"
        print("\nstarting the download for file:",_filename )
        with path.open() as fd:
            with open(_filename, "wb") as out:
                out.write(fd.read())
    else:
        print ("\n dtv file, .md5 file will not be downloaded ") #workaround for now to avoid md5 file check

def start_appsw_download_from_artifactory(_url,_filename):
    #Ralph suggested that the artifact need not be downloaded again , if it is already there, but md5 file to be downloaded and corsschecked against the md5 available
    #print ("url: \t", _url )
    
    if not os.path.isfile(_filename):
        print("\nstarting the download for file:",_filename ) 
        path = ArtifactoryPath(_url)       
        #path = ArtifactoryPath(_url, verify=False)       
        with open(_filename, 'wb') as f:
            with path.open() as repo_file:
                chunk_size=1024*1024
                while True:
                    chunk = repo_file.read(chunk_size)
                    if not chunk:
                        break
                    f.write(chunk)
        
        #print(_url, " successfully downloaded as ", _filename)   
    else:
        pass
        #print( "\n",_filename, "exists, will not be downloaded again " ) 
    #downloads the md5 file in any case
    #_file = _filename
    _url= _url + ".md5"
    path = ArtifactoryPath(_url)        
    _filename = _filename + ".md5"
    print("\nstarting the download for file:",_filename )
    with path.open() as fd:
        with open(_filename, "wb") as out:
            out.write(fd.read())

#def download_from_sharepoint(_url, _pd_ver_file):
def download_from_sharepoint(_url, _filename):
    _kerberos_auth = HTTPKerberosAuth(mutual_authentication=OPTIONAL)
    _file_url = _url
    #print(f"The file url:{file_url}")
    _response = requests.get(_file_url, auth=_kerberos_auth, verify=False) 
    #destination_path = _filename
    _statuscode = _response.status_code 
    #downloaded_file_path = destination_path
    
    if _statuscode != 200:
        print (f"Could not download the file, please check:Response code is {_statuscode}")
    else:
        #print(f"Response Code : {_statuscode}, OK" )
        _output = open(_filename, 'wb')
        _output.write(_response.content)
        #print (f"successfully downloaded {_filename}")
        _output.close()

            
def verify_md5sum(_md5_binary_key, _file ):
    #to cross verify the md5sum of the copied file against the md5sum of the source file which was already calculated via get_md5sum at check_artifacts function
    #print("\nat verify_md5sum for file:", _file)
    _source = _md5_binary_key
    _md5_binary_key_verify = _md5_binary_key + "_verify"
    _file = _file.replace('\\', '/') 
    _md5_temp_file = "verify_content.md5"    
    # _base_path = os.path.basename(_file)
    # _dir_path = os.path.dirname(_file)
    _cmd = 'md5sum '+ _file
    _result_code = os.system(_cmd + ' > verify_content.md5') #local file created
    if os.path.exists(_md5_temp_file):
        with open(_md5_temp_file, 'r') as fp:
            _md5_line = fp.read() #reading the local .md5 file
            _md5_line = (_md5_line.split(" ")[0]).strip()
            _md5_dest_dict[_md5_binary_key_verify] = _md5_line
            #print("_md5_line:",_md5_line)
            fp.close()
        
        
        _md5_source = _md5_dict[_source]
       
        if _md5_line != _md5_source:
            print("\n md5sum of the source file:", _md5_source)
            print("\n _md5sum of file:", _file, ":", _md5_line)
            print("ERROR: MD5 does not match between the actual file content and the .md5 file")  
            sys.exit(0)
            
        os.remove(_md5_temp_file)
            
  
def get_md5sum(_md5_binary_name, _file ): 
    #the md5sum of each binary is calculated everytime
    #even if the file already exists
    #or if the file is freshly downloaded
    #this is to ensure if there is any change in the md5sum of the file 
    #print ("at get_md5sum .. ")
    _file = _file.replace('\\', '/')
    #print("file:",_file)
    
   
    _base_path = os.path.basename(_file)
    _dir_path = os.path.dirname(_file)
    
    # print("_base_path:",_base_path)
    # print("_dir_path:",_dir_path)
     
    _md5_temp_file = "content.md5"
    
    _cmd = 'md5sum '+ _file
    #print ("cmd :", _cmd)
    
    _result_code = os.system(_cmd + ' > content.md5') #local file created
   
    #fj = open(_md5_file, "w")
    if os.path.exists(_md5_temp_file):
        with open(_md5_temp_file, 'r') as fp:
            _md5_line = fp.read() #reading the local content.md5 file
            _md5_line = (_md5_line.split(" ")[0]).strip()
            #print("_md5_line:",_md5_line)
            fp.close()
            
            #PD_config file_md5 and CFG file will not have the binary.md5 file
            #print("_md5_binary_name:",_md5_binary_name) 
            if _md5_binary_name == "PD_config_file_md5" or _md5_binary_name == "CFS_file_md5" or _md5_binary_name == "ublox_file_md5" or _md5_binary_name == "DTV_swfile_md5" or _md5_binary_name == "DTV_P1file_md5"  or _md5_binary_name == "DTV_P2file_md5" :
                _md5_dict[_md5_binary_name] = _md5_line
            
            #check as to ensure that the Lontium file placed in the server is not corrupted  , this version and md5 extraction from the filename may change in future if the naming convention is changed
            elif _md5_binary_name == "lont_file_md5":
                _lont_file_ver  = (os.path.splitext(_base_path)[0]).split("_")[-1]  #get the version from the filename
                _lont_md5 = (os.path.splitext(_base_path)[0]).split("_")[1]         #get the md5 from the filename
                _md5_10 = _md5_line[:10]
                # print ("_md5_line:", _md5_line)
                # print ("_md5_10:", _md5_10)
                # print ("_lont_md5:", _lont_md5)
                if _lont_md5 == _md5_10:
                    _md5_dict[_md5_binary_name] = _md5_line
                else:
                    sys.exit("MD5 of Lontium file does not match between the actual file content and the .md5 file")        
                
            else:
                _file_md5 = _base_path + ".md5"
                _file_md5 = os.path.join(_dir_path,_file_md5)
                #print ("_file_md5 : ",_file_md5)
                fm = open(_file_md5, "r")
                _file_md5_value= (fm.read()).strip()
                #print ("_file_md5_value : ",_file_md5_value)
                _md5_file = os.path.join(_dir_path,_md5_temp_file)
                fj = open(_md5_file, "w")
                if (_file_md5_value == _md5_line ):
                    #print ("\n md5 calculated matches with the value in .md5 file, updating the content.md5 file ")
                    #entry in the file is not required, as it is deleted through os.remove - added if in case it is required in future.
                    _md5_dict[_md5_binary_name] = _md5_line
                    _md5_file_entry = _base_path + ": " + _md5_line + "\n" 
                    fj.write(_md5_file_entry)
                    fj.close()
                else:
                    print ("filename: ", _base_path )
                    print ("file md5 :", _md5_line )
                    print("md5 file content :", _file_md5_value)
                    sys.exit("MD5 does not match between the actual file content and the .md5 file")
                os.remove(_md5_file)   # this file can be retained in future if needed , in that case make the changes to append the entries of aurix files, as it is called thrice from the same path                
    os.remove(_md5_temp_file)
    
    #print("md5 dictionary ", _md5_dict)                
    #exit()
    
                   

def set_filenames(_new_xml):
    # the filenames are formed and added to the dictionary filenames_dict
    # since a release is either BU or EU at any point of time, 
    # only app sw name differs between PNs for BU - btn gas and non gas, all the other filenames will be the same
    # aurix dnl filename is common , aurix srec with PN is for BU and with xxxxxxxxxx PN is for EU, aurix disableHsmDNL filename is common for BU and EU
    # _filenames_dict is formed from check_artifacts fucntion based on the availability of artifacts, the dictionary will contain only the files that are available
    # the filenames for artifacts are set based on the check if the file is present in the _filenames_dict, if it is not there , filenames will not be set
    # _setfilenames_dict will be set only for the filenames formed.
    # _ucb_config_filename will be added to the _setfilenames_dict
    
    print ("at set_filenames .. ")
    global _cts_filename
    global _app_gfilename
    global _app_ngfilename
    global _aurix_dnl_filename
    global _aurix_srec_filename
    global _aurix_srec_support_filename
    global _aurix_disHsmdnl_support_filename
    global _cts_config_filename
    global _setfilenames_dict
    global _ucb_config_filename
    global _ublox_filename
    global _sxm_filename
    global _dtv_swfilename
    global _lont_filename
    
    _setfilenames_dict = {}    
    _support_image_number = "xxxxxxxxxx"
    _ucb_config_filename = "Nissan_CCS2_UCB_Configuration_V0.9.1.xlsx"      #hardcoded as we have only this version file available" 
    
    
    #print ("\nSetting the filenames .." )
    #print ("\nmaster xml :", _new_xml )
    #print (" _filenames_dict: ", _filenames_dict )
    
    tree = ET1.ElementTree()
    tree.parse(_new_xml)
    root = tree.getroot()
    #_versions_tuple = (_cts_ver, _cts_cpver, _app_gver, _app_ngver, _aurix_ver)
    _cts_version = _cts_ver.lower()  
    _cts_version = (_cts_version.split("cts_")[1]).upper()
    
    for SET in root.iter("SET_Infos"):
        #print (SET.attrib )
        #print (SET.attrib["part_number"] )
        _PN = SET.attrib["part_number"]
        #print(type(_PN))
        _PN = _PN.split(",")[0]
        _PN.strip()
        #print ("filenames dict :", _filenames_dict)
        if "_cts_file" in _filenames_dict:
            if "_cts_filename" in _setfilenames_dict:  #considering that cts remains the same for all, set the name once is enough
                #print ("_cts_filename in _setfilenames_dict" )
                pass
            else:
                _cts_filename = SET.attrib["cts_image_number"] + "_" + SET.attrib["cts_file_md5"][:10] + "_"  + _cts_version + "_"  + _cts_file
                #print ("\n  cts filename ", _cts_filename )
                _setfilenames_dict["_cts_filename"] = _cts_filename
        else:
            print ( " _cts_file is not available, skipping the set filename .. " )
        
        if "_cts_cpfile" in _filenames_dict:
            _cts_cpfilename = _cts_cpfile.split(".zip")[0]
            _cts_cpfileextn = _cts_cpfile.split(".")[-1]
            # print ("\n  _cts_cpfilename ", _cts_cpfilename )
            # print ("\n  _cts_cpfileextn ", _cts_cpfileextn )
            if "_cts_config_filename" in _setfilenames_dict:  #considering that cts remains the same for all, set the name once is enough
                #print ("_cts_config_filename in _setfilenames_dict" )
                pass
            else:
                _cts_config_filename = _cts_cpfilename + "_"  + _cts_cpver + "." + _cts_cpfileextn
                _setfilenames_dict["_cts_config_filename"] = _cts_config_filename   #CTS config
                #print ("\n  cts config filename ", _cts_config_filename )
        else:
            print ( " _cts_cpfile is not available, skipping the set filename .. " )
        
        
        if "_PD_filename" in _filenames_dict:
            if "_PD_filename" in _setfilenames_dict:  #considering that cts remains the same for all, set the name once is enough
                #print ("_PD_filename in _setfilenames_dict" )
                pass
            else:
                _setfilenames_dict["_PD_filename"] = _PD_filename       #PD Config
                #print ("\n PD filename ", _PD_filename )
        else:
            print ( " _PD_filename is not available, skipping the set filename .. " )
        
        if "_CFS_filename" in _filenames_dict:
            if "_CFS_filename" in _setfilenames_dict:  #considering that cts remains the same for all, set the name once is enough
                #print ("_CFS_filename in _setfilenames_dict ")
                pass
            else:
                _setfilenames_dict["_CFS_filename"] = _CFS_filename     # CFS
                #print ("\n CFS filename ", _CFS_filename )
        else:
            print ( " _CFS_filename is not available, skipping the set filename .. " )
        
        
        # For end unit
        if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:
            
            if SET.attrib["device_type"] == "yes":
                if "_app_gfile" in _filenames_dict:
                    if "_app_gfilename" in _setfilenames_dict:  #considering that _app_gfile file remains the same for all, set the name once is enough
                        #print ("_app_gfilename in _setfilenames_dict ")
                        pass
                    else:
                        _app_gfilename = SET.attrib["app_image_number"] + "_" + SET.attrib["appsw_gas_file_md5"][:10] + "_"  + _app_gfile
                        #print ("\n  app gas filename ", _app_gfilename ) 
                        _setfilenames_dict["_app_gfilename"] = _app_gfilename 
                else:
                    print ( " _app_gfile is not available, skipping the set filename .. " )    
                
            elif SET.attrib["device_type"] == "non":
                if "_app_ngfile" in _filenames_dict:
                    if "_app_ngfilename" in _setfilenames_dict:  #considering that _app_ngfile file remains the same for all, set the name once is enough
                        #print ("_app_ngfilename in _setfilenames_dict" )
                        pass
                    else:
                        _app_ngfilename = SET.attrib["app_image_number"] + "_" + SET.attrib["appsw_nongas_file_md5"][:10] + "_"  + _app_ngfile
                        #print ("\n  app non gas gfilename ", _app_ngfilename ) 
                        _setfilenames_dict["_app_ngfilename"] = _app_ngfilename
                else:
                    print ( " _app_ngfile is not available, skipping the set filename .. " )

            if "_aurix_dnl" in _filenames_dict: 
                if "_aurix_dnl_filename" in _setfilenames_dict:  #considering that _aurix_dnl file remains the same for all, set the name once is enough for EU
                    #print ("_aurix_dnl_filename in _setfilenames_dict" )
                    pass
                else:
                    _aurix_dnl_filename = SET.attrib["aurix_dnl_image_number"] + "_" + SET.attrib["aurix_dnl_file_md5"][:10] + "_"  + _aurix_ver + "_"  + _aurix_dnl
                    _setfilenames_dict["_aurix_dnl_filename"] = _aurix_dnl_filename
                    #print ("\n  aurix dnl filename ", _aurix_dnl_filename )
            else:
                    print ( " _aurix_dnl is not available, skipping the set filename .. " )  
            
            if "_aurix_srec" in _filenames_dict:  
                if "_aurix_srec_support_filename" in _setfilenames_dict:  #considering that _aurix_srec file remains the same for all, set the name once is enough for EU
                    #print ("_aurix_srec_support_filename in _setfilenames_dict" )
                    pass
                else:
                    _aurix_srec_support_filename = _support_image_number + "_" + SET.attrib["aurix_srec_file_md5"][:10] + "_"  + _aurix_ver + "_"  + _aurix_srec
                    _setfilenames_dict["_aurix_srec_support_filename"] = _aurix_srec_support_filename
                    #print ("\n  aurix srec support filename ", _aurix_srec_support_filename )
            else:
                print ( " _aurix_srec is not available, skipping the set filename .. " )  
            
            if "_aurix_support" in _filenames_dict:        
                if "_aurix_disHsmdnl_support_filename" in _setfilenames_dict:  #considering that _aurix_support file remains the same for all, set the name once is enough for EU
                    #print ("_aurix_disHsmdnl_support_filename in _setfilenames_dict ")
                    pass
                else:
                    _aurix_disHsmdnl_support_filename = _support_image_number + "_" + SET.attrib["aurix_suport_file_md5"][:10] + "_"  + _aurix_ver + "_"  + _aurix_support
                    _setfilenames_dict["_aurix_disHsmdnl_support_filename"] = _aurix_disHsmdnl_support_filename
                    #print ("\n  aurix disable Hsm dnl support filename ", _aurix_disHsmdnl_support_filename )
            else:
                print ( " _aurix_support is not available, skipping the set filename .. " )  
            
            #ublox
            if "_ublox_filename" in _filenames_dict:
                # print ("\n filenames dict :", _filenames_dict )
                # print ("\n _setfilenames_dict dict :", _setfilenames_dict )
                if "_ublox_filename" in _setfilenames_dict:  #considering that cts remains the same for all, set the name once is enough
                    #print ("_ublox_filename in _setfilenames_dict" )
                    pass
                else:
                    if _ublox_info["version"] == "4.04":
                        _ublox_eu_md5 = "907640e53126add7de4708d754b575ef"
                        #hardcoding the name for 4.04 version as we have filename existing
                        if SET.attrib["ublox_image_number"] == "8609625164" and SET.attrib["ublox_file_md5"] == _ublox_eu_md5:
                            _ublox_filename = SET.attrib["ublox_image_number"] + "_" + SET.attrib["ublox_file_md5"][:10] + "_JU_EXT_404.bin"
                        else:  
                            _ublox_filename = SET.attrib["ublox_image_number"] + "_" + SET.attrib["ublox_file_md5"][:10] + "_"  + _ublox_info["version"] + "_"  + _ublox_file
                        
                    #print ("\n ublox filename ", _ublox_filename )
                _setfilenames_dict["_ublox_filename"] = _ublox_filename
            else:
                print ( " _ublox_file is not available, skipping the set filename .. " )
                
                
        # For base unit
        elif "Baseunit" in _tmpl_dir and "BU" in _sw_ver.split("_")[-1]:
            if "_aurix_srec" in _filenames_dict:
                if "_aurix_srec_filename" in _setfilenames_dict:  #considering that _aurix_srec file remains the same for all, set the name once is enough for BU
                    #print ("_aurix_srec_filename in _setfilenames_dict" )
                    pass
                else:
                    _aurix_srec_filename = SET.attrib["aurix_srec_image_number"] + "_" + SET.attrib["aurix_srec_file_md5"][:10] + "_"  + _aurix_ver + "_"  + _aurix_srec
                    #print ("\n  aurix srec filename ", _aurix_srec_filename )
                    _setfilenames_dict["_aurix_srec_filename"] = _aurix_srec_filename
            else:
                print ( " _aurix_srec is not available, skipping the set filename .. " ) 
                
            if "_aurix_dnl" in _filenames_dict:
                if "_aurix_dnl_filename" in _setfilenames_dict:  #considering that _aurix_dnl file remains the same for all, set the name once is enough for BU
                    #print ("_aurix_dnl_filename in _setfilenames_dict" )
                    pass
                else:
                    _aurix_dnl_filename = SET.attrib["aurix_dnl_image_number"] + "_" + SET.attrib["aurix_dnl_file_md5"][:10] + "_"  + _aurix_ver + "_"  + _aurix_dnl
                    #print ("\n  aurix dnl filename ", _aurix_dnl_filename )
                    _setfilenames_dict["_aurix_dnl_filename"] = _aurix_dnl_filename
            else:
                print ( " _aurix_dnl is not available, skipping the set filename .. " ) 
                
            if "_aurix_support" in _filenames_dict:
                if "_aurix_disHsmdnl_support_filename" in _setfilenames_dict:  #considering that _aurix_support file remains the same for all, set the name once is enough for BU
                    #print ("_aurix_disHsmdnl_support_filename in _setfilenames_dict" )
                    pass
                else:
                    _aurix_disHsmdnl_support_filename = _support_image_number + "_" + SET.attrib["aurix_suport_file_md5"][:10] + "_"  + _aurix_ver + "_"  + _aurix_support
                    #print ("\n  aurix disable Hsm dnl support filename ", _aurix_disHsmdnl_support_filename )
                    _setfilenames_dict["_aurix_disHsmdnl_support_filename"] = _aurix_disHsmdnl_support_filename
            else:
                print ( " _aurix_support is not available, skipping the set filename .. " ) 
            
            #ublox 
            if "_ublox_filename" in _filenames_dict:
                if "_ublox_filename" in _setfilenames_dict:  
                    #print ("_ublox_filename in _setfilenames_dict" )
                    pass
                else:
                    if _ublox_info["version"] == "4.04":
                        _ublox_bu_md5 = "80ff71bfcf34c8929b04aa9d16d9ede0"
                        #hardcoding the name for 4.04 version as we have filename existing
                        if SET.attrib["ublox_image_number"] == "8609624083" and SET.attrib["ublox_file_md5"] == _ublox_bu_md5:
                            _ublox_filename = SET.attrib["ublox_image_number"] + "_" + SET.attrib["ublox_file_md5"][:10] + "__UBX_M9_404_SPG_0x48_shifted.bin"
                    else:  
                        _ublox_filename = SET.attrib["ublox_image_number"] + "_" + SET.attrib["ublox_file_md5"][:10] + "_"  + _ublox_info["version"] + "_"  + _ublox_file
                    
                    #print ("\n ublox filename ", _ublox_filename )
                    _setfilenames_dict["_ublox_filename"] = _ublox_filename
            else:
                print ( " _ublox_file is not available, skipping the set filename .. " )
        #include sxm filename if part number has sxm, same applies for DTV and Lontium        
        #if "_sxm_filename" in _filenames_dict:
        if "_sxm_filename" in _filenames_dict and SET.attrib["sxm"] == "yes":
            if "_sxm_filename" in _setfilenames_dict:  
               #print ("_sxm_filename in _setfilenames_dict" )
                pass
            else:
                #_setfilenames_dict["_sxm_filename"] = _SXM_file
                _sxm_filename = SET.attrib["sxm_image_number"] + "_" + _SXM_file
                _setfilenames_dict["_sxm_filename"] = _sxm_filename 
        else:
            pass
            #print ( " _sxm_file is not available, skipping the set filename .. " )
            
        #if "_dtv_swfilename" in _filenames_dict:
        if "_dtv_swfilename" in _filenames_dict and SET.attrib["dtv"] == "yes":
            if "_dtv_swfilename" in _setfilenames_dict:  
               #print ("_sxm_filename in _setfilenames_dict" )
                pass
            else:
                #_setfilenames_dict["_dtv_swfilename"] = _DTV_SW_file
                _dtv_swfilename = SET.attrib["dtv_image_number"] + "_" + _DTV_SW_file
                _setfilenames_dict["_dtv_swfilename"] = _dtv_swfilename
        else:
            pass
            #print ( " _DTV_SW_file is not available, skipping the set filename .. " )
            
        if "_dtv_p1filename" in _filenames_dict and SET.attrib["dtv"] == "yes":
            if "_dtv_p1filename" in _setfilenames_dict:  
               #print ("_sxm_filename in _setfilenames_dict" )
                pass
            else:
                _setfilenames_dict["_dtv_p1filename"] = _DTV_P1_file
        else:
            pass
            #print ( " _DTV_P1_file is not available, skipping the set filename .. " )
            
        if "_dtv_p2filename" in _filenames_dict and SET.attrib["dtv"] == "yes":
            if "_dtv_p2filename" in _setfilenames_dict:  
               #print ("_sxm_filename in _setfilenames_dict" )
                pass
            else:
                _setfilenames_dict["_dtv_p2filename"] = _DTV_P2_file
        else:
            pass
            #print ( " _DTV_P2_file is not available, skipping the set filename .. " )
        
        _setfilenames_dict["_ucb_config_filename" ] = _ucb_config_filename        
            
        if "_lont_filename" in _filenames_dict and SET.attrib["dtv"] == "yes":
            if "_lont_filename" in _setfilenames_dict:  
                pass
            else:
                #_setfilenames_dict["_lont_filename"] = _lont_zip_file
                _lont_filename = SET.attrib["lont_image_number"] + "_" + (_lont_zip_file.split("_",1))[1]
                _setfilenames_dict["_lont_filename"] = _lont_filename
        else:
            pass
            #print ( " _lont_zip_file is not available, skipping the set filename .. " )
    
    
    #update xml can happen after the for loop completeion, as all the filenames are the same , except the gas and non-gas, the dictionary will get the keys appear only oce    
    update_sap_xml("filename_info", _new_xml)   
    #print ("set filename dictionary :", _setfilenames_dict )    


def copy_files(_new_xml ): 
    print ("at copy_files .. ")
    _copy_dict = {}
    #_target = r'\\bosch.com\dfsrb\DfsDE\DIV\CM\AI\SW_Production\Nissan\0047_RN_AIVI_7513750800\Test_folder'
    #print ("\n target : ", _target)
    
    #print ("set filename dictionary :", _setfilenames_dict )  
    for _key, _value in _setfilenames_dict.items():
        if ( "_cts_filename" in _key ):   
            _cts_dest = _target +  "\\QFIL_Download\\"
            _cts_dest_file = _cts_dest + _setfilenames_dict["_cts_filename"]
            _source = _cts_ver_file
            _dest = _cts_dest_file
            start_copy(_source, _dest) 
            verify_md5sum("cts_file_md5", _dest)
            
        if ( "_cts_config_filename" in _key ):   
            _cts_conf_dest = _target +  "\\CTS_Configuration\\"
            _cts_conf_dest_file = _cts_conf_dest + _setfilenames_dict["_cts_config_filename"]
            _source = _cts_cpver_file
            _dest = _cts_conf_dest_file
            start_copy(_source, _dest) 
            verify_md5sum("cts_config_file_md5", _dest)
            
        if ( "_PD_filename" in _key ):   
            _pd_dest = _target +  "\\PD_Configuration\\"
            _pd_dest_file = _pd_dest + _setfilenames_dict["_PD_filename"]
            _source = _pd_ver_file
            _dest = _pd_dest_file
            start_copy(_source, _dest)
            verify_md5sum("PD_config_file_md5", _dest) 
            
            #CFS file would not be copied as part of the release, it will already be avilable at the tools path
            #if ( "_CFS_filename" in _key ):   
           
        if ("_app_gfilename" in _key ):
            _gapp_dest_file = _cts_dest + _setfilenames_dict["_app_gfilename"]
            _source = _app_gver_file
            _dest = _gapp_dest_file
            start_copy(_source, _dest)
            verify_md5sum("appsw_gas_file_md5", _dest)
        if ("_app_ngfilename" in _key ):
            _ngapp_dest_file = _cts_dest + _setfilenames_dict["_app_ngfilename"]
            _source = _app_ngver_file
            _dest = _ngapp_dest_file
            start_copy(_source, _dest)
            verify_md5sum("appsw_nongas_file_md5", _dest)
        
        _aurix_supp_dest = _target + "\\Support_Files\\"        
        if ( "_aurix_dnl_filename" in _key ):   
            for _cont_name in _container_names_list:
                _aurix_dest = _target + "\\" + _cont_name +  "\\Product_SW\\"
                _aurix_dnl_dest_file = _aurix_dest + _setfilenames_dict["_aurix_dnl_filename"]
                _source = _aurix_dnl_file
                _dest = _aurix_dnl_dest_file
                start_copy(_source, _dest) 
                verify_md5sum("aurix_dnl_file_md5", _dest)
        if ( "_aurix_srec_support_filename" in _key ): 
            if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:    
                _aurix_srec_dest_file = _aurix_supp_dest + _setfilenames_dict["_aurix_srec_support_filename"]
                _source = _aurix_srec_file
                _dest = _aurix_srec_dest_file
                start_copy(_source, _dest)
                verify_md5sum("aurix_srec_file_md5", _dest) #using the same key for aurix srec in BU and EU as either of it exists at any point of time
       
        if ( "_aurix_srec_filename" in _key ):         
            if "Baseunit" in _tmpl_dir and "BU" in _sw_ver.split("_")[-1]:
                for _cont_name in _container_names_list:
                    _aurix_dest = _target + "\\" + _cont_name +  "\\Product_SW\\"
                    _aurix_srec_dest_file = _aurix_dest + _setfilenames_dict["_aurix_srec_filename"]
                    _source = _aurix_srec_file
                    _dest = _aurix_srec_dest_file
                    start_copy(_source, _dest) 
                    verify_md5sum("aurix_srec_file_md5", _dest) 
 
        if ( "_aurix_disHsmdnl_support_filename" in _key ):    
            aurix_disHsmdnl_dest_file = _aurix_supp_dest + _aurix_disHsmdnl_support_filename
            _source = _aurix_support_file
            _dest = aurix_disHsmdnl_dest_file
            start_copy(_source, _dest)
            verify_md5sum("aurix_suport_file_md5",_dest)
         
        if ( "_ucb_config_filename" in _key ):   
            _ucb_dest = _target +  "\\UCB_Configuration\\"
            _bin_str = "UCB_Configuration"
            _ucb_file = _setfilenames_dict["_ucb_config_filename"]
            _source_path = _prod_path +  "\\UCB_Configuration\\"
            _source = _source_path + "\\" + _ucb_file
            _dest = _ucb_dest + "\\" + _ucb_file
            start_copy(_source, _dest)    
        
        if ( "_ublox_filename" in _key ):         
            for _cont_name in _container_names_list:
                _ublox_dest = _target + "\\" + _cont_name +  "\\Product_SW\\"
                _ublox_dest_file = _ublox_dest + _setfilenames_dict["_ublox_filename"]
                _source = _ublox_ver_file
                _dest = _ublox_dest_file
                start_copy(_source, _dest)
                verify_md5sum("ublox_file_md5", _dest)
        
        if "_sxm_filename" in _key:    
            _sxm_container_list = [] 
            for _i in _set_dict:	
                _sxm_copy_status = _set_dict[_i]["sxm"].lower() 
                if _sxm_copy_status == "yes": 
                    _prod_name = _set_dict[_i]["product"]
                    _prod_name = "SW_Container_"+_prod_name
                    _sxm_container_list.append(_prod_name)
            if len(_sxm_container_list) >= 1:      
                for _cont_name in _sxm_container_list:
                    _sxm_dest = _target + "\\" + _cont_name +  "\\Product_SW\\"
                    _sxm_dest_file = _sxm_dest + _setfilenames_dict["_sxm_filename"]
                    _source = _SXM_fwver_file
                    _dest = _sxm_dest_file
                    start_copy(_source, _dest)
                    verify_md5sum("SXM_file_md5", _dest)
                    
                    
                    # print("source:", _source)
                    # print("destination file :", _dest)                
            
                
        if "_dtv_swfilename" in _key:
            _dtv_container_list = []
            for _i in _set_dict:	
                _dtv_copy_status = _set_dict[_i]["dtv"].lower() 
                if _dtv_copy_status == "yes": 
                    _prod_name = _set_dict[_i]["product"]
                    _prod_name = "SW_Container_"+_prod_name
                    _dtv_container_list.append(_prod_name)
            if len(_dtv_container_list) >= 1: 
                for _cont_name in _dtv_container_list:
                    _dtv_dest = _target + "\\" + _cont_name +  "\\Product_SW\\"
                    _dtvsw_dest_file = _dtv_dest + _setfilenames_dict["_dtv_swfilename"]
                    _source = _DTV_swver_file
                    _dest = _dtvsw_dest_file
                    start_copy(_source, _dest)
                    verify_md5sum("DTV_swfile_md5", _dest)
                                       
        
                #have to check if dtv parameter files needs part numbers and has to be product sw folder 
        if "_dtv_p1filename" in _key:
            if len(_dtv_container_list) >= 1:
                _dtvp1_dest_file = _aurix_supp_dest + _setfilenames_dict["_dtv_p1filename"] #_aurix_supp_dest -? same for ublox also
                _source = _DTV_P1ver_file
                _dest = _dtvp1_dest_file
                start_copy(_source, _dest)
                verify_md5sum("DTV_P1file_md5", _dest)
            
        if "_dtv_p2filename" in _key:
            if len(_dtv_container_list) >= 1:
                _dtvp2_dest_file = _aurix_supp_dest + _setfilenames_dict["_dtv_p2filename"] #_aurix_supp_dest -? same for ublox also
                _source = _DTV_P2ver_file
                _dest = _dtvp2_dest_file
                start_copy(_source, _dest)
                verify_md5sum("DTV_P2file_md5", _dest)
        if "_lont_filename" in _key:
            if len(_dtv_container_list) >= 1:
                for _cont_name in _dtv_container_list:
                    _dtv_dest = _target + "\\" + _cont_name +  "\\Product_SW\\"
                    _lont_dest_file = _dtv_dest + _setfilenames_dict["_lont_filename"]                    
                    _source = _lont_ver_file
                    _dest = _lont_dest_file
                    start_copy(_source, _dest)
                    verify_md5sum("lont_file_md5", _dest)
                                
    update_sap_xml("verify_md5_info", _new_xml)    
     # _copy_dict.update({_source: _dest})
    # shutil.copy(_cts_ver_file, _cts_dest_file )
    print ("copy completed .. ")
    
    
 

def start_copy ( _source, _dest ):
    #print ("at start copy .. ")
    # print ("source:", _source) 
    # print ("_dest:", _dest)
    
    #_copy_dict.update({_source: _dest})
    if not os.path.isfile(_dest):
        try:
            shutil.copy(_source, _dest )
        except PermissionError:
            print("Permission denied.")
        except:
            print("Error occurred while copying file.") 
    else:
        #print ("file ", _dest, "exists")
        pass
        
        

def create_productInfo_subElmt_sections(_attr_list):
    _col_list = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8", "Col9"]            
    root = ETS.Element('SW_Overview')
    _pos = 1
        
    if _attr_list[_pos] == "Checksum" :
        En = ETS.SubElement(E6,'Product_Info_SplitCol1')
        #print( "\n attribute at pos :", _attr_list[_pos])
        
    else:    
        En = ETS.SubElement(E6,'Product_Info')   
    
    _x = 0 
    print 
    for _i in _col_list:
        En.attrib[_i] = _attr_list[_x]
        _x +=1
        
def create_productInfo_subElmt_sections_os(_eachlist):
    #print ("\n \n create_productInfo_subElmt_sections_os ...  ", _eachlist)
    global _aurix_header_flag
    global _PD_Tooling_flag
    global _CTS_Parameter_flag
    global _App_Parameter_flag
    global _Sec_Parameter_flag
    global _VIP_Parameter_flag
    # _aurix_header_flag = "notset"
    _col_list = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8", "Col9"]            
    root = ETOS.Element('SW_Overview')
    _pos = 0
    #section_header_dict = {"UFS - CTS": "UFS", "uBlox": "uBlox", "Aurix Test Manager": "Test Manager", "Aurix Autosar": "Aurix", "Activate Aurix": "Lock Activation", "Production Tooling": "Production Download Tooling", "CTS Parameter": "CTS Parameter for Production Download Tooling", "Application Parameter": "Application Parameter for Production Download Tooling", "Secure": "Secure Boot Parameter for Production Download Tooling", "VIP DNL": "VIP DNL Parameter for Production Download Tooling", 
    #can add SBR also later
    if "UFS - CTS" in _eachlist[_pos]:
        Em_os_st = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_st.attrib["Col1"] = "UFS"
        #print ("\n at UFS")
    #root.append(E6_os)
    if "uBlox" in _eachlist[_pos]:
        Em_os_ub = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_ub.attrib["Col1"] = "uBlox"
        #print ("\n at Ublox")
    if "Aurix Test Manager" in _eachlist[_pos]:
        Em_os_tm = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_tm.attrib["Col1"] = "Test Manager"
        #print ("\n at test manager")
    if "Aurix Autosar" in _eachlist[_pos] and _aurix_header_flag == "notset":
        Em_os_au = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_au.attrib["Col1"] = "Aurix" 
        _aurix_header_flag = "set"
        #print ("\n at aurix")
    if "Activate Aurix" in _eachlist[_pos]:
        Em_os_aa = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_aa.attrib["Col1"] = "Lock Activation"
        #print ("\n at lock activation")
    if "Production Tooling" in _eachlist[_pos] and _PD_Tooling_flag == "notset":
        Em_os_pt = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_pt.attrib["Col1"] = "Production Download Tooling"
        _PD_Tooling_flag = "set"
        #print ("\n at PD Tooling")
    if "SXM" in _eachlist[_pos]:
        Em_os_s = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_s.attrib["Col1"] = "SXM"
        #print ("\n at SXM")
    if "DTV SW Version" in _eachlist[_pos]:
        Em_os_d = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_d.attrib["Col1"] = "DTV"
        #print ("\n at DTV")
    if "CTS Parameter" in _eachlist[_pos] and _CTS_Parameter_flag == "notset":
        Em_os_cp = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_cp.attrib["Col1"] = "CTS Parameter for Production Download Tooling"
        _CTS_Parameter_flag = "set"
        #print ("\n at CTS Parameter")
    if "Application Parameter" in _eachlist[_pos] and _App_Parameter_flag == "notset":
        Em_os_ap = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_ap.attrib["Col1"] = "Application Parameter for Production Download Tooling"
        _App_Parameter_flag = "set"
        #print ("\n at App Parameter")
    if "Secure" in _eachlist[_pos] and _Sec_Parameter_flag == "notset":
        Em_os_sp = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_sp.attrib["Col1"] = "Secure Boot Parameter for Production Download Tooling"
        _Sec_Parameter_flag = "set"
        #print ("\n at Secure boot Parameter")
    if "VIP DNL" in _eachlist[_pos] and _VIP_Parameter_flag == "notset":
        Em_os_vp = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_vp.attrib["Col1"] = "VIP DNL Parameter for Production Download Tooling"
        _VIP_Parameter_flag = "set"
        #print ("\n at VIP DNL Parameter")
    if "PD Delivery State" in _eachlist[_pos]:
        Em_os_pd = ETOS.SubElement(E6_os,'SW_Table_Header1')
        Em_os_pd.attrib["Col1"] = "Configuration"
        #print ("\n at Configuration")

    _pos = 1
    if _eachlist[_pos] == "Checksum" :
        En_os = ETOS.SubElement(E6_os,'Product_Info_SplitCol1')
        #print( "\n attribute at pos :", _eachlist[_pos])
        
    else:    
        En_os = ETOS.SubElement(E6_os,'Product_Info')   
    
    _x = 0 
    # print ("\n attributes : ", En_os.attrib )
    # print ("\n e6_os : ", E6_os )
    # print ( "e6 values : ", E6_os.items())
    # print ("_eachlist list  : ", _eachlist)
    for _i in _col_list:
        En_os.attrib[_i] = _eachlist[_x]
        _x +=1
     
def _list_append( _list_name):
    for _eachlist in _list_name:
       _list_of_sections.append(_eachlist)
       
def _list_append_os( _list_name):
    for _eachlist in _list_name:
       _list_of_sections_os.append(_eachlist)
    #print ("\n display _list_of_sections_os :", _list_of_sections_os )
        
def get_hex(_item):
    #hex coversion to get the diag version
    _item_array = list(_item)
    #print ("_item_array: " , _item_array)
    _item_hex=""
    #print ("isascii: ", _SW_ID.isascii())
    for char in _item_array:
       # print ("char: " , char)
        _hex = hex(ord(char))
        _item_hex = (_item_hex + " " + _hex.split("0x")[-1]).upper()
        _item_hex = _item_hex.strip()
       # print ("hex value: ",_item_hex)  
    return(_item_hex)
               

def sos_calls():
    global _Chng_NR
   
    _doc_dict = {}
    if _ecr != "" and _ecr !="None":
        _Chng_NR = _ecr
    elif _ecn.value !="" and _ecn.value != None:        #cell.value would be None if empty , not "None"
        _Chng_NR = _ecn.value
    _Chng_NR = str(_Chng_NR)  
    
    _sosall_filename = _Chng_NR + "_" + _sw_ver + "_SOS-CCS2.xml"
    _pdf_all_filename = _Chng_NR + "_" + _sw_ver + "_SOS-CCS2.pdf"
    
    for _i in _set_dict:
        _PN = _set_dict[_i]["part_number"]
        _set_cont = _set_dict[_i]["product"]
        _set_cont = "SW_Container_"+ _set_cont
        _SAP_PN = _set_dict[_i]["document_number"]
                   
        _sos_filename = _SAP_PN + "_" + _sw_ver + "_SOS-CCS2" + ".xml"  #versioning part can be checked laterw        
        _sosfile = _target + "\\" + _set_cont +  "\\Data_to_plant\\"+ _sos_filename
        Prepare_SOS(_PN, _sosfile)
        
    _overallsos_file = _target + "\\" + _sosall_filename
    Prepare_OverallSOS_EU(_overallsos_file)    
    
    

def Prepare_SOS(_PN, _sosfile):
    global E6
    global _list_of_sections
    _SW_Header_dict = {
        "Col1": "",
        "Col2": "",
        "Col3": "Version Database Key Name",
        "Col4": "SW Version",
        "Col5": "SW Filename",
        "Col6": "Version displayed on screen (or trace)",
        "Col7": "Version readable via production diagnosis",
        "Col8": "Read command",
        "Col9": "Write command"}
      
    _col_list = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8", "Col9"]            
    root = ETS.Element('SW_Overview')
    
    E1 = ETS.Element("BoschPrjName")
    E1.text = _prj
    
    root.append(E1)
   
    tree = ETS.ElementTree(root) 
    
    for _i in _set_dict:
        _PN_set = _set_dict[_i]["part_number"]
        if _PN == _PN_set :
            _SAP_PN = _set_dict[_i]["document_number"]
            _PN_type = _set_dict[_i]["device_type"].strip()
            _dtv_status = _set_dict[_i]["dtv"]
            
                
            _sxm_status = _set_dict[_i]["sxm"]
            _sxm_status = _sxm_status.lower()
            print ("\n at Prepare_SOS for part number :", _PN_set)
            E2 = ETS.Element("BoschPrjNameNr")
            _PN_SOS = _PN_set[0] + " " +  _PN_set[1:4] + " " +  _PN_set[4:7] + " " +  _PN_set[7:10] + " " +  _PN_set[10:]
            E2.text = _PN_SOS
            
            root.append(E2)
            E3 = ETS.Element("BoschSPLNrSet")
            _SPL_no = ETS.SubElement(E3,'BoschSPLNr')
            _SAP_PN_SOS = _SAP_PN[0] + " " +  _SAP_PN[1:4] + " " +  _SAP_PN[4:7] + " " +  _SAP_PN[7:10] + " " +  _SAP_PN[10:]
            _SPL_no.text = _SAP_PN_SOS
            root.append(E3)
            
            E4 = ETS.Element("ProjectInfo")
            #print ("\n _Project_Info_dict: ", _Project_Info_dict )
            for _key, _value in _Project_Info_dict.items():
                if _key == "CHNG_NR":
                 E4.attrib[_key] = _Chng_NR[0:10] + " " + _Chng_NR[10:12]
                else:
                    # print ("\n key: ", _key )
                    # print ("\n _PN_type: ", _PN_type )
                    if _PN_type == "yes":   
                        if _key != "SW_ID_NONGAS":
                            E4.attrib[_key] = _value            #exclude the non-gas sw-id if it is 
                           
                            
                    elif _PN_type ==  "non":
                        if _key != "SW_ID_GAS":
                            E4.attrib[_key] = _value 
                            #_SW_ID = E4.attrib["SW_ID_NONGAS"]
                            
                    else:                   #base unit
                        if "SW_ID_" not in _key:
                            E4.attrib[_key] = _value 
            if _PN_type == "yes":
                _SW_ID = E4.attrib["SW_ID_GAS"]
                _Ver_ID1 = _app_gver 
                print ("SWID value: ",_SW_ID)
                _app_filename =  _app_gfilename		#to uncomment
            elif _PN_type ==  "non": 
                _SW_ID = E4.attrib["SW_ID_NONGAS"]
                _Ver_ID1 = _app_ngver
                print ("SWID value: ",_SW_ID)
                _app_filename =  _app_ngfilename		#to uncomment
            if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:    
                E4.attrib["SW_ID"] = _SW_ID
                _SW_ID_hex = get_hex(_SW_ID)
            else:
                E4.attrib["SW_ID"] = ""
            #print ("\n attributes :", E4.attrib )
            # _SW_ID_hex = get_hex(_SW_ID)
            
            root.append(E4) 
                        
            E5 = ETS.Element("DocInfo")
            _today = dt.now()
            _date = _today.strftime("%a, %d.%m.%Y,%X")
            _Doc_Info_dict["Doc_Date"] = _date
            _Doc_Info_dict["Doc_State"] = "REL"
            for _key, _value in _Doc_Info_dict.items():
                E5.attrib[_key] = _value 
            
            root.append(E5)
            
            E6 = ETS.Element("SW_Versions")
            
            E7 = ETS.SubElement(E6,'SW_Table_Header')
            
            
            for _key, _value in _SW_Header_dict.items():
                E7.attrib[_key] = _value 
            #handling for md5 has to be added, may be mark it as MANUALLY
            # _app_filename = "test_app_filename" #test, lines 1017 and 1021 to be uncommented
            # _cts_filename = "test_filename"  #to be deleted , this is just for test, will be formed if set_filenames runs
            # _aurix_dnl_filename = "test_aurix_dn_filename" # to be commented later - added for test now
            # _cts_config_filename = "test_cts_config_filename"   # to be commented later  
            # _md5_dict["CFS_file_md5"] = "test_md5"
            # _md5_dict["PD_config_file_md5"] = "test_md5"
            # _md5_dict["cts_config_file_md5"] = "test_md5"
            
            _PD_config_md5 = _md5_dict["PD_config_file_md5"] + " [3]"
            _CFS_PTooling_md5 = _md5_dict["CFS_file_md5"] + " [3]"
            _cts_config_md5 = _md5_dict["cts_config_file_md5"] + " [3]"
            _ublox_sos_ver = "EXT CORE " + _ublox_info["version"] 
            _tm_title = _tm_info["title"] + " [1]"
            _PT_title = _PTooling_info["title"] + " [2] [12] [13]"
            _PT_BE_title = _PTooling_info["title_BE"] + " [2] [12] [13]"
            _PT_tar_file = _PTooling_info["tar_file"]
            _PD_Del_title = _PDDel_info["title"] + " [4]"
            _CTS_Del_title = _cts_info["title_del"] + " [5]"
            _UCB_Del_title = _UCB_Del_info["title"] + " [6]"
            _CTS_RXL_title = _CTSPar_info["title"] + " [10]"
            _App_RXL_title = _AppPar_info["title"] + " [10]"
            _Sec_RXL_title = _SecPar_info["title"] + " [10]"
            _VIP_RXL_title = _VIPPar_info["title"] + " [10]"
            
            _UCB_Del_md5 =  _UCB_Del_info["UCB_md5"] + "[3]"  #To be modified in future
            #_tm_diag = " 53 15 " #hardcoded now, has to find what conversion it is 
            _tm_diag_1 = int(_tm_info["version"].split(".0V")[0])
            _tm_diag_2 = int(_tm_info["version"].split(".0V")[-1])
            _tm_diag = hex(_tm_diag_1)[2:] + " " + hex( _tm_diag_2)[2:]
            
            
            _crypto_b_diag_ver = get_hex(_crypto_B_info["version"])
            _crypto_c_diag_ver = get_hex(_crypto_C_info["version"])
            _aurix_diag_ver = get_hex(_aurix_ver)
            _cts_ver_ci = _cts_ver.upper()
            _cts_ver_short = _cts_ver_ci.split("PRJ_CCS2_CTS_")[-1]         
            _cts_diag_ver = get_hex(_cts_ver.upper())
            if _PN_type == "non" or _PN_type == "yes":
                _app_diag_ver = get_hex(_Ver_ID1.upper())
            _sbr_diag_ver = get_hex(_SBR_info["label"]) + " 00"
            _sbr_version = _SBR_info["label"].split(" ", 1)[0]   #split based on the 1st occurence - removes the date part to just have the version
            #in the update master for sap xml -> add the date to sbrinfo date 
            _sxm_diag_ver = get_hex(_sxm_info["version"])
            _F4Linux_diag_ver = get_hex(_F4Linux_info["version"]) + " 0A"  #0A set as default as it is linix line feed 
            
            #lengthstarts from 1
            #positioning starts from 0
            #append 00 to aurix diag version to fill up 50 bytes
            while len(_aurix_diag_ver) <= 148:
                _aurix_diag_ver = _aurix_diag_ver + " 00"
                
            
            if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:    
                _ublox_info["diag_ver"] = ""
                _SW_info_list = ["SW ID", "", _app_sw_key, _SW_ID, "","-", _SW_ID_hex, _app_read_comm,"n.a."]
                _app_list = [_app_info["title"], "", _app_info["db_key"], _Ver_ID1.upper(), _app_filename, "-", _app_diag_ver, _app_info["db_key_read"], "n.a."]
                _aurix_srec_list = [ _aurix_info["title"], "", _aurix_info["db_key_aurix_srec"], _aurix_ver, "n.a.", "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]
                _ucb_lock_list = [ _UCB_Lock_info["title"], "", _UCB_Lock_info["db_key"], _UCB_Lock_info["status"], "n.a.", "n.a.", "n.a.", "n.a.","-"]
                _sec_lock_list = [ _Sec_Lock_info["title"], "", _Sec_Lock_info["db_key"], _Sec_Lock_info["status"], "n.a.", "n.a.", "n.a.", "n.a.","-"]        
                _APP_LUN_list = [ _AppPar_info["title"], "", _AppPar_info["db_key_erase_lun"], _AppPar_info["app_param_erase_lun"], "-", "n.a.", "n.a.", "n.a.","-"]
                _APP_OP_list = [ _AppPar_info["title"], "", _AppPar_info["db_key_op"], _AppPar_info["app_param_op"], "-", "n.a.", "n.a.", "n.a.","-"]
                _APP_RXL_list = [ _App_RXL_title, "", _AppPar_info["db_key_rxl"], _AppPar_info["app_param_rxl"], "-", "n.a.", "n.a.", "n.a.","-"]
                _APP_VA_list = [ _AppPar_info["title"], "", _AppPar_info["db_key_va"], _AppPar_info["app_param_va"], "-", "n.a.", "n.a.", "n.a.","-"]
                _APP_PXL_list = [ _App_RXL_title, "", _AppPar_info["db_key_pxl"], _AppPar_info["app_param_pxl"], "-", "n.a.", "n.a.", "n.a.","-"]
                _SEC_OP_list = [ _SecPar_info["title"], "", _SecPar_info["db_key_op"], _SecPar_info["sec_param_op"], "-", "n.a.", "n.a.", "n.a.","-"]
                _SEC_RXL_list = [ _Sec_RXL_title, "", _SecPar_info["db_key_rxl"], _SecPar_info["sec_param_rxl"], "-", "n.a.", "n.a.", "n.a.","-"]
                _SEC_VA_list = [ _SecPar_info["title"], "", _SecPar_info["db_key_va"], _SecPar_info["sec_param_va"], "-", "n.a.", "n.a.", "n.a.","-"]
                _SEC_PXL_list = [ _Sec_RXL_title, "", _SecPar_info["db_key_pxl"], _SecPar_info["sec_param_pxl"], "-", "n.a.", "n.a.", "n.a.","-"]
                _VIP_OP_list =  [ _VIPPar_info["title"], "",_VIPPar_info["db_key_op"], _VIPPar_info["vip_param_op"], "-", "n.a.", "n.a.", "n.a.", "-"]
                _VIP_RXL_list =  [ _VIP_RXL_title, "",_VIPPar_info["db_key_rxl"], _VIPPar_info["vip_param_rxl"], "-", "n.a.", "n.a.", "n.a.", "-"]
                _VIP_VA_list =  [ _VIPPar_info["title"], "",_VIPPar_info["db_key_va"], _VIPPar_info["vip_param_va"], "-", "n.a.", "n.a.", "n.a.", "-"]
                _VIP_PXL_list =  [ _VIP_RXL_title, "",_VIPPar_info["db_key_pxl"], _VIPPar_info["vip_param_pxl"], "-", "n.a.", "n.a.", "n.a.", "-"]
                
                    
            else:
                _SW_info_list = ["SW ID", "", _app_sw_key, "-", "","-", "-", _app_read_comm,"n.a."]
                _app_list = [_app_info["title"], "", _app_info["db_key"], "-", "not-delivered", "-", "-", _app_info["db_key_read"], "n.a."]
                _aurix_srec_list = [ _aurix_info["title"], "", _aurix_info["db_key_aurix_srec"], _aurix_ver, _aurix_srec_filename, "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]
                
            _Overall_SW_info_list =  ["Overall SW Version", "", _ecu_sw_key, _sw_ver, "-", "n.a.", "n.a.", "n.a.", "-"]
            _android_version_list = [_android_version_info["title"], "", _android_version_info["db_key"], _android_version_info["version"],"n.a.","n.a.","n.a.","n.a.","-"]    
            _Crypto_Bosch_list = [_crypto_B_info["title"], "", _crypto_B_info["db_key"], _crypto_B_info["version"],"n.a.","n.a.", _crypto_b_diag_ver, _crypto_B_info["db_key_read"], "-"]
            _Crypto_Cust_list = [_crypto_C_info["title"], "", _crypto_C_info["db_key"], _crypto_C_info["version"],"n.a.","n.a.", _crypto_c_diag_ver, _crypto_C_info["db_key_read"], "-"]
            _oem_list = [_oem_info["title"], "", _oem_info["db_key"], _oem_info["version"],"n.a.","n.a.","n.a.","n.a.","-"]
            _cts_list = [_cts_title, "", _cts_info["db_key"], _cts_ver_short, _cts_filename, _cts_ver.upper() , _cts_diag_ver, _cts_info["db_key_read"], "-"]
            #_app_list = [_app_info["title"], "", _app_info["db_key"], _Ver_ID1.upper(), _app_filename, "", _app_diag_ver, _app_info["db_key_read"], "n.a."]
            _ublox_list = [_ublox_info["title"], "", _ublox_info["db_key"], _ublox_sos_ver, _ublox_filename, "n.a.", _ublox_info["diag_ver"], _ublox_info["db_key_read"], "-"]
            _ublox_split_list = ["", "Checksum", _ublox_info["db_key_crc"], _ublox_info["product"], "-", "n.a.", "MANUALLY", _ublox_info["db_key_crc_read"], "-"]
            _SBR_list = [_SBR_info["product"], "", _SBR_info["db_key"], _sbr_version.strip(), "-", _SBR_info["label"], _sbr_diag_ver, _SBR_info["db_key_read"], "-"]
            _F4Linux_list = [_F4Linux_info["product"], "", _F4Linux_info["db_key"], _F4Linux_info["version"], "-", _F4Linux_info["label"], _F4Linux_diag_ver, _F4Linux_info["db_key_read"], "-"]
            if _sxm_status == "yes":
                _SXM_list = [ _sxm_info["key"], "", _sxm_info["db_key"], _sxm_info["version"], _sxm_filename, "n.a.", _sxm_diag_ver, _sxm_info["db_key_read"], "-"] 
            if _dtv_status == "yes":
                _JPN_PN = "yes"
                _DTV_FW_title = _DTV_SW_info["title"] + " [7] [8] [14]"
                _DTV_BSW_title = _DTV_BSW_info["title"] + " [7]"
                _DTV_HW_title = _DTV_HSW_info["title"] + " [7] [9]"
                _DTV_P1_title = _DTV_Par1_info["title"] + " [7] [14]"
                _DTV_P2_title = _DTV_Par2_info["title"] + " [7] [14]"
                _Lont_title = _Lont_info["title"] + " [11]"
                _dtv_fw_diag_ver = get_hex(_DTV_SW_info["version"])
                _dtv_bsw_diag_ver = get_hex(_DTV_BSW_info["version"])
                _dtv_hw_diag_ver = get_hex(_DTV_HSW_info["version"])
                _dtv_hw_ver = _DTV_HSW_info["version"]
                _dtv_bsw_ver = _DTV_BSW_info["version"]
                _dtv_fw_ver = _DTV_SW_info["version"]
                _lont_ver = _Lont_info["version"]
                _lont_diag_ver = _lont_ver[1:3] + " "  + _lont_ver[3:5] + " " + _lont_ver[5:]
                # _dtv_par1_ver = _DTV_Par1_info["version"]
                # _dtv_par2_ver = _DTV_Par2_info["version"]
                _dtv_hw_ver_label = _DTV_HSW_info["label"].replace("DTV_HW_VER",_dtv_hw_ver )
                _dtv_bsw_ver_label = _DTV_BSW_info["label"].replace("DTV_BOOT_VER",_dtv_bsw_ver )
                _dtv_fw_ver_label = _DTV_SW_info["label"].replace("DTV_SW_VER",_dtv_fw_ver )
                _DTV_P1_md5 = _md5_dict["DTV_P1file_md5"] + " [3]"
                _DTV_P2_md5 = _md5_dict["DTV_P2file_md5"] + " [3]"
                
                #_DTV_FW_list = [_DTV_FW_title, "", _DTV_SW_info["db_key"], _DTV_SW_info["version"], _DTV_SW_info["file"], _dtv_fw_ver_label, _dtv_fw_diag_ver, _DTV_SW_info["db_key_read"],""] 
                _DTV_FW_list = [_DTV_FW_title, "", _DTV_SW_info["db_key"], _DTV_SW_info["version"], _dtv_swfilename, _dtv_fw_ver_label, _dtv_fw_diag_ver, _DTV_SW_info["db_key_read"],""] 
                _DTV_BSW_list = [_DTV_BSW_title, "", _DTV_BSW_info["db_key"], _DTV_BSW_info["version"], "", _dtv_bsw_ver_label, _dtv_bsw_diag_ver, _DTV_BSW_info["db_key_read"],""]
                _DTV_HSW_list = [_DTV_HW_title, "", _DTV_HSW_info["db_key"], _DTV_HSW_info["version"], "", _dtv_hw_ver_label, _dtv_hw_diag_ver, _DTV_HSW_info["db_key_read"], ""]
                _DTV_P1_list = [_DTV_P1_title, "", _DTV_Par1_info["db_key"], _DTV_Par1_info["version"], _DTV_Par1_info["file"], _DTV_P1_md5, _DTV_Par1_info["diag_ver"], "-", "-"]
                _DTV_P2_list = [_DTV_P2_title, "", _DTV_Par2_info["db_key"], _DTV_Par2_info["version"], _DTV_Par2_info["file"], _DTV_P2_md5, _DTV_Par2_info["diag_ver"], "-", "-"]
                _Lont_list = [ _Lont_title, "", _Lont_info["db_key"], _Lont_info["version"], _lont_filename, "-", _lont_diag_ver, _Lont_info["db_key_read"], ""]
               
            else:
                _JPN_PN = "no"
            
            
                
            _aurix_TM_list = [ _tm_title, "", _tm_info["db_key"], _tm_info["version"], "n.a.", "n.a.", _tm_diag, _tm_info["db_key_read"], "-"]
            _aurix_list = [ _aurix_info["title"], "", _aurix_info["db_key"], _aurix_ver, "-", "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]            
            _aurix_dnl_list = [ _aurix_info["title"], "", _aurix_info["db_key_aurix_dnl"], _aurix_ver, _aurix_dnl_filename, "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]
            _PTooling_list = [ _PT_title, "", _PTooling_info["db_key"], _PTooling_info["version"], _PTooling_info["file"], "n.a.", "n.a.", "n.a.", "-"]
            _PTooling_md5_list = [ _PT_title, "", _PTooling_info["db_key_crc"], _PTooling_info["version"], _PTooling_info["file"], _CFS_PTooling_md5, "n.a.", "-", "-"]
            _PTooling_BE_list = [ _PT_BE_title, "", _PTooling_info["db_key_ProdTool_BE"], _PTooling_info["BE_tool"], "-", "-", "-", "-", "-"]
            _CTS_LUN_list = [ _CTSPar_info["title"], "", _CTSPar_info["db_key_erase_lun"], "n.a.", "-", "n.a.", "n.a.", "n.a.","-"]
            _CTS_OP_list = [ _CTSPar_info["title"], "", _CTSPar_info["db_key_op"], _CTSPar_info["cts_param_op"], "-", "n.a.", "n.a.", "n.a.","-"]
            _CTS_RXL_list = [ _CTS_RXL_title, "", _CTSPar_info["db_key_rxl"], _CTSPar_info["cts_param_rxl"], "-", "n.a.", "n.a.", "n.a.","-"]
            _CTS_VA_list = [ _CTSPar_info["title"], "", _CTSPar_info["db_key_va"], _CTSPar_info["cts_param_va"], "-", "n.a.", "n.a.", "n.a.","-"]
            _CTS_PXL_list = [ _CTS_RXL_title, "", _CTSPar_info["db_key_pxl"], _CTSPar_info["cts_param_pxl"], "-", "n.a.", "n.a.", "n.a.","-"]
            _PD_Del_list = [ _PD_Del_title, "", _PDDel_info["db_key_crc"], _PDDel_info["version"], _PDDel_info["file"], _PD_config_md5, "n.a.","-","-"] 
            _CTS_Del_list = [ _CTS_Del_title, "", _cts_info["db_key_crc"], _cts_info["CTS_config_version"], _cts_config_filename, _cts_config_md5, "n.a.", "-", "-"]
            _UCB_Del_list = [ _UCB_Del_title, "", _UCB_Del_info["db_key_crc"], _UCB_Del_info["version"], _UCB_Del_info["file"], _UCB_Del_md5, "n.a.", "-", "-"]
            
                    
            _PD_target = _target + "\\PD_Configuration"
            _CTS_target = _target + "\\CTS_Configuration"
            _UCB_target = _target + "\\UCB_Configuration"
            _Support_target = _target + "\\Support_Files"
            
            _test_manager_ver = _test_manager.attrib["file"].split(".zip")[0]
            
            _com_1 = "[1] TM is not preprogrammed. TM is RAM loaded by CTS during operation. Used Version in CTS SW is " + _test_manager_ver + "(see SW Version and Version readable via production diagnosis).."
            _com_2 = "[2] Base path Production Tooling - \\\\bosch.com\\dfsrb\\DfsDE\\DIV\\CM\\AI\\SW_Production\\Nissan\\0060_CCS2_7515752366\\01_Tools\\production_tooling"
            _com_3 = "[3] MD5 checksum"
            _com_4 = "[4] PD Configuration can be found in: " + _PD_target
            _com_5 = "[5] CTS Configuration can be found in: " + _CTS_target
            _com_6 = "[6] UCB Configuration can be found in: " + _UCB_target
            _com_7 = "[7] DTV Versions see https://hi-dms.de.bosch.com/docushare/dsweb/Services/Document-1179901"
            _com_8 = "[8] To update the Firmware Version with the Version provided in the eMMC-Image (Firmware-Partition) use the routine to update all submodules: 31 01 11 10"
            _com_9 = "[9] Config Parameter and System (ASS_DTV_CONFIG) see https://hi-dms.de.bosch.com/docushare/dsweb/GetVersion/File-1034235/4"
            _com_10 = "[10] These keys should actually be in quotes when passed to the command line of the Production Download Tooling. This should be described in PAVE and implemented in UTS"
            _com_11 = "[11] Only valid for JPN-Variants."
            _com_12 = "[12] Relation between CFS CLI and SOS and PAVE and Station Setup is described in Docupedia: https://inside-docupedia.bosch.com/confluence/display/GG/4.2+CCS2+Project+-+CFS+CLI+to+PAVE+Description"
            _com_13 = "[13] According to CFS Documentation additional Qualcomm Packages are needed: " + _PT_tar_file 
            _com_14 = "[14] DTV FW and Parameter can be found in: " + _Support_target
            
            if _JPN_PN == "yes":
                _com_sec_list = [_com_1, _com_2, _com_3, _com_4, _com_5, _com_6, _com_7, _com_8, _com_9, _com_10, _com_11, _com_12, _com_13, _com_14]                
            else:
                _com_sec_list = _com_sec_list = [_com_1, _com_2, _com_3, _com_4, _com_5, _com_6, _com_10, _com_12, _com_13]
                       
            _list_of_sections = []
            
            _list_items_1 = [ _SW_info_list, _Overall_SW_info_list, _android_version_list, _Crypto_Bosch_list, _Crypto_Cust_list, _oem_list, _cts_list, _app_list,_ublox_list,_ublox_split_list, _SBR_list, _F4Linux_list ]
            
            _list_append( _list_items_1 )
            
            
            if _sxm_status == "yes":
                _list_of_sections.append(_SXM_list)
            
            if _dtv_status == "yes":
                _list_items_2 = [ _DTV_FW_list, _DTV_BSW_list, _DTV_HSW_list, _DTV_P1_list, _DTV_P2_list ] 
                for _eachlist in _list_items_2:
                    _list_of_sections.append(_eachlist)
            
            _list_items_3 = [ _aurix_TM_list, _aurix_list, _aurix_srec_list, _aurix_dnl_list]
            for _eachlist in _list_items_3:
                _list_of_sections.append(_eachlist)
            
            
            if _dtv_status == "yes":
                _list_of_sections.append(_Lont_list)
            if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:  
                _list_items_4 = [ _ucb_lock_list, _sec_lock_list]
                for _eachlist in _list_items_4:
                    _list_of_sections.append(_eachlist)
            
            _list_items_5 = [_PTooling_list, _PTooling_md5_list, _PTooling_BE_list, _CTS_LUN_list, _CTS_OP_list, _CTS_RXL_list, _CTS_VA_list, _CTS_PXL_list ]            
            for _eachlist in _list_items_5:
                _list_of_sections.append(_eachlist)
            
            if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:              
                _list_items_6 = [ _APP_LUN_list, _APP_OP_list, _APP_RXL_list, _APP_VA_list, _APP_PXL_list, _SEC_OP_list, _SEC_RXL_list, _SEC_VA_list, _SEC_PXL_list, _VIP_OP_list, _VIP_RXL_list, _VIP_VA_list, _VIP_PXL_list ]
                for _eachlist in _list_items_6:
                    _list_of_sections.append(_eachlist)
                
            _list_items_7 = [ _PD_Del_list, _CTS_Del_list, _UCB_Del_list ]
            for _eachlist in _list_items_7:
                _list_of_sections.append(_eachlist)
           
            
            for _each in _list_of_sections:
                create_productInfo_subElmt_sections(_each)
            
            #Comments section
            E8 = ETS.SubElement(E6,'SW_Comments')
            for _com in range(len(_com_sec_list)):
                Em = ETS.SubElement(E8,'Comment')
                Em.text = _com_sec_list[_com ]
            
            
            root.append(E6)
            
            
       
    ETS.indent(tree, space='  ', level=0)
    tree.write(_sosfile, encoding='utf-8', xml_declaration=True)
    _pdffile = _sosfile.split(".xml")[0] 
    _pdffile = _pdffile + ".pdf"
    cmd = ["perl", "xml2odxe.pl", "-xml", _sosfile ]
    subprocess.run(cmd)
    print ("\n odx-e file created for _sosfile")
    cmd2 = ["sh", fop_path, "-xsl", "SOS_pdf_ccs2.xsl", "-xml", _sosfile, "-pdf", _pdffile ]
	# cmd2 = [fop_path, "-xsl", "SOS_pdf_ccs2.xsl", "-xml", _sosfile, "-pdf", _pdffile ]
    subprocess.run(cmd2)
    print ("\n pdf file created for _sosfile ")
    
    
def Prepare_OverallSOS_EU(_overallsos_file):
    print("\n at Prepare_OverallSOS_EU ... ")
    global _list_of_sections_os
    global E6_os
    _SW_Header_dict = {
        "Col1": "",
        "Col2": "",
        "Col3": "Version Database Key Name",
        "Col4": "SW Version",
        "Col5": "SW Filename",
        "Col6": "Version displayed on screen (or trace)",
        "Col7": "Version readable via production diagnosis",
        "Col8": "Read command",
        "Col9": "Write command"}
      
    _col_list = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7", "Col8", "Col9"]            
    Set_Info_col_list  = [ "Col1", "Col2", "Col3" ] 
    Set_Info_col_list_2 = [ "Col1", "Col2", "Col3", "Col4", "Col5", "Col6" ]
    
    root = ETOS.Element('SW_Overview')   
    #E1_os = ETOS.Element("BoschPrjName")    #os =>overall sos
    E1_os = ETOS.SubElement(root, "BoschPrjName")
    E1_os.text = _prj
    
    tree = ETOS.ElementTree(root) 
    _PN_list = [] 
    _SAP_PN_list = []
    _PN_type_list = []
    _cont_list = []
    Set_Info_Col4_dict = {} #for keys
    _Set_Info_PNInfoList = []
    _Set_Info_PNInfoList_2 = []
    _Set_Info_PNInfoList_all = []
    _Set_Info_PNInfoList_all_2 = []
    # _dtv_status = "no"
    # _sxm_status = "no"
    _add_sxm_section = "no"
    _add_dtv_section = "no"
    _set_gasid_status = "no"
    _set_ngasid_status = "no"
    
    
    for _i in _set_dict:
        _dtv_status = "no"
        _sxm_status = "no"
        _pn_f = _set_dict[_i]["part_number"]
        _PN_format = _pn_f[0:4] + "." +  _pn_f[4:7] + "." +  _pn_f[7: ]
        _Set_Info_PNInfoList = [ _set_dict[_i]["product"], _PN_format, _set_dict[_i]["document_number"]]
        _Set_Info_PNInfoList_all.append(_Set_Info_PNInfoList)
        _PN_list.append(_set_dict[_i]["part_number"])
        _cont_list.append(_set_dict[_i]["product"])
        _SAP_PN_list.append(_set_dict[_i]["document_number"])
        _PN_type_list.append(_set_dict[_i]["device_type"])
        
        _x = 0 
        if _set_dict[_i]["device_type"] == "yes"  and _set_gasid_status == "no":
           _set_gasid_status = "yes"
           _SW_ID_gas = _Project_Info_dict["SW_ID_GAS"]
           _Ver_ID_gas = _app_gver
           _SW_gID_hex = get_hex(_SW_ID_gas)
           
            
        if _set_dict[_i]["device_type"] == "non"  and _set_ngasid_status == "no":
           _set_ngasid_status = "yes"
           _SW_ID_ngas = _Project_Info_dict["SW_ID_NONGAS"]
           _Ver_ID_ngas = _app_ngver
           _SW_ngID_hex = get_hex(_SW_ID_ngas)
       
        # if _dtv_status != "yes":
            # _dtv_status = _set_dict[_i]["dtv"]     #set only once
        if _set_dict[_i]["dtv"] == "yes":
            _dtv_status = (_set_dict[_i]["dtv"]).lower()
            _add_dtv_section = "yes"
            
        #if _sxm_status != "yes" and _set_dict[_i]["sxm"] == "yes":
        if _set_dict[_i]["sxm"] == "yes":
            _sxm_status = (_set_dict[_i]["sxm"]).lower()
            _add_sxm_section = "yes"
            #_sxm_status = _sxm_status.lower()
        
        _Set_Info_PNInfoList_2 = [  _set_dict[_i]["product"],"cts", "test_manager", "uBlox", _sxm_status, _dtv_status]
        _Set_Info_PNInfoList_all_2.append(_Set_Info_PNInfoList_2)

    
    for _pn in _PN_list:  
        _PN_SOS = _pn[0] + " " +  _pn[1:4] + " " +  _pn[4:7] + " " +  _pn[7:10] + " " +  _pn[10:]
        ETOS.SubElement(root,"BoschPrjNameNr").text = _PN_SOS
        
        
   
    E3_os = ETOS.Element("BoschSPLNrSet")  
    for _spl_no in _SAP_PN_list:
        _SAP_PN_SOS = _spl_no[0] + " " +  _spl_no[1:4] + " " +  _spl_no[4:7] + " " +  _spl_no[7:10] + " " +  _spl_no[10:]
        ETOS.SubElement(E3_os,'BoschSPLNr').text = _SAP_PN_SOS    
    root.append(E3_os)
    
           
    
    E4_os = ETOS.Element("ProjectInfo")
   
    for _key, _value in _Project_Info_dict.items():
        if _key == "CHNG_NR":
            E4_os.attrib[_key] = _Chng_NR[0:10] + " " + _Chng_NR[10:12]
        elif _key != "SW_ID_NONGAS" and _key != "SW_ID_GAS":
            E4_os.attrib[_key] = _value
        if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:    
            E4_os.attrib["SW_ID"] = "See part number specific SOS"
            
        else:
            E4_os.attrib["SW_ID"] = ""
    root.append(E4_os) 
    

    E5_os = ETOS.Element("DocInfo")
    _today = dt.now()
    _date = _today.strftime("%a, %d.%m.%Y,%X")
    _Doc_Info_dict["Doc_Date"] = _date
    _Doc_Info_dict["Doc_State"] = "REL"
   
    for _key, _value in _Doc_Info_dict.items():
        E5_os.attrib[_key] = _value 
      
    root.append(E5_os)
    E6_os = ETOS.Element("SW_Versions")        
    E7_os = ETOS.SubElement(E6_os,'SW_Table_Header')
    
    for _key, _value in _SW_Header_dict.items():
        E7_os.attrib[_key] = _value 
    root.append(E7_os)
    
    
    # _app_gfilename  = "test_app_filename"      # to be commented later 
    # _app_ngfilename = "test_app_filename"       # to be commented later
    # _app_filename = "test_app_filename" #test, lines 1017 and 1021 to be uncommented
    # _cts_filename = "test_filename"  #to be deleted , this is just for test, will be formed if set_filenames runs
    # _aurix_dnl_filename = "test_aurix_dn_filename" # to be commented later - added for test now
    # _cts_config_filename = "test_cts_config_filename"   # to be commented later 
    # _md5_dict["CFS_file_md5"] = "test_md5"
    # _md5_dict["PD_config_file_md5"] = "test_md5"
    # _md5_dict["cts_config_file_md5"] = "test_md5"
            
    _PD_config_md5 = _md5_dict["PD_config_file_md5"] + " [3]"
    _CFS_PTooling_md5 = _md5_dict["CFS_file_md5"] + " [3]"
    _cts_config_md5 = _md5_dict["cts_config_file_md5"] + " [3]"
    _ublox_sos_ver = "EXT CORE " + _ublox_info["version"] 
    _tm_title = _tm_info["title"] + " [1]"
    _PT_title = _PTooling_info["title"] + " [2] [12] [13]"
    _PT_BE_title = _PTooling_info["title_BE"] + " [2] [12] [13]"
    _PT_tar_file = _PTooling_info["tar_file"]
    _PD_Del_title = _PDDel_info["title"] + " [4]"
    _CTS_Del_title = _cts_info["title_del"] + " [5]"
    _UCB_Del_title = _UCB_Del_info["title"] + " [6]"
    _CTS_RXL_title = _CTSPar_info["title"] + " [10]"
    _App_RXL_title = _AppPar_info["title"] + " [10]"
    _Sec_RXL_title = _SecPar_info["title"] + " [10]"
    _VIP_RXL_title = _VIPPar_info["title"] + " [10]"
          
    _UCB_Del_md5 =  _UCB_Del_info["UCB_md5"] + "[3]"  #To be modified in future
    _tm_diag_1 = int(_tm_info["version"].split(".0V")[0])
    _tm_diag_2 = int(_tm_info["version"].split(".0V")[-1])
    _tm_diag = hex(_tm_diag_1)[2:] + " " + hex( _tm_diag_2)[2:]
    #_tm_diag = " 53 15 " #hardcoded now, has to find what conversion it is 
          
            
    _crypto_b_diag_ver = get_hex(_crypto_B_info["version"])
    _crypto_c_diag_ver = get_hex(_crypto_C_info["version"])
    _aurix_diag_ver = get_hex(_aurix_ver)
    _cts_ver_ci = _cts_ver.upper()
    _cts_ver_short = _cts_ver_ci.split("PRJ_CCS2_CTS_")[-1]
    _cts_diag_ver = get_hex(_cts_ver.upper())
    if _set_ngasid_status == "yes":
        _app_ngdiag_ver = get_hex(_Ver_ID_ngas.upper())
    if _set_gasid_status == "yes":
        _app_gdiag_ver = get_hex(_Ver_ID_gas.upper())
    _sbr_diag_ver = get_hex(_SBR_info["label"]) + " 00"
    _sbr_version = _SBR_info["label"].split(" ", 1)[0]
    _sxm_diag_ver = get_hex(_sxm_info["version"])
    _F4Linux_diag_ver = get_hex(_F4Linux_info["version"]) + " 0A"     #0A set as default as it is linix line feed
            
    #lengthstarts from 1
    #positioning starts from 0
    #append 00 to aurix diag version to fill up 50 bytes
    while len(_aurix_diag_ver) <= 148:
       _aurix_diag_ver = _aurix_diag_ver + " 00"
               
            
    if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:    
        _ublox_info["diag_ver"] = ""
        if _set_gasid_status == "yes":
            _SW_ginfo_list = ["SW ID", "", _app_sw_key, _SW_ID_gas, "","-", _SW_gID_hex, _app_read_comm,"n.a."]
            _app_gas_list = [_app_info["title"], "", _app_info["db_key"], _Ver_ID_gas.upper(), _app_gfilename, "-", _app_gdiag_ver, _app_info["db_key_read"], "n.a."]
        if _set_ngasid_status == "yes":
            _SW_nginfo_list = ["SW ID", "", _app_sw_key, _SW_ID_ngas, "","-", _SW_ngID_hex, _app_read_comm,"n.a."]
            _app_ngas_list = [_app_info["title"], "", _app_info["db_key"], _Ver_ID_ngas.upper(), _app_ngfilename, "-", _app_ngdiag_ver, _app_info["db_key_read"], "n.a."]
        _aurix_srec_list = [ _aurix_info["title"], "", _aurix_info["db_key_aurix_srec"], _aurix_ver, "n.a.", "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]
        _ucb_lock_list = [ _UCB_Lock_info["title"], "", _UCB_Lock_info["db_key"], _UCB_Lock_info["status"], "n.a.", "n.a.", "n.a.", "n.a.","-"]
        _sec_lock_list = [ _Sec_Lock_info["title"], "", _Sec_Lock_info["db_key"], _Sec_Lock_info["status"], "n.a.", "n.a.", "n.a.", "n.a.","-"]        
        _APP_LUN_list = [ _AppPar_info["title"], "", _AppPar_info["db_key_erase_lun"], _AppPar_info["app_param_erase_lun"], "-", "n.a.", "n.a.", "n.a.","-"]
        _APP_OP_list = [ _AppPar_info["title"], "", _AppPar_info["db_key_op"], _AppPar_info["app_param_op"], "-", "n.a.", "n.a.", "n.a.","-"]
        _APP_RXL_list = [ _App_RXL_title, "", _AppPar_info["db_key_rxl"], _AppPar_info["app_param_rxl"], "-", "n.a.", "n.a.", "n.a.","-"]
        _APP_VA_list = [ _AppPar_info["title"], "", _AppPar_info["db_key_va"], _AppPar_info["app_param_va"], "-", "n.a.", "n.a.", "n.a.","-"]
        _APP_PXL_list = [ _App_RXL_title, "", _AppPar_info["db_key_pxl"], _AppPar_info["app_param_pxl"], "-", "n.a.", "n.a.", "n.a.","-"]
        _SEC_OP_list = [ _SecPar_info["title"], "", _SecPar_info["db_key_op"], _SecPar_info["sec_param_op"], "-", "n.a.", "n.a.", "n.a.","-"]
        _SEC_RXL_list = [ _Sec_RXL_title, "", _SecPar_info["db_key_rxl"], _SecPar_info["sec_param_rxl"], "-", "n.a.", "n.a.", "n.a.","-"]
        _SEC_VA_list = [ _SecPar_info["title"], "", _SecPar_info["db_key_va"], _SecPar_info["sec_param_va"], "-", "n.a.", "n.a.", "n.a.","-"]
        _SEC_PXL_list = [ _Sec_RXL_title, "", _SecPar_info["db_key_pxl"], _SecPar_info["sec_param_pxl"], "-", "n.a.", "n.a.", "n.a.","-"]
        _VIP_OP_list =  [ _VIPPar_info["title"], "",_VIPPar_info["db_key_op"], _VIPPar_info["vip_param_op"], "-", "n.a.", "n.a.", "n.a.", "-"]
        _VIP_RXL_list =  [ _VIP_RXL_title, "",_VIPPar_info["db_key_rxl"], _VIPPar_info["vip_param_rxl"], "-", "n.a.", "n.a.", "n.a.", "-"]
        _VIP_VA_list =  [ _VIPPar_info["title"], "",_VIPPar_info["db_key_va"], _VIPPar_info["vip_param_va"], "-", "n.a.", "n.a.", "n.a.", "-"]
        _VIP_PXL_list =  [ _VIP_RXL_title, "",_VIPPar_info["db_key_pxl"], _VIPPar_info["vip_param_pxl"], "-", "n.a.", "n.a.", "n.a.", "-"]
                
                    
    else:
        _SW_info_list = ["SW ID", "", _app_sw_key, "-", "","-", "-", _app_read_comm,"n.a."]
        _app_list = [_app_info["title"], "", _app_info["db_key"], "-", "not-delivered", "-", "-", _app_info["db_key_read"], "n.a."]
        _aurix_srec_list = [ _aurix_info["title"], "", _aurix_info["db_key_aurix_srec"], _aurix_ver, _aurix_srec_filename, "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]
               
    _Overall_SW_info_list =  ["Overall SW Version", "", _ecu_sw_key, _sw_ver, "-", "n.a.", "n.a.", "n.a.", "-"] 
    _android_version_list = [_android_version_info["title"], "", _android_version_info["db_key"], _android_version_info["version"],"n.a.","n.a.","n.a.","n.a.","-"]        
    _Crypto_Bosch_list = [_crypto_B_info["title"], "", _crypto_B_info["db_key"], _crypto_B_info["version"],"n.a.","n.a.", _crypto_b_diag_ver, _crypto_B_info["db_key_read"], "-"]
    _Crypto_Cust_list = [_crypto_C_info["title"], "", _crypto_C_info["db_key"], _crypto_C_info["version"],"n.a.","n.a.", _crypto_c_diag_ver, _crypto_C_info["db_key_read"], "-"]
    _oem_list = [_oem_info["title"], "", _oem_info["db_key"], _oem_info["version"],"n.a.","n.a.","n.a.","n.a.","-"]
    _cts_list = [_cts_title, "", _cts_info["db_key"], _cts_ver_short.upper(), _cts_filename, _cts_ver.upper() , _cts_diag_ver, _cts_info["db_key_read"], "-"]
    _ublox_list = [_ublox_info["title"], "", _ublox_info["db_key"], _ublox_sos_ver, _ublox_filename, "n.a.", _ublox_info["diag_ver"], _ublox_info["db_key_read"], "-"]
    _ublox_split_list = ["", "Checksum", _ublox_info["db_key_crc"], _ublox_info["product"], "-", "n.a.", "MANUALLY", _ublox_info["db_key_crc_read"], "-"]
    _SBR_list = [_SBR_info["product"], "", _SBR_info["db_key"], _sbr_version.strip(), "-", _SBR_info["label"], _sbr_diag_ver, _SBR_info["db_key_read"], "-"]
    _F4Linux_list = [_F4Linux_info["product"], "", _F4Linux_info["db_key"], _F4Linux_info["version"], "-", _F4Linux_info["label"], _F4Linux_diag_ver, _F4Linux_info["db_key_read"], "-"]
    
    if _add_sxm_section == "yes":
       _SXM_list = [ _sxm_info["key"], "", _sxm_info["db_key"], _sxm_info["version"], _sxm_filename, "n.a.", _sxm_diag_ver, _sxm_info["db_key_read"], "-"] 
    if _add_dtv_section == "yes":
        _JPN_PN = "yes"
        _DTV_FW_title = _DTV_SW_info["title"] + " [7] [8] [14]"
        _DTV_BSW_title = _DTV_BSW_info["title"] + " [7]"
        _DTV_HW_title = _DTV_HSW_info["title"] + " [7] [9]"
        _DTV_P1_title = _DTV_Par1_info["title"] + " [7] [14]"
        _DTV_P2_title = _DTV_Par2_info["title"] + " [7] [14]"
        _Lont_title = _Lont_info["title"] + " [11]"
        _dtv_fw_diag_ver = get_hex(_DTV_SW_info["version"])
        _dtv_bsw_diag_ver = get_hex(_DTV_BSW_info["version"])
        _dtv_hw_diag_ver = get_hex(_DTV_HSW_info["version"])
        _dtv_hw_ver = _DTV_HSW_info["version"]
        _dtv_bsw_ver = _DTV_BSW_info["version"]
        _dtv_hw_ver_label = _DTV_HSW_info["label"].replace("DTV_HW_VER",_dtv_hw_ver )
        _dtv_bsw_ver_label = _DTV_BSW_info["label"].replace("DTV_BOOT_VER",_dtv_bsw_ver )
        _dtv_fw_ver = _DTV_SW_info["version"]     
        _dtv_fw_ver_label = _DTV_SW_info["label"].replace("DTV_SW_VER",_dtv_fw_ver )
        _DTV_P1_md5 = _md5_dict["DTV_P1file_md5"] + " [3]"
        _DTV_P2_md5 = _md5_dict["DTV_P2file_md5"] + " [3]"
        _lont_ver = _Lont_info["version"]
        _lont_diag_ver = _lont_ver[1:3] + " "  + _lont_ver[3:5] + " " + _lont_ver[5:]
        
        _DTV_FW_list = [_DTV_FW_title, "", _DTV_SW_info["db_key"], _DTV_SW_info["version"], _dtv_swfilename, _dtv_fw_ver_label, _dtv_fw_diag_ver, _DTV_SW_info["db_key_read"],""] 
        #_DTV_FW_list = [_DTV_FW_title, "", _DTV_SW_info["db_key"], _DTV_SW_info["version"], "", _dtv_fw_ver_label, _dtv_fw_diag_ver, _DTV_SW_info["db_key_read"],""] 
        _DTV_BSW_list = [_DTV_BSW_title, "", _DTV_BSW_info["db_key"], _DTV_BSW_info["version"], "", _dtv_bsw_ver_label, _dtv_bsw_diag_ver, _DTV_BSW_info["db_key_read"],""]
        _DTV_HSW_list = [_DTV_HW_title, "", _DTV_HSW_info["db_key"], _DTV_HSW_info["version"], "", _dtv_hw_ver_label, _dtv_hw_diag_ver, _DTV_HSW_info["db_key_read"], ""]
        _DTV_P1_list = [_DTV_P1_title, "", _DTV_Par1_info["db_key"], _DTV_Par1_info["version"], _DTV_Par1_info["file"], _DTV_P1_md5, _DTV_Par1_info["diag_ver"], "-", "-"]
        _DTV_P2_list = [_DTV_P2_title, "", _DTV_Par2_info["db_key"], _DTV_Par2_info["version"], _DTV_Par2_info["file"], _DTV_P2_md5, _DTV_Par2_info["diag_ver"], "-", "-"]
        _Lont_list = [ _Lont_title, "", _Lont_info["db_key"], _Lont_info["version"], _lont_filename, "-", _lont_diag_ver, _Lont_info["db_key_read"], ""]
               
    else:
        _JPN_PN = "no"
            
            
    _aurix_TM_list = [ _tm_title, "", _tm_info["db_key"], _tm_info["version"], "n.a.", "n.a.", _tm_diag, _tm_info["db_key_read"], "-"]
    _aurix_list = [ _aurix_info["title"], "", _aurix_info["db_key"], _aurix_ver, "-", "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"] 
    _aurix_dnl_list = [ _aurix_info["title"], "", _aurix_info["db_key_aurix_dnl"], _aurix_ver, _aurix_dnl_filename, "n.a.", _aurix_diag_ver, _aurix_info["db_key_read"], "-"]
    _PTooling_list = [ _PT_title, "", _PTooling_info["db_key"], _PTooling_info["version"], _PTooling_info["file"], "n.a.", "n.a.", "n.a.", "-"]
    _PTooling_md5_list = [ _PT_title, "", _PTooling_info["db_key_crc"], _PTooling_info["version"], _PTooling_info["file"], _CFS_PTooling_md5, "n.a.", "-", "-"]
    _PTooling_BE_list = [ _PT_BE_title, "", _PTooling_info["db_key_ProdTool_BE"], _PTooling_info["BE_tool"], "-", "-", "-", "-", "-"]
    _CTS_LUN_list = [ _CTSPar_info["title"], "", _CTSPar_info["db_key_erase_lun"], "n.a.", "-", "n.a.", "n.a.", "n.a.","-"]
    _CTS_OP_list = [ _CTSPar_info["title"], "", _CTSPar_info["db_key_op"], _CTSPar_info["cts_param_op"], "-", "n.a.", "n.a.", "n.a.","-"]
    _CTS_RXL_list = [ _CTS_RXL_title, "", _CTSPar_info["db_key_rxl"], _CTSPar_info["cts_param_rxl"], "-", "n.a.", "n.a.", "n.a.","-"]
    _CTS_VA_list = [ _CTSPar_info["title"], "", _CTSPar_info["db_key_va"], _CTSPar_info["cts_param_va"], "-", "n.a.", "n.a.", "n.a.","-"]
    _CTS_PXL_list = [ _CTS_RXL_title, "", _CTSPar_info["db_key_pxl"], _CTSPar_info["cts_param_pxl"], "-", "n.a.", "n.a.", "n.a.","-"]
    _PD_Del_list = [ _PD_Del_title, "", _PDDel_info["db_key_crc"], _PDDel_info["version"], _PDDel_info["file"], _PD_config_md5, "n.a.","-","-"] 
    _CTS_Del_list = [ _CTS_Del_title, "", _cts_info["db_key_crc"], _cts_info["CTS_config_version"], _cts_config_filename, _cts_config_md5, "n.a.", "-", "-"]
    _UCB_Del_list = [ _UCB_Del_title, "", _UCB_Del_info["db_key_crc"], _UCB_Del_info["version"], _UCB_Del_info["file"], _UCB_Del_md5, "n.a.", "-", "-"]
            
                            
    _PD_target = _target + "\\PD_Configuration"
    _CTS_target = _target + "\\CTS_Configuration"
    _UCB_target = _target + "\\UCB_Configuration"
    _Support_target = _target + "\\Support_Files"
            
    _test_manager_ver = _test_manager.attrib["file"].split(".zip")[0]
    _com_1 = "[1] TM is not preprogrammed. TM is RAM loaded by CTS during operation. Used Version in CTS SW is " + _test_manager_ver + "(see SW Version and Version readable via production diagnosis).."
    _com_2 = "[2] Base path Production Tooling - \\\\bosch.com\\dfsrb\\DfsDE\\DIV\\CM\\AI\\SW_Production\\Nissan\\0060_CCS2_7515752366\\01_Tools\\production_tooling"
    _com_3 = "[3] MD5 checksum"
    _com_4 = "[4] PD Configuration can be found in: " + _PD_target
    _com_5 = "[5] CTS Configuration can be found in: " + _CTS_target
    _com_6 = "[6] UCB Configuration can be found in: " + _UCB_target
    _com_7 = "[7] DTV Versions see https://hi-dms.de.bosch.com/docushare/dsweb/Services/Document-1179901"
    _com_8 = "[8] To update the Firmware Version with the Version provided in the eMMC-Image (Firmware-Partition) use the routine to update all submodules: 31 01 11 10"
    _com_9 = "[9] Config Parameter and System (ASS_DTV_CONFIG) see https://hi-dms.de.bosch.com/docushare/dsweb/GetVersion/File-1034235/4"
    _com_10 = "[10] These keys should actually be in quotes when passed to the command line of the Production Download Tooling. This should be described in PAVE and implemented in UTS"
    _com_11 = "[11] Only valid for JPN-Variants."
    _com_12 = "[12] Relation between CFS CLI and SOS and PAVE and Station Setup is described in Docupedia: https://inside-docupedia.bosch.com/confluence/display/GG/4.2+CCS2+Project+-+CFS+CLI+to+PAVE+Description"
    _com_13 = "[13] According to CFS Documentation additional Qualcomm Packages are needed: " + _PT_tar_file 
    _com_14 = "[14] DTV FW and Parameter can be found in: " + _Support_target
           
    if _JPN_PN == "yes":
        _com_sec_list = [_com_1, _com_2, _com_3, _com_4, _com_5, _com_6, _com_7, _com_8, _com_9, _com_10, _com_11, _com_12, _com_13, _com_14]                
    else:
        _com_sec_list = _com_sec_list = [_com_1, _com_2, _com_3, _com_4, _com_5, _com_6, _com_10, _com_12, _com_13]
                
            
                       
    _list_of_sections_os = []
    
    
    if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]: 
        if _set_gasid_status == "yes" and _set_ngasid_status == "yes" :
            _list_items_sw =  [_SW_ginfo_list,_SW_nginfo_list] 
            _list_items_app = [ _app_gas_list,_app_ngas_list,_ublox_list,_ublox_split_list, _SBR_list, _F4Linux_list ]
        elif _set_gasid_status == "yes" and _set_ngasid_status == "no" :
            _list_items_sw =  [_SW_ginfo_list]
            _list_items_app = [ _app_gas_list,_ublox_list,_ublox_split_list, _SBR_list, _F4Linux_list ]
        elif _set_gasid_status == "no" and _set_ngasid_status == "yes" :
            _list_items_sw =  [_SW_nginfo_list]
            _list_items_app = [ _app_ngas_list,_ublox_list,_ublox_split_list, _SBR_list, _F4Linux_list ]
    else:
        _list_items_sw =  [_SW_info_list]
        _list_items_app = [_app_list,_ublox_list,_ublox_split_list, _SBR_list, _F4Linux_list ]
    _list_items_1 = [ _Overall_SW_info_list, _android_version_list, _Crypto_Bosch_list, _Crypto_Cust_list, _oem_list, _cts_list]
    _list_append_os( _list_items_sw )
    _list_append_os( _list_items_1 )
    _list_append_os( _list_items_app )
    
    if _add_sxm_section == "yes":
        _list_of_sections_os.append(_SXM_list)
            
    if _add_dtv_section == "yes":
        _list_items_2 = [ _DTV_FW_list, _DTV_BSW_list, _DTV_HSW_list, _DTV_P1_list, _DTV_P2_list ] 
        for _eachlist in _list_items_2:
            _list_of_sections_os.append(_eachlist)
            
    _list_items_3 = [ _aurix_TM_list, _aurix_list, _aurix_srec_list, _aurix_dnl_list]
    for _eachlist in _list_items_3:
        _list_of_sections_os.append(_eachlist)
                      
    if _add_dtv_section == "yes":
        _list_of_sections_os.append(_Lont_list)
    if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:  
        _list_items_4 = [ _ucb_lock_list, _sec_lock_list]
        for _eachlist in _list_items_4:
            _list_of_sections_os.append(_eachlist)
            
    _list_items_5 = [_PTooling_list, _PTooling_md5_list, _PTooling_BE_list, _CTS_LUN_list, _CTS_OP_list, _CTS_RXL_list, _CTS_VA_list, _CTS_PXL_list ]            
    for _eachlist in _list_items_5:
       _list_of_sections_os.append(_eachlist)
            
    if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:              
        _list_items_6 = [ _APP_LUN_list, _APP_OP_list, _APP_RXL_list, _APP_VA_list, _APP_PXL_list, _SEC_OP_list, _SEC_RXL_list, _SEC_VA_list, _SEC_PXL_list, _VIP_OP_list, _VIP_RXL_list, _VIP_VA_list, _VIP_PXL_list ]
 
        for _eachlist in _list_items_6:
            _list_of_sections_os.append(_eachlist)
                    
    _list_items_7 = [ _PD_Del_list, _CTS_Del_list, _UCB_Del_list ]
    for _eachlist in _list_items_7:
        _list_of_sections_os.append(_eachlist)
    
   
    for _eachlist in _list_of_sections_os:
        create_productInfo_subElmt_sections_os(_eachlist)
           
    #Comments section
    E8_os = ETOS.SubElement(E6_os,'SW_Comments')
    for _com in range(len(_com_sec_list)):
        Em = ETOS.SubElement(E8_os,'Comment')
        Em.text = _com_sec_list[_com ]
            
    root.append(E6_os)
   
    E9_os = ETOS.SubElement(root, "Set_Versions_Partnumbers")
    
    for _thislist in _Set_Info_PNInfoList_all:
        _x = 0   
        Ep = ETOS.SubElement(E9_os,'Set_Info_Col4')
        for _keys in Set_Info_col_list:
            Ep.attrib[_keys] = _thislist[_x]
            _x += 1
            
    E10_os = ETOS.SubElement(root, "Set_Versions_Products")    
    for _thislist in _Set_Info_PNInfoList_all_2:
        _x = 0   
        Eq = ETOS.SubElement(E10_os,'Set_Info_Col13')
        for _keys in Set_Info_col_list_2:
            Eq.attrib[_keys] = _thislist[_x]
            _x += 1
            
    ETOS.indent(tree, space='  ', level=1)
   
    tree.write(_overallsos_file, encoding='utf-8', xml_declaration=True) 
    _pdffile = _overallsos_file.split(".xml")[0] 

    _pdffile = _pdffile + ".pdf"
    cmd = ["perl", "xml2odxe.pl", "-xml", _overallsos_file ]
    subprocess.run(cmd)
    print ("\n odx-e file created for _overallsos_file")
    cmd2 = ["sh", fop_path, "-xsl", "SOS_pdf_ccs2.xsl", "-xml", _overallsos_file, "-pdf", _pdffile ]
    subprocess.run(cmd2)
    print ("\n pdf file created for _overallsos_file ")
 
def spl_call ():
    print("\n at spl_call ... ")
    _spl_template = "CCS2_SPL_TMPL_V01.00.xlsx"
    _Plant = "B6"
    _PN_cell = "B7"
    _setno_cell = "H4"
    _state_cell = "H5"
    _date_cell = "H6"
    _purpose_cell = "H7"
    _pcm_resp_cell = "B12"       #to be incorporated in master xml
    _pm_cell = "B13"
    _sw_pm_cell = "B14"
    _hw_pm_cell = "B15"
    _ecn_cell = "B16"
    _ecr_cell = "B17"
    _ucb_conf_cell = "B19"
    _cust_sw_cell = "B20"
    _overall_sw_cell = "B21"
    _sw_cont_cell = "B22"
    _doc_cell = "B23"
    _dnl_cell = "B24"
    _doc_no_cell = "H23"
    _dnl_no_cell = "H24" 
    
    _row_start = 25
    
    #_xls_File = _xls_File
    # wb1 =  openpyxl.load_workbook(_spl_template)     #load the work book
    # ws1 =  wb1['SPL'] 

    # wb2 = openpyxl.Workbook()
    # ws2 = wb2.active
    #ws2 = wb2["container"]
    
    _PN_list = [] 
    _SAP_PN_list = []
    _PN_type_list = []
    _cont_list = []
    
    
    for _i in _set_dict:
        r = _row_start
        _pn_f = _set_dict[_i]["part_number"]
        _PN_format = _pn_f[0:4] + "." +  _pn_f[4:7] + "." +  _pn_f[7: ]
        
        _PN =_set_dict[_i]["part_number"]
        _cont = _set_dict[_i]["product"]
        _set_cont = "SW_Container_"+ _cont
        _SAP_PN = _set_dict[_i]["document_number"]
        _PN_type = _set_dict[_i]["device_type"].lower()
        _set_no = "SET" + _set_dict[_i]["set_number"]
        _ufs_s = "UFS3.1_" + _set_dict[_i]["ufs_info"].upper() + " TMPL"  # form the search string
        #print ("\n ufs string: ",_ufs_s) 
        
        _aurix_s = _set_dict[_i]["aurix_spl_id"] + " TMPL"
        _ublox_s = _set_dict[_i]["gnss_spl_id"] + " TMPL"
        
        # if _ecr != "" and _ecr !="None":
            # _Chng_NR = _ecr
            # ws1[_ecr_cell ] = _Chng_NR
            # print ("_Chng_NR ecr : ", _Chng_NR )
        # elif _ecn.value !="" and _ecn.value != "None":
            # _Chng_NR = _ecn.value
            # ws1[_ecn_cell ] = _Chng_NR
            # print ("_Chng_NR ecn : ", _Chng_NR )
        #_Chng_NR = str(_Chng_NR) 
        
        if "Baseunit" in _tmpl_dir and "BU" in _sw_ver.split("_")[-1]:
            _spl_search_list = [_ufs_s, _aurix_s, _ublox_s ]   # base unit requires aurix srec and dnl  
            _BU_dnl_add_flag = "no"
        elif "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]: 
            _spl_search_list = [_ufs_s, _aurix_s, _ublox_s ] 
            _aurix_entry = "no"
            
            
        #print ("search strings : ", _spl_search_list )
        
        #form SPL name
        _spl_filename = _SAP_PN + "_" + _sw_ver + "_SPL-CCS2_" + _cont + "_" + _set_no + ".xlsx"  #versioning part can be checked laterw        
        _pdf_filename = _SAP_PN + "_" + _sw_ver + "_SPL-CCS2_" + _cont + "_" + _set_no + ".pdf"
        _splfile = _target + "\\" + _set_cont +  "\\Data_to_plant\\" + _spl_filename
        _pdffile = _target + "\\" + _set_cont +  "\\Data_to_plant\\" + _pdf_filename
        _sos_doc = _SAP_PN + "_" + _sw_ver + "_SOS-CCS2"
        copy(_spl_template, _splfile)
        wb1 =  openpyxl.load_workbook(_splfile, read_only=False, data_only=True, keep_links=False)     #load the work book
        ws1 =  wb1['SPL'] 
        _plants = _set_dict[_i]["plant"]
        _both = "9050 / 982W"
        _pgp = "9050"
        if _plants == _pgp:
            ws1[_Plant] = "RBMA (9050)"          
        elif _plants == _both:
            ws1[_Plant] = "RBMA (9050), AdP (982W)"
        #print ("plant : ", ws1[_Plant])
        
        ws1[_PN_cell] = _PN
        ws1[_setno_cell] = _set_dict[_i]["set_number"]
        ws1[_date_cell] = dt.now().strftime("%d/%m/%Y")
        ws1[_purpose_cell] = _purpose
        ws1[_pm_cell] = _pm.split(";")[0]
        ws1[_sw_pm_cell] = _sw_pm.split(";")[0]
        ws1[_hw_pm_cell] = _hw_pm.split(";")[0]
        ws1[_pcm_resp_cell] = _pcm.split(";")[0]
        ws1[_ucb_conf_cell] = "Refer the UCB Delivery State row in SOS"
        ws1[_cust_sw_cell] = "n.a."
        ws1[_overall_sw_cell] = "n.a."
        ws1[_sw_cont_cell] = _set_cont
        ws1[_doc_cell] = _sos_doc
        ws1[_dnl_cell] = "n.a."
        ws1[_doc_no_cell] = _SAP_PN[0:4] + "." +  _SAP_PN[4:7] + "." +  _SAP_PN[7:]
        ws1[_dnl_no_cell] = "n.a."
        
        if _ecr != "" and _ecr !="None":
            _Chng_NR = _ecr
            ws1[_ecr_cell ] = _Chng_NR
            ws1[_state_cell] = "withECR"
            #print ("_Chng_NR ecr : ", _Chng_NR )
        elif _ecn.value !="" and _ecn.value != None:
            _Chng_NR = _ecn.value 
            ws1[_ecn_cell ] = _Chng_NR[0:10] + " " + _Chng_NR[10:]
            #print ("_Chng_NR ecn : ", _Chng_NR )
            ws1[_state_cell] = "withECN"
        
        
        while ( r < ws1.max_row ):
            _spl_id = ws1.cell(row=r,column=1).value
            #print ("\n cell : ", _spl_id)
            _elec_pn_cell = ws1.cell(row = r + 3, column = 6) 
            _blank_devpn_cell = ws1.cell(row = r + 3, column = 2)
            _prog_devpn_cell = ws1.cell(row = r + 3, column = 8)
            _image_file_cell = ws1.cell(row = r + 3, column = 3)
            _image_pno_cell = ws1.cell(row = r + 3, column = 5)
            _remarks_cell = ws1.cell(row = r + 1, column = 7)
            if _spl_id in _spl_search_list:
                _elec_pn_cell.value = _set_dict[_i]["electrical_part_number"]
                
                _cts_ino = _set_dict[_i]["cts_image_number"]
                _cts_ino = _cts_ino[0:4] + "." +  _cts_ino[4:7] + "." +  _cts_ino[7:]
                
                _ublox_ino = _set_dict[_i]["ublox_image_number"]
                _ublox_ino = _ublox_ino[0:4] + "." +  _ublox_ino[4:7] + "." +  _ublox_ino[7:] 
                
                if "Baseunit" in _tmpl_dir and "BU" in _sw_ver.split("_")[-1]:  # this check is to avoid run time error of name not defined for _BU_dnl_add_flag
                    #print ("\n 2062 : ", _spl_id)
                    if _spl_id == _ufs_s:
                        _ufs_image_file = _set_dict[_i]["_cts_filename"]
                        _ufs_images_nos = _cts_ino 
                        _image_file_cell.value = _ufs_image_file
                        _image_pno_cell.value = _ufs_images_nos
                        _blank_devpn_cell.value = _set_dict[_i]["ufs_blank_device_number"]
                        r = r + 4
                        # print ("\n 2070 : ", _spl_id)
                        # print ("\n _aurix_s : ", _aurix_s)
                    elif _spl_id == _aurix_s and _BU_dnl_add_flag == "no":
                        _BU_dnl_add_flag = "yes"       #flag set, extra entry in spl for aurix only for BU, iterates twice as row not incremented
                        _image_file_cell.value = _set_dict[_i]["_aurix_srec_filename"]
                        _aurix_ino = _set_dict[_i]["aurix_srec_image_number"]
                        _aurix_ino = _aurix_ino[0:4] + "." +  _aurix_ino[4:7] + "." +  _aurix_ino[7:] 
                        _image_pno_cell.value = _aurix_ino
                        _prog_pno = _set_dict[_i]["scc_programmed_device_number"]
                        _blank_devpn_cell.value = _set_dict[_i]["scc_blank_device_number"]
                        _prog_devpn_cell.value = _prog_pno[0:4] + "." +  _prog_pno[4:7] + "." +  _prog_pno[7:]
                        r = r + 4
                        # print ("r at 2079 : ", r)
                        # print("_BU_dnl_add_flag:", _BU_dnl_add_flag)
                        # print ("\n 2079 : ", _spl_id)
                    elif _spl_id == _aurix_s and _BU_dnl_add_flag == "yes":
                        _image_file_cell.value = _set_dict[_i]["_aurix_dnl_filename"]
                        _blank_devpn_cell.value = "n.a."
                        _prog_devpn_cell.value = "n.a."
                        _aurix_ino = _set_dict[_i]["aurix_dnl_image_number"]
                        _aurix_ino = _aurix_ino[0:4] + "." +  _aurix_ino[4:7] + "." +  _aurix_ino[7:] 
                        _image_pno_cell.value = _aurix_ino
                        
                        #_remarks_cell.value = "To be programmed via CTS"
                        r = r + 4
                        # print ("r at 2089 : ", r)
                        # print("_BU_dnl_add_flag:", _BU_dnl_add_flag)
                        # print ("\n 2089 : ", _spl_id)
                        # exit()
                    elif _spl_id == _ublox_s:
                        _image_file_cell.value = _setfilenames_dict["_ublox_filename"]
                        _image_pno_cell.value = _ublox_ino
                        _prog_pno = _set_dict[_i]["ublox_programmed_device_number"]
                        _prog_devpn_cell.value = _prog_pno[0:4] + "." +  _prog_pno[4:7] + "." +  _prog_pno[7:]
                        _blank_devpn_cell.value = _set_dict[_i]["ublox_blank_device_number"]
                        r = r + 4
                        
                    # else:
                        # r = r + 4
                else:
                        #print("\n line 2123, else ")
                        _blank_devpn_cell.value = "n.a."
                         
                        _prog_devpn_cell.value = "n.a."
                         
                        if _spl_id == _ufs_s:
                            if _PN_type == "yes":
                                _app_filename = _set_dict[_i]["_app_gfilename"]
                            elif _PN_type == "non":
                                _app_filename = _set_dict[_i]["_app_ngfilename"]
                            _ufs_image_file = _set_dict[_i]["_cts_filename"] + "\n\n" + _app_filename
                            
                            _cts_ino = _set_dict[_i]["cts_image_number"]
                            _cts_ino = _cts_ino[0:4] + "." +  _cts_ino[4:7] + "." +  _cts_ino[7:]
                            _app_ino = _set_dict[_i]["app_image_number"]
                            _app_ino = _app_ino[0:4] + "." +  _app_ino[4:7] + "." +  _app_ino[7:]
                            _ufs_images_nos = _cts_ino + "\n" + _app_ino
                            
                            #_ufs_images_nos = _ufs_images_nos[0:4] + "." +  _ufs_images_nos[4:7] + "." +  _ufs_images_nos[7:] 
                            
                            #print ("ufs image number :", _ufs_images_nos )
                            _image_file_cell.value = _ufs_image_file
                            _image_pno_cell.value = _ufs_images_nos
                            _remarks_cell.value = "programmed to assembled device"
                            r = r + 4
                        elif _spl_id == _aurix_s:
                            if _aurix_entry == "no":
                                #print("\n line 2149, else, r ", r)
                                _aurix_entry = "yes"
                                _image_file_cell.value = _set_dict[_i]["_aurix_dnl_filename"]
                                _aurix_ino = _set_dict[_i]["aurix_dnl_image_number"]
                                _aurix_ino = _aurix_ino[0:4] + "." +  _aurix_ino[4:7] + "." +  _aurix_ino[7:] 
                                _image_pno_cell.value = _aurix_ino
                                _remarks_cell.value = "To be programmed via CTS " 
                                r = r + 4
                            else:
                                ws1.delete_rows(r,4)
                                #print("\n line 2161, else, r ", r)
                        elif _spl_id == _ublox_s:
                            _image_file_cell.value = _setfilenames_dict["_ublox_filename"]
                            _image_pno_cell.value = _ublox_ino
                            _remarks_cell.value = "To be programmed via CTS (functionality currently neither needed nor available)"
                            r = r + 4
            else:
                ws1.delete_rows(r,4)
                
        wb1.save(_splfile)  
        wb1.close()
        
        try:
            pythoncom.CoInitialize()
            #pdf generation from excel with PyWin32
            #excel = client.Dispatch("Excel.Application") 
            excel = win32com.client.Dispatch("Excel.Application")
            #excel.Visible =  True
            # Read Excel File 
            wb3 = excel.Workbooks.Open(_splfile) 
            ws3 = wb3.Worksheets[0]  
            # Convert into PDF File 
            ws3.ExportAsFixedFormat(0, _pdffile)
            #wb3.ExportAsFixedFormat(0, _pdffile)
            #xlsx2pdf(_splfile, _pdffile)
            #subprocess.run(['xlsx2pdf', _splfile, _pdffile])
            #wb3.ExportAsFixedFormat(0, _pdffile)
            wb3.Close(False) 
            excel.Quit()
  
        except Exception as e:
            print (f"\n Error occured during the generation of pdf : {e}")
  
    
def update_sap_xml(_update_type, _new_xml):
    global _PN_count
    global _ecr
    global _target
    global _target_prod
    #global _prod_path
    global _scope
    global _spl_tmpl
    global _sw_ver
    #global _sw_full_ver
    global _swid_g
    global _swid_ng
    global _tmpl_dir
    global _pd_dir
    global _pd_ver
    global _pd_file
    global _DTV_BSW_info
    global _DTV_HSW_info
    global _DTV_SW_info
    global _SBR_info
    global _F4Linux_info
    global _crypto_B_info
    global _crypto_C_info
    global _sxm_info
    
    global _cts_info
    global _cts_title
    global _cts_file 
    global _cts_src 
    global _cts_ver 
    global _cts_cpver 
    global _cts_conf_src 
    global _cts_cpfile 
    
    global _app_info
    global _app_ver
    global _app_gver
    global _app_ngver
    global _app_gfile
    global _app_ngfile
    global _app_gsrc 
    global _app_ngsrc
    
    global _aurix_info
    global _aurix_ver
    global _aurix_src
    global _aurix_srec
    global _aurix_dnl
    global _aurix_support

    global _ublox_info
    global _tm_info
    global _Lont_info
    global _PTooling_info
    global _CTSPar_info
    global _AppPar_info
    global _SecPar_info
    global _VIPPar_info
    global _PDDel_info
    global _UCB_Del_info
    global _UCB_Lock_info
    global _Sec_Lock_info
    global _DTV_Par1_info
    global _DTV_Par2_info
    global _set_dict
    global _device_type_dict
    global _prj
    global _prj_t
    global _Doc_Info_dict
    global _Project_Info_dict
    global _ecu_sw_key
    global _app_sw_key
    global _ecu_read_comm
    global _app_read_comm
    global _app_db_key
    global _sw_db_key
    global _test_manager
    global _pcm
    global _pm
    global _sw_pm 
    global _hw_pm
    global _purpose
    global _crypto_aurix_env
    global _aurixcryptosign
    global _oem_info
    global _android_version_info
        
    
    
    _device_type_dict = {}
    _PN_position = 4
        
    tree = ET.ElementTree()
    tree.parse(_new_xml)
    root = tree.getroot()
    
    
    #for i in root.findall("SET_Infos"):
        #count = i.count()
    #print ("count of part numbers : " )
    _PN_count = len(root.findall("SET_Infos"))
    if (_update_type == "sap_info"):
        print("\n at update_sap_xml with sap_info ")
        read_sap_sheet()     
        i=0
        _sap_sheet_tag_list = ["ecn_ecr", "container_name", "set_number", "document_number","part_number","ufs_blank_device_number", "scc_programmed_device_number", "scc_blank_device_number", "ublox_programmed_device_number", "ublox_blank_device_number","cts_image_number","app_image_number","aurix_srec_image_number", "aurix_dnl_image_number","ublox_image_number","sxm_image_number", "dtv_image_number", "lont_image_number", "electrical_part_number"]                              
        parent=ET.Element(root.tag)
        #child = ET.SubElement(parent, 'SET_Infos')
        for SET in root.iter("SET_Infos"):
            #print (SET.attrib )
            print (SET.attrib["part_number"] )
            _PN = SET.attrib["part_number"]
            _dev_type = SET.attrib["device_type"]
            #print ("device type: ", _dev_type)
            if _dev_type == "yes" :
                if "gas" in _device_type_dict:
                    pass    #one  entry is enough
                else:
                    _device_type_dict["gas"] = "yes"
            elif _dev_type == "non":
                if "non-gas" in _device_type_dict:
                    pass   #one  entry is enough
                else:
                    _device_type_dict["non-gas"] = "yes"
            elif _dev_type == "n.a":
                pass;
            
            #print(type(_PN))
            _PN = _PN.split(",")[0]
            _PN.strip()
            # print("_PN: ", _PN)
            # print("type:", type(_PN))
            _availability_status = "no"
            #print("\n part nu,ber list: ", _PN_list)
            for i in range(len(_PN_list)):
                _curr_PN = _PN_list[i][4]
                _curr_PN = str(_curr_PN)
                # print("current pn:", _curr_PN)
                # print("_curr_PN type:", type(_curr_PN))
                
                if ( _PN in _curr_PN ):
                    _availability_status = "yes"
                    for j in range(len(_sap_sheet_tag_list)):
                        _item_name = _sap_sheet_tag_list[j]
                        #child.set(_item_name,str(_PN_list[i][j])) # adds a new set_info section
                        SET.attrib[_item_name ] = str(_PN_list[i][j])#v# appends to the existing set_info   
                
                _i=+1
            if ( _availability_status == "no" ):
                print ("part number",_PN,"not available in the sap sheet")
            
        #root.append(child)     #not required as it adda an extra set_info tag to the xml
        _ecn_no = root.find("./Overall_Infos[@Col1='ecn']")
                    
        if _ecn.value != None:
             _ecn_no.attrib["Col2"] = _ecn.value
        else:   
            _ecn_no.attrib["Col2"] = ""
        
        ET.indent(tree,' ')            
        tree.write(_new_xml, encoding="utf-8")    
        
        _prj = root.find("./BoschPrjName")
        _prj = _prj.text
        

        for _element in root:
            if _element.tag == "DocInfo":
                #print ("DocInfo")
                _Doc_Info_dict = _element.attrib
            if _element.tag == "ProjectInfo":
                #print ("ProjectInfo")
                _Project_Info_dict = _element.attrib
                
        
        _app_db_key = root.find("./Overall_Infos[@Col1='db_key_app']")
        _app_sw_key = _app_db_key.attrib["Col2"]
        
        _app_read_key = root.find("./Overall_Infos[@Col1='db_read_app']")
        _app_read_comm = _app_read_key.attrib["Col2"]
        
        _sw_db_key = root.find("./Overall_Infos[@Col1='db_key_sw']")
        _ecu_sw_key = _sw_db_key.attrib["Col2"]
        
        _sw_read_key = root.find("./Overall_Infos[@Col1='db_read_sw']")
        _ecu_read_comm = _sw_read_key.attrib["Col2"]
        
        _ecr_no = root.find("./Overall_Infos[@Col1='ecr']")
        _ecr = _ecr_no.attrib["Col2"]
        
        _target_dir = root.find("./Overall_Infos[@Col1='target_dir']")
        _target = _target_dir.attrib["Col2"]
        #_target = r'\\bosch.com\dfsrb\DfsDE\DIV\CM\AI\SW_Production\Nissan\0047_RN_AIVI_7513750800\Test_folder'  #hardcoded for now
        
        _target_prod_dir = root.find("./Overall_Infos[@Col1='target_prod_dir']")
        _target_prod = _target_prod_dir.attrib["Col2"]
        #print("_target_prod: ", _target_prod)
        
        # _prod_src_dir = root.find("./Overall_Infos[@Col1='prod_src_dir']")
        # _prod_path = _prod_src_dir.attrib["Col2"]
        
        _scope_info = root.find("./Overall_Infos[@Col1='scope_info']")
        _scope = _scope_info.attrib["Col2"]
        
        _spl_tmpl_p2 = root.find("./Overall_Infos[@Col1='spl_tmpl_p2']")
        _spl_tmpl = _spl_tmpl_p2.attrib["Col2"]
        
        _sw_full_ver = root.find("./Overall_Infos[@Col1='sw_full_ver']")
        _sw_ver = _sw_full_ver.attrib["Col2"]
         
            
        _swid_gas = root.find("./Overall_Infos[@Col1='swid_gas']")
        _swid_g = _swid_gas.attrib["Col2"]
        
        _swid_nongas = root.find("./Overall_Infos[@Col1='swid_nongas']")
        _swid_ng = _swid_nongas.attrib["Col2"]
        
        _template_dir = root.find("./Overall_Infos[@Col1='template_dir']")
        _tmpl_dir = _template_dir.attrib["Col2"]
        
        _pcm_resp = root.find("./Overall_Infos[@Col1='del_resp']")
        _pcm = _pcm_resp.attrib["Col2"]
        
        _proj_man = root.find("./Overall_Infos[@Col1='PM']")
        _pm = _proj_man.attrib["Col2"]
        
        _sw_man = root.find("./Overall_Infos[@Col1='SW_PM']")
        _sw_pm = _sw_man.attrib["Col2"]
        
        _hw_man = root.find("./Overall_Infos[@Col1='HW_PM']")
        _hw_pm = _hw_man.attrib["Col2"]
        
        _rel_purpose = root.find("./Overall_Infos[@Col1='purpose']")
        _purpose = _rel_purpose.attrib["Col2"]
        
       
        _kds_pd_dir = root.find("./Overall_Infos[@Col1='kds_pd_dir']")
        _pd_dir = _kds_pd_dir.attrib["Col2"]
        _kds_pd_ver = root.find("./Overall_Infos[@Col1='kds_pd_ver']")
        _pd_ver = _kds_pd_ver.attrib["Col2"]
        _kds_pd_file = root.find("./Overall_Infos[@Col1='kds_pd_file']")
        _pd_file = _kds_pd_file.attrib["Col2"]
        
        _DTV_B_SW = root.find("./SW_Versions[@key='DTV_Boot_SW']")
        _DTV_BSW_info = _DTV_B_SW.attrib
        _DTV_H_SW = root.find("./SW_Versions[@key='DTV_HW']")
        _DTV_HSW_info = _DTV_H_SW.attrib
        _DTV_SW = root.find("./SW_Versions[@key='DTV_SW']")
        _DTV_SW_info = _DTV_SW.attrib
        _DTV_Par1 = root.find("./SW_Versions[@key='DTV_Parameter_1']")
        _DTV_Par1_info = _DTV_Par1.attrib
        _DTV_Par2 = root.find("./SW_Versions[@key='DTV_Parameter_2']")
        _DTV_Par2_info = _DTV_Par2.attrib
        #hardcoding DTV parameter file names for DT versions TH010792 and TH010793
        if _DTV_SW_info["version"] == "TH010793" or _DTV_SW_info["version"] == "TH010792" or _DTV_SW_info["version"] == "TH010790":
            _DTV_Par1_info["version"] = "TH1.0.0"
            _DTV_Par2_info["version"] = "TH1.0.1"
            _DTV_Par1_info["file"] = "e5caa2bb76_dtv_param1_TH1.0.0.zip"
            _DTV_Par2_info["file"] = "036c7b1271_dtv_param2_TH1.0.1.zip"
            
        
        #print (" dtv info:", _DTV_SW_info )
            
        _SBR = root.find("./SW_Versions[@key='SBR']")
        _SBR_info = _SBR.attrib
        
        _SBR.attrib["date"] = _SBR_info["label"].split(" ", 1)[1]
        
        _F4Linux = root.find("./SW_Versions[@key='Flash4Linux']")
        _F4Linux_info = _F4Linux.attrib
            
        _crypto_B = root.find("./SW_Versions[@key='cryptoenv_bosch']")
        _crypto_B_info = _crypto_B.attrib
            
        _crypto_C = root.find("./SW_Versions[@key='cryptoenv_cust']")
        _crypto_C_info = _crypto_C.attrib
        
        _oem = root.find("./SW_Versions[@key='oemcrypto']")
        _oem_info = _oem.attrib
        
        _android_version = root.find("./SW_Versions[@key='android_version']")
        _android_version_info = _android_version.attrib
            
        _sxm = root.find("./SW_Versions[@key='SXM']")
        _sxm_info = _sxm.attrib
      
        _cts = root.find("./SW_Versions[@key='cts']")
        _cts_info = _cts.attrib
        _cts_title = _cts.attrib['title']
        _cts_file = _cts_info['CTS_file']
        _cts_src = _cts_info['src_file']
        _cts_ver = _cts_info['version']   
        _cts_cpver = _cts_info['CTS_config_version']
        _cts_conf_src = _cts_info['CTS_config_src_file']
        _cts_cpfile = _cts_info['config_product']
       
        
        _app = root.find("./SW_Versions[@key='android_app_sw']")
        _app_info = _app.attrib
        _app_ver = _app_info['version']
        _app_gver = _app_info['app_sw_gas_version']
        _app_ngver = _app_info['app_sw_nongas_version']
        _app_gfile = _app_info['gas_file']
        _app_ngfile = _app_info['nongas_file']
        _app_gsrc = _app_info['gas_src_file']
        _app_ngsrc = _app_info['nongas_src_file']
        
        _aurix = root.find("./SW_Versions[@key='aurix_autosar_app_sw']")
        _aurix_info = _aurix.attrib
        _aurix_ver = _aurix_info['version']
        _aurix_src = _aurix_info['src_file']
        _aurix_srec = _aurix_info['srec_file']
        _aurix_dnl= _aurix_info['dnl_file']
        _aurix_support= _aurix_info['support_file']
        _aurixcryptosign = _aurix_info['crypto_env']
        
        _ublox = root.find("./SW_Versions[@key='ublox']")
        _ublox_info = _ublox.attrib
        
        _test_manager = root.find("./SW_Versions[@key='aurix_tm']")
        _tm_info = _test_manager.attrib
        
        _Lontium_SW = root.find("./SW_Versions[@key='Lontium_SW']")
        _Lont_info = _Lontium_SW.attrib
        
        _Prod_Tooling = root.find("./SW_Versions[@key='Prod_Tooling']")
        _PTooling_info = _Prod_Tooling.attrib
        
        _CTS_Parameter = root.find("./SW_Versions[@key='CTS_Parameter']")
        _CTSPar_info = _CTS_Parameter.attrib
        
        _App_Parameter = root.find("./SW_Versions[@key='Application_Parameter']")
        _AppPar_info = _App_Parameter.attrib
        
        _Sec_Parameter = root.find("./SW_Versions[@key='Secure_Boot_Parameter']")
        _SecPar_info = _Sec_Parameter.attrib
        
        _VIP_Parameter = root.find("./SW_Versions[@key='VIP_DNL_Parameter']")
        _VIPPar_info = _VIP_Parameter.attrib
        
        _PD_Del_state = root.find("./SW_Versions[@key='PD_Delivery_State']")
        _PDDel_info = _PD_Del_state.attrib
        
        _UCB_Del_state = root.find("./SW_Versions[@key='UCB_Delivery_State']")
        _UCB_Del_info = _UCB_Del_state.attrib
       
        _UCB_Lock_status = root.find("./SW_Versions[@key='UCB_Lock_Activate']")
        _UCB_Lock_info = _UCB_Lock_status.attrib
        
        _Sec_Lock_status = root.find("./SW_Versions[@key='Sec_Lock_Activate']")
        _Sec_Lock_info = _Sec_Lock_status.attrib
        
        
        
        #print ("fetching all the set infos to a dictionary")
        #fetching all the set infos to a dictionary
        #dict array - working , an extra set is added- check and fix -> because an extra setinfo section is there in master xml
        _set_dict ={}
        i=0
        for SET in root.iter("SET_Infos"):
            
            _SET_Infos_all = SET.attrib
            #print ("_SET_Infos_all", _SET_Infos_all)
            _set_dict[i] = _SET_Infos_all
            i+=1
            
            #print (SET.attrib)
        #print("**************************display set_dict******************************")
        #print (_set_dict)
        #dictionary will not work as update dictionary will replace the values, cos it will not allow same keys with mutiple values
        #exit()
       
    elif (_update_type == "md5_info"):
        #print("at update_sap_xml with md5_info ")
        
        for SET in root.iter("SET_Infos"):
            _md5_temp_dict = _md5_dict.copy()           
            _PN = SET.attrib['part_number']
            _d_type = SET.attrib["device_type"]
            # print ("part number : ", _PN) 
            # print ("device_type : ", _d_type )
            
            _PN = _PN.split(",")[0]
            _PN.strip()
        #please re-visit this area
            if _d_type == "yes" and "appsw_nongas_file_md5" in _md5_temp_dict.keys():    
                _md5_temp_dict.pop("appsw_nongas_file_md5")
            elif _d_type == "non" and "appsw_gas_file_md5" in _md5_temp_dict.keys():
                _md5_temp_dict.pop("appsw_gas_file_md5")
            #print ("temp dict updated : ", _md5_temp_dict )
        
            for _md5_bin, _md5_val in _md5_temp_dict.items():
                SET.attrib[_md5_bin] = _md5_val
        tree.write(_new_xml, encoding="utf-8")
        
    elif (_update_type == "verify_md5_info"):
        #print("at update_sap_xml with md5_info ")
        
        for SET in root.iter("SET_Infos"):
            _md5_temp_dict = _md5_dest_dict.copy()           
            _PN = SET.attrib['part_number']
            _d_type = SET.attrib["device_type"]
            # print ("part number : ", _PN) 
            # print ("device_type : ", _d_type )
            
            _PN = _PN.split(",")[0]
            _PN.strip()
        #please re-visit this area
            if _d_type == "yes" and "appsw_nongas_file_md5" in _md5_temp_dict.keys():    
                _md5_temp_dict.pop("appsw_nongas_file_md5")
            elif _d_type == "non" and "appsw_gas_file_md5" in _md5_temp_dict.keys():
                _md5_temp_dict.pop("appsw_gas_file_md5")
            #print ("temp dict updated : ", _md5_temp_dict )
        
            for _md5_bin, _md5_val in _md5_temp_dict.items():
                SET.attrib[_md5_bin] = _md5_val
        
        tree.write(_new_xml, encoding="utf-8")
        
    elif (_update_type == "filename_info"):
        #print("at update_sap_xml with set filename_info ")
        for SET in root.iter("SET_Infos"):
            _filename_temp_dict = _setfilenames_dict.copy()           
            _PN = SET.attrib['part_number']
            _d_type = SET.attrib["device_type"]
            # print ("part number : ", _PN) 
            # print ("device_type : ", _d_type )
            
            _PN = _PN.split(",")[0]
            _PN.strip()
            if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:
                #if _d_type == "yes" :
                if _d_type == "yes"  and "_app_ngfilename" in _filename_temp_dict:
                    _filename_temp_dict.pop("_app_ngfilename")
                elif _d_type == "non"and "_app_gfilename" in _filename_temp_dict:
                    _filename_temp_dict.pop("_app_gfilename")
                #print ("temp dict updated : ", _md5_temp_dict )
            
            if SET.attrib['sxm'] != "yes" and "_sxm_filename" in _filename_temp_dict:
                    _filename_temp_dict.pop("_sxm_filename")
            
            if SET.attrib['dtv'] != "yes":
                if "_dtv_swfilename" in _filename_temp_dict:
                    _filename_temp_dict.pop("_dtv_swfilename")
                if "_dtv_p1filename"in _filename_temp_dict:
                    _filename_temp_dict.pop("_dtv_p1filename")
                if "_dtv_p2filename"in _filename_temp_dict:
                    _filename_temp_dict.pop("_dtv_p2filename")
                if "_lont_filename" in _filename_temp_dict:
                    _filename_temp_dict.pop("_lont_filename")
                    
                               
            
            
            for _artifact, _name in _filename_temp_dict.items():
                SET.attrib[_artifact] = _name
            
        tree.write(_new_xml, encoding="utf-8") 
        i=0
        #_set_dict is to be updated again here to add in the filenames and the md5sums to the dictionary
        for SET in root.iter("SET_Infos"):
            
            _SET_Infos_all = SET.attrib
            #print ("_SET_Infos_all", _SET_Infos_all)
            _set_dict[i] = _SET_Infos_all
            i+=1
           
        
    
    
    
if __name__ == '__main__':
    parser = optparse.OptionParser()
    parser.add_option('-x', '--xml', dest='xml', default=None,
                      help='Master xml')
    parser.add_option('-t', '--template', dest='template', default=None,
                      help='SAP Request sheet')
                     
    (options, args) = parser.parse_args()
    if not options.xml:
        #input_masterxml = input("\nPlease enter masterxml filename\n")
        print ("\n Please enter the master xml name as input ")
        print ("\n please run the script in the form : ")
        print ("\t create_prod_ccs2.py -x <xml_name> -t <sap_request_sheet_name> ")
        sys.exit(0) 
    else:
        input_masterxml = options.xml
        _xml_Filename = input_masterxml
        print ("\n Master xml ", _xml_Filename," will be used ")
		
	
    if not options.template:
        #input_template = input("\nPlease enter SAP request sheet name\n")
        print ("\n Please enter the SAP Request sheet name as input ")
        print ("\n please run the script in the form : ")
        print ("\t create_prod_ccs2.py -x <xml_name> -t <sap_request_sheet_name> ")
        sys.exit(0)
        
    else:
        input_template = options.template
        #print ("\n SAP request sheet :", input_template )
        _xls_Filename = input_template
        print ("\n Sap Request sheet ", _xls_Filename ," will be used ")
	

    
    _curr_Dir = os.getcwd()
    _xls_File = os.path.join(_curr_Dir,_xls_Filename)   
    _xml_File = os.path.join(_curr_Dir,_xml_Filename) 
    
    _xml_base = _xml_Filename.split(".xml")[0]
    _new_xml_basename = "SAP_"+_xml_base
    _new_xml_name = "SAP_"+_xml_base + ".xml"
    _new_xml = os.path.join(_curr_Dir,_new_xml_name)
    # print ("\n _new_xml_name : ",_new_xml_name)
    # print ("\n _new_xml : ",_new_xml)
    
    # print ("\n master xml new_xml_name: ",_new_xml_name)
    _new_xml = os.path.join(_curr_Dir,_new_xml_name)
    print ("\n _new_xml : ",_new_xml)
    if os.path.exists(_new_xml):
        _suffix = dt.now().strftime("%Y%m%d_%H%M%S") #creates suffix with current date and exact time
        #print("\n date suffix: ", _suffix)
        _file_rename = _new_xml_basename + "." + _suffix + ".xml"
        _xml_rename = os.path.join(_curr_Dir,_file_rename)
        os.rename(_new_xml, _xml_rename) 
        # print("\n renamed the existing xml .. ")
            
    copy(_xml_File, _new_xml) 
    
    update_sap_xml("sap_info", _new_xml)
    print ("\n Please check and download the PD Config file from sharepoint to the corresponding folder in ", _target_prod )
    print ("\n \t PD config file : ", _pd_file )
    print ("\n please check if the production tooling CFS file is available in the 01_Tools path ")
    print ("\n \t Production Tooling (CFS) file: ", _PTooling_info["file"] )
    #print ("\n \t Please confirn if these files are available ")
    _userinput = input("\n Please confirm if these files are available : Y/N \t").upper()
    
    if _userinput == "Y":
        print ("\n")
        print("****************************************************************************************************************************************")
        if not os.path.exists(_target):
            print("\n The production folder", _target, " will be created" )
        else:
            print("\n The production folder : ", _target, "exists already" )
            print("\n Please note that the update will replace the files ")
        _userinput = input("\n Please confirm to proceed : Y/N \t").upper()
        if _userinput == "Y":
            print("****************************************************************************************************************************************")
            create_prod_folders(_new_xml) 
            file_url = "url"
            filename ="filename"
            #download_from_sharepoint(file_url, filename)
            check_artifacts() 
            set_filenames(_new_xml)
            copy_files(_new_xml)
            sos_calls()
            spl_call()
            print("****************************************************************************************************************************************")
            print ("\ncreate production folder is completed  ")
            print("****************************************************************************************************************************************")
        else:
            sys.exit(0)
    else:
        sys.exit(0)
	

        
        