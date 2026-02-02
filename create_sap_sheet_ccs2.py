#################################################################################################################
#
# FILE:         create_sap_sheet_ccs2.py
# DESCRIPTION:  This script can be used to create the SAP request sheet for CCS2 project
# USAGE:        see help_text
# PREREQUISITE: The files CCS2_SPL_TMPL_V01.00.xlsx, SAP_Request_template.xlsx has to be in the same directory as this script
# HISTORY:
# Date         | Author          		| Modification
# 29.08.2024   | Nisharani C  			| Initial version
# 18.09.2024   | Nisharani C  			| HW Mapping sheet update
# 24.10.2024   | Nisharani C            | Corrected the blank device printing in the HW sheet
# 24.06.2025   | Abinaya M	  			| Sync with latest changes in Set definition tmpl (to add Supplier feed column and change from col name ublox to gnss)
# 25.06.2025   | Nisharani C            | Imagename changes for SXM,DTV,Lontium
# 11.07.2025   | Nisharani C            | Bugfix for reuse check - when sxm, dtv, lont versions are blank - when SAP sheet is created from set-def without any versions
# 20.08.2025   | Nisharani C            | Sync with latest changes in Set definition tmpl (to add Plant column) 
# 06.10.2025   | Nisharani C			| Sync with latest changes in Set definition tmpl (to add BU column)
# 08.10.2025   | Nisharani C			| Sync with latest changes - to include app_ID column in prod deliveries sheet
# 25.11.2025   | Nisharani C            | Sync with latest changes - to consider Aurix_Sub_Path and app_filename for re-usage, and bug fixes on sxm, dtv and lontium re-usage
#################################################################################################################

import os
import sys
import re
import argparse
#import optparse
import subprocess
import pandas as pd # for dataframes
import xml.etree.ElementTree as ET
import xlwings as xw
# import xml.etree.ElementTree as ET1
# import xml.etree.ElementTree as ETS
# import xml.etree.ElementTree as ETOS
 
import openpyxl 
import numpy as np

import shutil
from datetime import datetime as dt
from shutil import copy
from artifactory import ArtifactoryPath
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from win32com import client         #for excel to pdf generation during spl call


class _SAP_request:


    def __init__(self):
        self._del_sheet = ""
        self._set_def_no = ""
        self._cts_ver = ""
        self._app_gver = ""
        self._app_ngver = ""
        self._aurix_ver = ""
        self._aurix_sub_path = ""
        self._aurixcryptosign = ""
        self._ublox_ver = ""
        #self._xml = _xml_File
        #self._set_info_dict = ""
        self._set_info_list = ""
        self._rel = ""
        self._PN_len = ""
        self._rel_type = ""
        self._sap_sheet = "SAP_Request_template.xlsx"
        self._sap_request_sheet = ""
        self._sw_ver = ""
       
    def process_xml(self, _xml_File):
        print ("\n initializing as xml ")
        self._xml = _xml_File
    def process_xls(self, _xls_File):
        print ("\n initializing as xls ")
        self._xls = _xls_File
        
    def parse_master_xml( self ): 
        _xml  = self._xml
        print ("\n at parse_master_xml with xml : ", self._xml)
        tree = ET.ElementTree()
        tree.parse(_xml)
        root = tree.getroot()
        
        for _element in root:
            if _element.tag == "DocInfo":
                #print ("DocInfo")
                _Doc_Info_dict = _element.attrib
                self._set_def_no = _element.attrib["Doc_SetDef"]
        _target_prod_dir = root.find("./Overall_Infos[@Col1='target_prod_dir']")
        _target_prod = _target_prod_dir.attrib["Col2"]
        self._del_sheet =  _target_prod + "\\Reference_sheets\\ProdDeliveries_Overview_CCS2.xlsx"
        
        _ecr_no = root.find("./Overall_Infos[@Col1='ecr']")
        _ecr = _ecr_no.attrib["Col2"]
        
        _ecn_no = root.find("./Overall_Infos[@Col1='ecn']")
        _ecn = _ecn_no.attrib["Col2"]
        
        if _ecr != "" and _ecr !="None":
            _Chng_NR = _ecr
            _Chng_NR_type = "withECR"
        elif _ecn != "" and _ecn !="None":
            _Chng_NR = ""
            _Chng_NR_type = "withECN"
        self._Chng_NR = _Chng_NR
        self._Chng_NR_type = _Chng_NR_type
        
        _template_dir = root.find("./Overall_Infos[@Col1='template_dir']")
        _tmpl_dir = _template_dir.attrib["Col2"]
        
        _sw_full_ver = root.find("./Overall_Infos[@Col1='sw_full_ver']")
        _sw_ver = _sw_full_ver.attrib["Col2"]
        self._sw_ver = _sw_ver
        
        if "Endunit" in _tmpl_dir and "EU" in _sw_ver.split("_")[-1]:
            self._rel = "EU"
        elif "Baseunit" in _tmpl_dir and "BU" in _sw_ver.split("_")[-1]:
            self._rel = "BU"
         
        _cts = root.find("./SW_Versions[@key='cts']")
        _cts_info = _cts.attrib
        _cts_ver = _cts_info['version']
        _cts_ver = _cts_ver.lower()  
        _cts_ver = (_cts_ver.split("cts_")[1]).upper()
    
        self._cts_ver = _cts_ver.strip()
            
        _app = root.find("./SW_Versions[@key='android_app_sw']")
        _app_info = _app.attrib
        _app_ver = _app_info['version']
        self._app_gver = _app_info['app_sw_gas_version']
        self._app_ngver = _app_info['app_sw_nongas_version']
        self._app_gfile = _app_info['gas_file']
        self._app_ngfile = _app_info['nongas_file']
            
        _aurix = root.find("./SW_Versions[@key='aurix_autosar_app_sw']")
        _aurix_info = _aurix.attrib
        self._aurix_ver = _aurix_info['version']
        self._aurix_sub_path = _aurix_info['_sub_path']
        #self._aurixcryptosign = _aurix_info['crypto_env']
            
        _ublox = root.find("./SW_Versions[@key='ublox']")
        _ublox_info = _ublox.attrib
        self._ublox_ver = _ublox_info['version'].strip()
        
        _DTV_SW = root.find("./SW_Versions[@key='DTV_SW']")
        _DTV_SW_info = _DTV_SW.attrib
        self._dtv_ver = _DTV_SW_info['version'].strip()
        
        _sxm = root.find("./SW_Versions[@key='SXM']")
        _sxm_info = _sxm.attrib
        self._sxm_ver = _sxm_info['version'].strip()
        self._sxm_fwver = _sxm_info['fw_version'].strip()
        
        _Lontium_SW = root.find("./SW_Versions[@key='Lontium_SW']")
        _Lont_info = _Lontium_SW.attrib
        self._lont_ver = _Lont_info['version'].strip()
        
        _i = 0 
        #_set_info_dict = {}
        _set_info_list = []
        _rel_type = "not_set"
        _PN_len = ""
        for SET in root.iter("SET_Infos"):
            #print (SET.attrib )
            #print (SET.attrib["part_number"] )
            _PN = SET.attrib["part_number"]
            if len(_PN) == 10:
                #print("here at 10")
                _PN_len = "sop"
            elif len(_PN) == 13:
                #print("here at 13")
                _PN_len = "tsb"
            if _rel_type == "not_set" :    
                _rel_type = _PN_len
            elif _PN_len != _rel_type:
                print ("\n Please note : PNs with different length is not possible within a release ")
                print ("\n exiting ... ")
                sys.exit(0)
            #print ("\n release type: ",  _rel_type )   
            _dev_type = SET.attrib["device_type"]
            _set_info_list.append(SET.attrib)
            #_set_info_dict[_i] = SET.attrib
            # _i +=  1
         
            # print ("set info list : ", _set_info_list )

        self._rel_type = _rel_type
        #print ("set info list: ", _set_info_list)
        self._set_info_list =_set_info_list
        # self._set_info_dict = _set_info_dict
    
    def get_blank_dev_numbers(self, _spl_temp):
        _spl_template = _spl_temp
        #_spl_template = "CCS2_SPL_TMPL_V01.00.xlsx"
        print ("\nat get_blank_dev_numbers .." )
        print ("Using ", _spl_template, "to get the blank device numbers .. ")
        
        _i = 0 
        _pn_splid_dict = {}
        _blank_dev_dict = {}
        _blank_dev_dict_pn = {}
        for _set in self._set_info_list:
            _ufs_s = "UFS3.1_" + _set["ufs_info"].upper() + " TMPL"  # form the search string
            _aurix_s = _set["aurix_spl_id"] + " TMPL"
            _ublox_s = _set["gnss_spl_id"] + " TMPL"            
            _pn_splid_list = [_ufs_s, _aurix_s, _ublox_s ]
            
            _row_start = 25
            r = _row_start
            wb1 =  openpyxl.load_workbook(_spl_template)     #load the work book
            ws1 =  wb1['SPL'] 
            while ( r < ws1.max_row ):
                _spl_id = ws1.cell(row=r,column=1).value 
                _blank_devpn_cell = ws1.cell(row = r + 3, column = 2)
                if _spl_id in _pn_splid_list: 
                    if _spl_id == _ufs_s:
                        _blank_dev_dict["ufs_blank_device"] = _blank_devpn_cell.value
                    elif _spl_id == _aurix_s:
                        _blank_dev_dict["aurix_blank_device"] = _blank_devpn_cell.value
                    elif _spl_id == _ublox_s:
                        _blank_dev_dict["ublox_blank_device"] = _blank_devpn_cell.value
                r = r + 4
            _set.update(_blank_dev_dict)
            #print ("\n set now : ", _set )
    
    def read_SetDefinition_sheet(self):
        print ("\nat read_SetDefinition_sheet ")
        
        _set_def_cell = "B1"        
        _rel_version_cell = "B3"
        _ecn_ecr_cell = "B4"
        _ecr_cell = "C4"
        _CTS_ver_cell = "B8"
        _AppSw_ver_cell = "B9"
        _GAS_ver_cell = "B10"
        _NonGAS_ver_cell = "B11"
        _Aurix_ver_cell = "B12"
        _Aurix_sub_path_cell = "C12"
        _GNSS_cell = "B14"
        _SXM_cell = "B15"
        _SXM_fw_cell = "C15"
        _DTV_cell = "B16"
        _Lont_cell = "B17"
       
        wb2 =  openpyxl.load_workbook(self._xls, data_only=True)     #load the work book
        ws5 =  wb2['Set Definition']
        
        _target_prod = r'\\bosch.com\dfsrb\DfsDE\DIV\CM\AI\SW_Production\Nissan\0060_CCS2_7515752366\00_SW\_PCM_internal'
        self._del_sheet =   _target_prod + "\\Reference_sheets\\ProdDeliveries_Overview_CCS2.xlsx"
        
        self._set_def_no = ws5[_set_def_cell].value
        #print ("\n self._set_def_no: ", self._set_def_no )
        _Chng_NR_type = ws5[_ecn_ecr_cell].value
        if _Chng_NR_type == "ECR":
            _Chng_NR_type = "withECR"
            _Chng_NR = ws5[_ecr_cell].value
        elif _Chng_NR_type == "ECN":
            _Chng_NR_type = "withECN"
            _Chng_NR = ""
        else:
            print("\n please select ECN or ECR .. ")
            sys.exit(0)
         
        self._Chng_NR = _Chng_NR
        self._Chng_NR_type = _Chng_NR_type
        # print ("\n self._Chng_NR_type: ", self._Chng_NR_type )
        # print ("\n self._Chng_NR: ", self._Chng_NR )
        
        self._sw_ver = ws5[_rel_version_cell].value
        #print ("\n self._sw_ver: ", self._sw_ver )
        
        if "EU" in self._sw_ver.split("_")[-1] and "EU" in self._set_def_no.split("_")[-1]:
            self._rel = "EU"
        elif "BU" in self._sw_ver.split("_")[-1] and "BU" in self._set_def_no.split("_")[-1]:
            self._rel = "BU"
        #print ("\n self._sw_ver: ", self._sw_ver )
       
        _cts_ver = ws5[_CTS_ver_cell].value
        #print("\n cts version: ", _cts_ver)
        if _cts_ver != None:
            _cts_ver = (_cts_ver.split("cts_")[1]).upper()
        else:
            _cts_ver = ""
        self._cts_ver = _cts_ver
        #print ("\n self._cts_ver: ", self._cts_ver )
        
        _app_ver = ws5[_AppSw_ver_cell].value
        self._app_ver = _app_ver
        self._app_gver = ws5[_GAS_ver_cell].value
        self._app_ngver = ws5[_NonGAS_ver_cell].value
        # print ("\n self._app_ver: ", self._app_ver )
        # print ("\n self._app_gver: ", self._app_gver )
        # print ("\n self._app_ngver: ", self._app_ngver )
        
        _aurix_ver = ws5[_Aurix_ver_cell].value
        self._aurix_ver = _aurix_ver
        self._aurix_sub_path = ws5[_Aurix_sub_path_cell].value
        #print ("\n self._aurix_ver: ", self._aurix_ver )
        
        _ublox_ver = ws5[_GNSS_cell].value
        if _ublox_ver != None:
            self._ublox_ver = str(_ublox_ver).strip()
            #self._ublox_ver = _ublox_ver.strip()
        else:
            _ublox_ver = ""
            self._ublox_ver = _ublox_ver
        
        _sxm_ver = ws5[_SXM_cell].value
        if _sxm_ver != None:
            self._sxm_ver = str(_sxm_ver).strip()
        else:
            _sxm_ver = ""
            self._sxm_ver = _sxm_ver
            
        _sxm_fw_ver = ws5[_SXM_fw_cell].value
        if _sxm_fw_ver != None:
            self._sxm_fwver = str(_sxm_fw_ver).strip()
        else:
            _sxm_fwver = ""
            self._sxm_fwver = _sxm_fwver
            
        _dtv_ver = ws5[_DTV_cell].value
        if _dtv_ver != None:
            self._dtv_ver = str(_dtv_ver).strip()
        else:
            _dtv_ver = ""
            self._dtv_ver = _dtv_ver

        _lont_ver = ws5[_Lont_cell].value
        if _lont_ver != None:
            self._lont_ver = str(_lont_ver).strip()
        else:
            _lont_ver = ""
            self._lont_ver = _lont_ver        
        
        # print ("\n self._sxm_ver: ", self._sxm_ver )
        # print ("\n self._dtv_ver: ", self._dtv_ver )        
        # print ("\n self._lont_ver: ", self._lont_ver )
        
              
        _PN_count = 0
        _PN_details = []  
        df1 = pd.read_excel(self._xls, header=None, sheet_name='Set Definition', skiprows=22)
        df1 = df1.fillna('')  #replace the "NAN"values from columns ( empty string )
        # Convert DataFrame to a list using the 'values' attribute
        df1_list = df1.values.tolist()
        #print(df1_list) 
        _Set_keys_list = ['part_number','product','scope','ufs_info','ufs_spl_id','aurix_info','aurix_spl_id','gnss_prod','gnss_spl_id','fpga_prod','sxm','dtv','adr_type','device_type','supplier feed','plant','bu','variant']
        _set_info_list = []
        _rel_type = "not_set"
        _PN_len = ""
        for i in range(len(df1_list)):
            _del_req = df1_list[i][0]
            # print ("\n _del_required value ",_del_req)
            #print ("i value ", i )
            if _del_req in ["X","x"]:
                #k = 0
                print ("\n _del_required for PN ",df1_list[i][1] )
                df1_list[i].pop(0)
                #print(df1_list[i])
                #print("\n")
                _PN = str(df1_list[i][0]).strip()
                # print("\n PN: ", _PN )
                # print("\n length of PN:",len(_PN))
                if len(_PN) == 10:
                    #print("here at 10")
                    _PN_len = "sop"
                elif len(_PN) == 13:
                    #print("here at 13")
                    _PN_len = "tsb"
                if _rel_type == "not_set" :    
                    _rel_type = _PN_len
                elif _PN_len != _rel_type:
                    print ("\n Please note : PNs with different length is not possible within a release ")
                    print ("\n exiting ... ")
                    sys.exit(0)
                #print ("\n release type: ",  _rel_type ) 
                
                _col_count = len(df1_list[i])
                #print ("\n _col_count:", _col_count)
                _Set_info_dict = {};
                #print ("k value ", k )
                for j in range(0,_col_count):
                   _item_name = _Set_keys_list[j]
                   _Set_info_dict[_item_name ] = df1_list[i][j]
                #print("\n _Set_info_dict:", _Set_info_dict)    
                _set_info_list.append(_Set_info_dict)
                #print ("\n _set_info_list:", _set_info_list)
                #k = k + 1
        self._set_info_list =_set_info_list
        self._rel_type = _rel_type
        
        
    def check_del_sheet(self):
        print ("\nat check_del_sheet ")
        #the below variables resembles that from update_prod_deliveries script, but its col values are just 1 value lesser here as we use list here, so it starts from 0
        _row_start = 3
        _col_start = 2
        _rel_name_col = 1
        _ecn_ecr_col = 2
        _pn_col = 3
        _ufs_BNo_col = 4
        _cts_ver_col = 5
        _cts_DNo_col = 6
        _cts_file_col = 7
        _app_id_col = 8
        _app_ver_col = 9
        _app_DNo_col = 10
        _app_file_col = 11
        _aurix_BNo_col = 12
        _aurix_PNo_col = 13
        _aurix_ver_col = 14
        _aurix_sub_path = 15
        _aurix_srec_DNo_col = 16
        _aurix_srec_file_col = 17
        _aurix_dnl_DNo_col = 18
        _aurix_dnl_file_col = 19
        _ublox_BNo_col =  20
        _ublox_PNo_col = 21
        _ublox_ver_col = 22
        _ublox_DNo_col = 23
        _ublox_file_col = 24
        _sxm_ver_col = 25
        _sxm_fwver_col = 26         #6thJune - add this check also for reusage - add 
        _sxm_DNo_col = 27
        _sxm_file_col = 28
        _dtv_ver_col = 29
        _dtv_DNo_col = 30
        _dtv_file_col = 31
        _lont_ver_col =32
        _lont_DNo_col = 33
        _lont_file_col = 34
        _set_def_col = 35
        _cts_ino_reuse_status = "no"
        _app_igno_reuse_status = "no"
        _app_ingno_reuse_status = "no"
        _aurix_isrecno_reuse_status = "no"
        _aurix_idnlno_reuse_status = "no"
        _ublox_pno_reuse_status = "no"
        _ublox_ino_reuse_status = "no"
        _sxm_ino_reuse_status = "no"
        _dtv_ino_reuse_status = "no"
        _lont_ino_reuse_status = "no"
            
        #saying openpyxl to ignore non-data cells, setting data_only=True
        wb1 =  openpyxl.load_workbook(self._del_sheet, data_only=True)     #load the work book , data_only flag helps to get the value instead of the formaula from a cell in the excel
        ws1 =  wb1['Sheet1']
        _r_start = ws1.max_row
        #_col_list = [_rel_name_col, _ecn_ecr_col, _pn_col, _ufs_BNo_col, _cts_ver_col, _cts_DNo_col, _cts_file_col, _app_ver_col, _app_DNo_col, _app_file_col, _scc_BNo_col, _scc_PNo_col, _aurix_ver_col, _aurix_srec_DNo_col, _aurix_srec_file_col, _aurix_dnl_DNo_col, _aurix_dnl_file_col, _ublox_BNo_col, _ublox_PNo_col, _ublox_ver_col, _ublox_DNo_col, _ublox_file_col, _set_def_col]
        df = pd.read_excel(self._del_sheet, sheet_name=0, skiprows=4)
        #print ("\n data frame :",df)
        df = df.fillna('')  #replace the "NAN"values from columns ( empty string )
        # Convert DataFrame to a list using the 'values' attribute
        df_list = df.values.tolist()
        df_list.reverse()       #.reverse() doesn’t return reversed list, the list is reversed in place. 
        #print(df_list)
        for _set in self._set_info_list:
            _reuse_dict = {}
            _rel_type_each_row = ""
            #print ("\n set for PN :", _set["part_number"])
            for _each_row in range(len(df_list)):
                _rel_full_name = df_list[_each_row][_rel_name_col]
                _rel_each_row = df_list[_each_row][_rel_name_col].split("_")[-1].strip()
                _set_def_each_row = df_list[_each_row][_set_def_col]
                _comment = "re-used from " + _set_def_each_row + " in " + _rel_full_name
                _pn_len = len(str(df_list[_each_row][_pn_col]).strip())
                if _pn_len == 10:
                    _rel_type_each_row = "sop"
                elif _pn_len == 13:
                    _rel_type_each_row = "tsb"
                else:
                    #print("\n PN length not 10 or 13, skip the row .. ")
                    continue

                
                #for cts  - just version match is required, ufs requires no programmed number, so no other match needed, can match btn BU and EU and btn different release types
                #once reuse is found, bo further checks, as cts is same for all PNs
                if "cts_ino" not in _reuse_dict.keys():         #once the key is set, this will avoid further checks for re-usage in the upcoming rows
                    #print ("\n at cts")
                    if _cts_ino_reuse_status == "no":
                        if self._cts_ver == df_list[_each_row][_cts_ver_col]:
                            _reuse_dict["cts_ino"] = int(df_list[_each_row][_cts_DNo_col])
                            _cts_reused_no = _reuse_dict["cts_ino"] 
                            _reuse_dict["cts_comment"] = _comment
                            _cts_ino_reuse_status = "yes"
                    #the check below is for cases where the cts is once set for one pn and for the next pn the reuse_dict is again set to empty,so it will not have the key, but based on the _cts_ino_reuse_status the number already reused can be taken for the next pn also       
                    elif _cts_ino_reuse_status == "yes" and self._cts_ver == df_list[_each_row][_cts_ver_col]:
                        _reuse_dict["cts_ino"] = _cts_reused_no
                        _reuse_dict["cts_comment"] = _comment
                
                #new change for sxm - 15th May  
                #for sxm, dtv, lontium  - just version match is required, no other match needed, can match btn BU and EU and btn different release types  
                if _set["sxm"] == "yes" and self._sxm_ver != "" and self._sxm_fwver != "":                    
                    if "sxm_ino" not in _reuse_dict.keys():                        
                        if _sxm_ino_reuse_status == "no" and self._sxm_ver == df_list[_each_row][_sxm_ver_col] and self._sxm_fwver == df_list[_each_row][_sxm_fwver_col]:
                            _reuse_dict["sxm_ino"] = int(df_list[_each_row][_sxm_DNo_col])
                            _sxm_reused_no = _reuse_dict["sxm_ino"] 
                            _reuse_dict["sxm_comment"] = _comment
                            _sxm_ino_reuse_status = "yes"
                            #the check below is for cases where the cts is once set for one pn and for the next pn the reuse_dict is again set to empty,so it will not have the key, but based on the _cts_ino_reuse_status the number already reused can be taken for the next pn also       
                        elif _sxm_ino_reuse_status == "yes" and self._sxm_ver == df_list[_each_row][_sxm_ver_col] and self._sxm_fwver == df_list[_each_row][_sxm_fwver_col]:
                            _reuse_dict["sxm_ino"] = _sxm_reused_no
                            _reuse_dict["sxm_comment"] = _comment 
                else:
                    pass
                  
                if _set["dtv"] == "yes":  
                    if "dtv_ino" not in _reuse_dict.keys():    
                        if _dtv_ino_reuse_status == "no" and self._dtv_ver == df_list[_each_row][_dtv_ver_col] and self._dtv_ver != "":
                            #print ("version at row: ", df_list[_each_row][_dtv_ver_col])
                            _reuse_dict["dtv_ino"] = int(df_list[_each_row][_dtv_DNo_col])
                            _dtv_reused_no = _reuse_dict["dtv_ino"] 
                            _reuse_dict["dtv_comment"] = _comment
                            _dtv_ino_reuse_status = "yes" 
                        elif _dtv_ino_reuse_status == "yes" and self._dtv_ver == df_list[_each_row][_dtv_ver_col]:
                            _reuse_dict["dtv_ino"] = _dtv_reused_no
                            _reuse_dict["dtv_comment"] = _comment
                        
                        if "lont_ino" not in _reuse_dict.keys(): 
                            if _lont_ino_reuse_status == "no" and self._lont_ver == df_list[_each_row][_lont_ver_col] and self._lont_ver != "":
                                _reuse_dict["lont_ino"] = int(df_list[_each_row][_lont_DNo_col])
                                _lont_reused_no = _reuse_dict["lont_ino"] 
                                _reuse_dict["lont_comment"] = _comment
                                _lont_ino_reuse_status = "yes"
                            elif _lont_ino_reuse_status == "yes" and self._lont_ver == df_list[_each_row][_lont_ver_col]:
                                _reuse_dict["lont_ino"] = _lont_reused_no
                                _reuse_dict["lont_comment"] = _comment                             
                else:
                    pass    

                    
                if self._rel_type == _rel_type_each_row:
                 #app sw check, matches only between same release types , like within sop, tsbs etc, cross matching not considered, images differ for app sw images btn tsb and sop
                 #there may be cases where cross matching can happen, like a tsb might have used a production flat image, so image name should be considered
                 #the same applies for aurix as well - for aurix paths differ btn dev and prod normally
                 #but there can be rare cases where cross matching happens where a tsb uses a kms signed image and then in future the same aurix version non kms signed version is used for another release
                 #as to avoid accidential re-use of such cases, aurix sub path is considered now for re-use
                    if self._rel == "EU":
                        _each_row_appfile = (df_list[_each_row][_app_file_col]).split("_", 2)[-1].strip()
                        if _set["device_type"].lower() == "yes":
                            if "app_igno" not in _reuse_dict.keys():
                                if _app_igno_reuse_status == "no":
                                    if self._app_gver == df_list[_each_row][_app_ver_col] and self._app_gfile == _each_row_appfile and self._rel == _rel_each_row:
                                        _reuse_dict["app_igno"] = int(df_list[_each_row][_app_DNo_col])     #GAS image no
                                        _gapp_reused_no = _reuse_dict["app_igno"]
                                        _reuse_dict["app_comment"] = _comment
                                        _app_igno_reuse_status = "yes"
                            
                                elif _app_igno_reuse_status == "yes" and self._app_gver == df_list[_each_row][_app_ver_col] and self._app_gfile == _each_row_appfile and self._rel == _rel_each_row:
                                    _reuse_dict["app_igno"] = _gapp_reused_no
                                    _reuse_dict["app_comment"] = _comment                                   
                    
                        elif _set["device_type"].lower() == "non":
                            if "app_ingno" not in _reuse_dict.keys():
                                if _app_ingno_reuse_status == "no":
                                    #print ("\n non gas version:", self._app_ngver)
                                    if self._app_ngver == df_list[_each_row][_app_ver_col] and self._app_ngfile == _each_row_appfile and self._rel == _rel_each_row:
                                        print ("\n at here non gas version matches .. ")
                                        _reuse_dict["app_ingno"] = int(df_list[_each_row][_app_DNo_col])     #NONGAS image no
                                        print ("\n at here non gas version image number .. ")
                                        _ngapp_reused_no = _reuse_dict["app_ingno"]
                                        _reuse_dict["app_comment"] = _comment
                                        _app_ingno_reuse_status = "yes"
                            
                                elif _app_ingno_reuse_status == "yes" and self._app_ngver == df_list[_each_row][_app_ver_col]and self._app_ngfile == _each_row_appfile and self._rel == _rel_each_row:
                                    print ("\n at here non gas version matches and reuse status yes .. ")
                                    _reuse_dict["app_ingno"] = _ngapp_reused_no
                                    _reuse_dict["app_comment"] = _comment
                                    print ("\n at here non gas image number .. ",_reuse_dict["app_ingno"] )
                                        
                                    print ("\n app comment : ", _reuse_dict["app_comment"])
                    
                    #dnl number is required in any case, re-use also depends on the release type - tsb or sop , because of the path change in binaries
                    #if self._aurix_ver == df_list[_each_row][_aurix_ver_col] and self.aurix_blank_device == df_list[_each_row][_aurix_BNo_col] and self._rel == _rel_each_row :
                    # print("\n at aurix reuse dict: ", _reuse_dict)
                    if "aurix_idnlno" not in _reuse_dict.keys():
                        if _aurix_idnlno_reuse_status == "no":
                            if self._aurix_ver == df_list[_each_row][_aurix_ver_col] and self._aurix_sub_path == df_list[_each_row][_aurix_sub_path] and self._aurix_sub_path != "":
                                # print ("\n self.rel", self._rel)
                                # print ("\n _rel_each_row : ", _rel_each_row)
                                # print ("\n self._aurix_ver", self._aurix_ver)
                                # print ("\n _rel_each_row : ", df_list[_each_row][_aurix_ver_col])
                                # print ("\n current row df_list[_each_row] at aurix dnl:", df_list[_each_row])
                                # print ("\n aurix version from set : ", self._aurix_ver )
                                _reuse_dict["aurix_idnlno"] = int(df_list[_each_row][_aurix_dnl_DNo_col])
                                _aurixdnl_reused_no = _reuse_dict["aurix_idnlno"]
                                _reuse_dict["aurix_dnl_comment"] = _comment
                                _aurix_idnlno_reuse_status = "yes"
                                #print ("\n aurix dnl reuse status set to yes")
                        elif _aurix_idnlno_reuse_status == "yes" and self._aurix_ver == df_list[_each_row][_aurix_ver_col]and self._aurix_sub_path == df_list[_each_row][_aurix_sub_path]and self._aurix_sub_path != "":
                        #elif _aurix_idnlno_reuse_status == "yes" and self._aurix_ver == df_list[_each_row][_aurix_ver_col]:
                            _reuse_dict["aurix_idnlno"] = _aurixdnl_reused_no
                            _reuse_dict["aurix_dnl_comment"] = _comment

                         
                        
                    if self._rel == "BU":
                        #print("\n at aurix reuse dict: ", _reuse_dict)
                        if "aurix_pno" not in _reuse_dict.keys() and "aurix_isrecno" not in _reuse_dict.keys():
                            #print ("\n at aurix srec")
                            if _aurix_isrecno_reuse_status == "no":
                                if self._aurix_ver == df_list[_each_row][_aurix_ver_col] and _set["aurix_blank_device"] == df_list[_each_row][_aurix_BNo_col] and self._aurix_sub_path == df_list[_each_row][_aurix_sub_path] and self._rel == _rel_each_row and self._aurix_sub_path != "":
                                    #print ("\n current row df_list[_each_row] at aurix srec:", df_list[_each_row])
                                    # print ("\n aurix version from set at srec : ", self._aurix_ver )
                                    #print ("\n aurix df type : ", type(df_list[_each_row][_aurix_PNo_col] ))
                                    _reuse_dict["aurix_pno"] = df_list[_each_row][_aurix_PNo_col]     #pno required only for baseunits
                                    _reuse_dict["aurix_pno_comment"] = _comment
                                    _reuse_dict["aurix_isrecno"] = int(df_list[_each_row][_aurix_srec_DNo_col])
                                    _reuse_dict["aurixsrec_comment"] = _comment
                                    _aurixsrec_reused_no = _reuse_dict["aurix_isrecno"]
                                    _aurixpno_reused_no = _reuse_dict["aurix_pno"]
                                    _aurix_isrecno_reuse_status = "yes"

                            elif _aurix_isrecno_reuse_status == "yes" and self._aurix_ver == df_list[_each_row][_aurix_ver_col] and self._aurix_sub_path == df_list[_each_row][_aurix_sub_path] and self._rel == _rel_each_row and self._aurix_sub_path != "":
                                _reuse_dict["aurix_pno"] = _aurixpno_reused_no
                                _reuse_dict["aurix_pno_comment"] = _comment 
                                _reuse_dict["aurix_isrecno"] = _aurixsrec_reused_no
                                _reuse_dict["aurixsrec_comment"] = _comment 

                                   
                                   
                    #ublox has shifted and normal version, shifted is used for baseunits and the other is used for end-unit (reflash)
                    #baseunit  - endor/bosch/firmware/u-blox/production, endunit - endor/bosch/firmware/u-blox
                    #for reuse, match can happen either btn baseunits or btn endunits. 
                    #for baseunits match should happen btn same release type( i.e - btn BU or btn EU ) as programming PNs differ
                    # print ("\n ublox version : ",type(self._ublox_ver ))
                    # print ("\n ublox at df :", type(str(df_list[_each_row][_ublox_ver_col])))
                    # print ("\n blank device: ", _set["ublox_blank_device"] )
                    #int(df_list[_each_row][_ublox_ver_col]).strip()
                    #print("\n at ublox reuse dict: ", _reuse_dict)
                    if self._ublox_ver == str(df_list[_each_row][_ublox_ver_col]):
                        if self._rel == "BU":
                            if "ublox_pno" not in _reuse_dict.keys():
                                if _ublox_pno_reuse_status == "no":
                                    if _set["ublox_blank_device"] == df_list[_each_row][_ublox_BNo_col] and self._rel == _rel_each_row :
                                        _reuse_dict["ublox_pno"] = df_list[_each_row][_ublox_PNo_col]
                                        _ubloxpno_reusedno = _reuse_dict["ublox_pno"]
                                        _reuse_dict["ubloxpno_comment"] = _comment
                                        _ublox_pno_reuse_status = "yes"
                                       
                                elif _ublox_pno_reuse_status == "yes" and _set["ublox_blank_device"] == df_list[_each_row][_ublox_BNo_col] and self._rel == _rel_each_row :
                                    _reuse_dict["ublox_pno"] = _ubloxpno_reusedno
                                    _reuse_dict["ubloxpno_comment"] = _comment

                        
                        if self._rel == _rel_each_row:
                            if "ublox_ino" not in _reuse_dict.keys():
                                if _ublox_ino_reuse_status == "no":
                                    _reuse_dict["ublox_ino"] = df_list[_each_row][_ublox_DNo_col]
                                    _ubloxino_reusedno = _reuse_dict["ublox_ino"]
                                    _reuse_dict["ubloxino_comment"] = _comment
                                    _ublox_ino_reuse_status = "yes"
                                elif _ublox_ino_reuse_status == "yes" :
                                    _reuse_dict["ublox_ino"] = _ubloxino_reusedno
                                    _reuse_dict["ubloxino_comment"] = _comment
  
            _set.update(_reuse_dict)   
    
    def prepare_sap_sheet_template(self,_sap_sheet_temp ):  
       # _PN_count = _PN_count
       
        _sap_sheet = _sap_sheet_temp
       
        print ("at prepare_sap_sheet_template .." )
        print ("Using ", _sap_sheet, "to prepare the sap request sheet template .. ")
        
        global _PN_list
        global _ecn
        _sheet1 = "SAP Partnumbers"
        _sheet2 = "HW Mapping"
        
        
        self._sap_request_sheet = self._set_def_no + "_SAP_Request.xlsx"
        _curr_Dir = os.getcwd()
        _sap_File = os.path.join(_curr_Dir,self._sap_request_sheet) 
        #print ("\n SAP file : ",_sap_File)
   
        if os.path.exists(_sap_File):
            #_suffix = datetime.now().strftime("%Y%m%d_%H%M%S") #creates suffix with current date and exact time
            _suffix = dt.now().strftime("%Y%m%d_%H%M%S") #creates suffix with current date and exact time
            #print("\n date suffix: ", _suffix)
            # _a = os.path.basename(_sap_File)
            # print ("\n a:", _a)
            _sap_basename = os.path.splitext(os.path.basename(_sap_File))[0]
            #print ("\n _sap_basename:", _sap_basename)
            _file_rename = _sap_basename + "." + _suffix + ".xlsx"
            _sap_rename = os.path.join(_curr_Dir,_file_rename)
            os.rename(self._sap_request_sheet, _sap_rename) 
            print("\n renamed the existing sap sheet .. ")
            
    
        
        #may not require the below   
        _space_btn_sets = 3 
        _space_to_minor = 2  #may not need now as we do not have minor PNs concept now
       
        print ("\nat prepare_sap_sheet_template .." )
        _xls_File = _sap_sheet
        _app = xw.App(visible=False)
        wb = xw.Book(_xls_File)         #xlwings object xw is used to enable copy and paste as in windows 
        ws1 =  wb.sheets['SAP Partnumbers']
        ws2 =  wb.sheets['HW Mapping']  
        ws1.range('C:C').copy()                     #copy commandlike cntrl C
        #ws2.range('A2:F6').copy()
        _len =  len(self._set_info_list)
        _row_start = 6
        _col_start = 3
        _row_start_hm = 2
        _totcols  = ( _col_start + _len ) - 1       # 1 col is already there in the templae, so 1 is reduced
        _tot_rows = 4
        #print ("\n total cols to add on to the sap sheet : ", _totcols)
        for _col in range(_col_start, _totcols):
            if ( 1 <= _col <= 26):
                _alp = chr(_col + 65)
                ws1.range(f'{_alp}1').paste(paste='all')
        
        ws2.range('A2:F6').copy()                   #copy lines of HW mapping sheet
        _i = 1
        while _i < ( _len):
            _row_start_hm = _row_start_hm + 5
            ws2.range(f'A{_row_start_hm}').paste(paste='all')
            _i += 1
            
                
        wb.save(self._sap_request_sheet)
        wb.close()  
        _app.quit()
        
        # for _set in self._set_info_list:
            #print ("\n _set", _set)
        self.fill_sap_sheet(self)
          
    
    
    def fill_sap_sheet(self,_set):
        print ("\nat fill_sap_sheet .." )
       # _PN_count = _PN_count
        global _PN_list
        global _ecn
        _sheet1 = "SAP Partnumbers"
        _sheet2 = "HW Mapping"
        _grey_cell_list = []
        
        _relname_cell = "B1"
        _sample_cell = "A2"
        _ecn_cell = "C3"
        _ecr_cell = "C4"
        
        _ecn_row =  3
        _ecr_row = 4 
        _cont_row = 6
        _setnum_row = 7
        _docnum_row = 8
        _pn_row = 9
        _UFS_blank_row = 12
        _aurix_prog_row = 13
        _aurix_blank_row = 14
        _ublox_prog_row = 15
        _ublox_blank_row = 16
        
        _cts_ino_row = 19
        _appsw_ino_row = 20
        _aurix_srec_ino_row = 22
        _aurix_dnl_ino_row = 23
        _ublox_ino_row = 25
        _sxm_ino_row = 27
        _dtv_ino_row = 29
        _lont_ino_row = 31
        
        _cont_start_col = "C"
        _cont_start_colno = 3
        
        _col_start = 3
        _start_col = 3
        _len =  len(self._set_info_list)
        _c = _col_start
        _col_max = ( _len + _col_start )
        
        #hw sheet    
        _elec_pn_row = 4
        _cont_name_col = 1
        _pn_col = 2
        _dev_col = 3
        _pos_col = 4
        _blank_dev_pn_col = 5
        _elec_pn_col = 6
        _sapsheet_cont_col = _col_start
        
        _set_gasino_status = "no"
        _set_nongasino_status = "no"
        _gino_cell_val = ""
        _ngino_cell_val = ""
        _set_sxmino_status = "no"
        _set_dtvino_status = "no"
        _set_lontino_status = "no"
    
        
        _grey_fill = PatternFill(start_color='D3D3D3', fill_type='solid')
        _reuse_fill = PatternFill(start_color='CCFFFF', fill_type='solid')      #pale blue
        _setreuse_fill = PatternFill(start_color='CCFFCC', fill_type='solid')   #pale green
        _newno_fill = PatternFill(start_color='FFFFCC', fill_type='solid')  
        
        _notes_dict = {}   
        
        wb1 =  openpyxl.load_workbook(self._sap_request_sheet, data_only=True)     #load the work book , data_only flag helps to get the value instead of the formaula from a cell in the excel
        ws3 =  wb1['SAP Partnumbers']
        ws4 =  wb1['HW Mapping'] 
        ws3[_relname_cell] = self._sw_ver
		

        for _set in self._set_info_list:
            _notes_dict = {} 
            if "cts_ino" in _set.keys():
                _cts_ino = _set["cts_ino"]
                _cts_com = _set["cts_comment"]
                _notes_dict[_cts_ino_row] = _cts_com
            else:
               _cts_ino = ""
               _cts_com = ""
                   
            if "aurix_idnlno" in _set.keys():
                _aur_idnlno = _set["aurix_idnlno"]
                _aur_idnl_com = _set["aurix_dnl_comment"]
                _notes_dict[_aurix_dnl_ino_row] = _aur_idnl_com
            else:
                _aur_idnlno = ""
                _aur_idnl_com = ""
                    
            if "ublox_ino" in _set.keys():
                _ublox_ino = _set["ublox_ino"]
                _ublox_ino_com = _set["ubloxino_comment"]
                _notes_dict[_ublox_ino_row] = _ublox_ino_com
            else:
                _ublox_ino = ""
                _ublox_ino_com = ""
                
            #change start Nisha = ImageNameChange
            if "sxm_ino" in _set.keys():
                _sxm_ino = _set["sxm_ino"]
                _sxm_com = _set["sxm_comment"]
                _notes_dict[_sxm_ino_row] = _sxm_com
            else:
               _sxm_ino = ""
               _sxm_com = ""

            if "dtv_ino" in _set.keys():
                _dtv_ino = _set["dtv_ino"]
                _dtv_com = _set["dtv_comment"]
                _notes_dict[_dtv_ino_row] = _dtv_com
            else:
                _dtv_ino = ""
                _dtv_com = ""
               
            if "lont_ino" in _set.keys():
                _lont_ino = _set["lont_ino"]
                _lont_com = _set["lont_comment"]
                _notes_dict[_lont_ino_row] = _lont_com
            else:
                _lont_ino = ""
                _lont_com = ""
               
               
               
                
            if self._rel == "BU":
                if "aurix_pno" in _set.keys():
                    _aur_pno = _set["aurix_pno"]
                    _aur_pno_com = _set["aurix_pno_comment"]
                    _notes_dict[_aurix_prog_row] = _aur_pno_com
                else:
                    _aur_pno = ""
                    _aur_pno_com = ""
                   
                if "ublox_pno" in _set.keys():
                    _ublox_pno = _set["ublox_pno"]
                    _ublox_pno_com = _set["ubloxpno_comment"]
                    _notes_dict[_ublox_prog_row] = _ublox_pno_com
                else:
                    _ublox_pno = ""
                    _ublox_pno_com = ""
            
                if "aurix_isrecno" in _set.keys():
                    _aur_isrecno = _set["aurix_isrecno"]
                    _aur_isrec_com = _set["aurixsrec_comment"]
                    _notes_dict[_aurix_srec_ino_row] = _aur_isrec_com
                else:
                    _aur_isrecno = ""
                    _aur_isrec_com = ""
                
                _setlist = [_set["product"], "", "", _set["part_number"], "", "", _set["ufs_blank_device"], _aur_pno, _set["aurix_blank_device"], _ublox_pno, _set["ublox_blank_device"], "", "", _cts_ino, "", "", _aur_isrecno, _aur_idnlno, "", _ublox_ino, "", _sxm_ino, "", _dtv_ino, "", _lont_ino]									    
                
                
            elif self._rel == "EU":
                if _set["device_type"] == "yes":
                    if "app_igno" in _set.keys():
                        _app_igno = _set["app_igno"]
                        _app_igno_com = _set["app_comment"]
                        _notes_dict[_appsw_ino_row] = _app_igno_com
                    else:
                        _app_igno = ""
                        _app_igno_com = ""
                    _setlist = [_set["product"], "", "", _set["part_number"], "", "", _set["ufs_blank_device"], "", _set["aurix_blank_device"], "", _set["ublox_blank_device"], "", "", _cts_ino, _app_igno, "", "", _aur_idnlno, "", _ublox_ino, "", _sxm_ino, "", _dtv_ino, "", _lont_ino]	
                
                elif _set["device_type"] == "non":
                    if "app_ingno" in _set.keys():
                        print ("\n non gas in set keys ")
                        _app_ingno = _set["app_ingno"]
                        _app_ingno_com = _set["app_comment"]
                        _notes_dict[_appsw_ino_row] = _app_ingno_com
                    else:
                        print ("\n non gas not in set keys ")
                        _app_ingno = ""
                        _app_ingno_com = ""   
                    _setlist = [_set["product"], "", "", _set["part_number"], "", "", _set["ufs_blank_device"], "", _set["aurix_blank_device"], "", _set["ublox_blank_device"], "", "", _cts_ino, _app_ingno, "", "", _aur_idnlno, "", _ublox_ino, "", _sxm_ino, "", _dtv_ino, "", _lont_ino]	
                
            #print ("\n set_list : ", _setlist)
            if self._rel == "EU":
                _grey_cell_list = [_aurix_prog_row, _ublox_prog_row,_aurix_srec_ino_row ]
                _ino_row_list = [_cts_ino_row, _aurix_dnl_ino_row, _ublox_ino_row ]
                
            else:
                _ino_row_list = [_aurix_prog_row, _ublox_prog_row, _cts_ino_row, _aurix_srec_ino_row, _aurix_dnl_ino_row, _ublox_ino_row ] 
                _grey_cell_list = [_appsw_ino_row]
            
   
            if _set["sxm"] != "yes":
                _grey_cell_list.append(_sxm_ino_row)

            if _set["dtv"] != "yes":    
                _grey_cell_list.append(_dtv_ino_row)
                _grey_cell_list.append(_lont_ino_row)
                
            _row_start = 6
            _r_max = 32
            for _item in _notes_dict:
                _notes_dict[_item] = Comment(text =_notes_dict[_item], author="")

            _x = 0
            for _r in range (_row_start , _r_max ):
                    _col_letter = get_column_letter(_c)
                    _cell_pos = _col_letter + str(_r)
                    
                    if _r in _grey_cell_list and self._rel == "EU":
                        ws3[_cell_pos].fill = _grey_fill
                    elif _r in _grey_cell_list and self._rel == "BU":
                        ws3[_cell_pos].fill = _grey_fill 

                    if _r == _appsw_ino_row:
                        if self._rel == "EU" and _set["device_type"] == "yes":
                            if _set_gasino_status == "no":
                                _set_gasino_status = "yes"
                                _app_gino = "=" + _cell_pos 
                                ws3[_cell_pos] = _setlist[_x]
 
                            elif _set_gasino_status == "yes":
                                ws3[_cell_pos] = _app_gino
                                ws3[_cell_pos].fill = _setreuse_fill
                        
                        elif self._rel == "EU" and _set["device_type"] == "non":
                            if _set_nongasino_status == "no":
                                _set_nongasino_status = "yes"
                                _app_ngino = "=" + _cell_pos 
                                ws3[_cell_pos] = _setlist[_x]
                               
                            elif _set_nongasino_status == "yes":
                                ws3[_cell_pos] = _app_ngino
                                ws3[_cell_pos].fill = _setreuse_fill
                    
                    #for sxm
                    elif _r == _sxm_ino_row and _set["sxm"] == "yes":            
                        if _set_sxmino_status == "no":
                            _set_sxmino_status = "yes"
                            _sxm_i_value = "=" + _cell_pos 
                            ws3[_cell_pos] = _setlist[_x]
   
                        elif _set_sxmino_status == "yes":
                            ws3[_cell_pos] = _sxm_i_value
                            ws3[_cell_pos].fill = _setreuse_fill
                    #for dtv
                    elif _r == _dtv_ino_row and _set["dtv"] == "yes":             
                        if _set_dtvino_status == "no":
                            _set_dtvino_status = "yes"
                            _dtv_i_value = "=" + _cell_pos 
                            ws3[_cell_pos] = _setlist[_x]
   
                        elif _set_dtvino_status == "yes":
                            ws3[_cell_pos] = _dtv_i_value
                            ws3[_cell_pos].fill = _setreuse_fill
                    #for lontium
                    elif _r == _lont_ino_row and _set["dtv"] == "yes":            
                        if _set_lontino_status == "no":
                            _set_lontino_status = "yes"
                            _lont_i_value = "=" + _cell_pos 
                            ws3[_cell_pos] = _setlist[_x]
                        elif _set_lontino_status == "yes":
                            ws3[_cell_pos] = _lont_i_value
                            ws3[_cell_pos].fill = _setreuse_fill
                    
                    elif _r in _ino_row_list and _c != _cont_start_colno:
                        _ino_cell_val = "=" + _cont_start_col + str(_r) 
                        ws3[_cell_pos ] = _ino_cell_val
                        ws3[_cell_pos].fill = _setreuse_fill
                        #for numbers which is new and the numbers that gets reused also ex: ublox, it gets the cell formula from here
                    else:
                        ws3[_cell_pos ] = _setlist[_x] 
                    
                    for _item in _notes_dict:
                        if _r == _item:
                            # for numbers that gets reused ex: ublox, it gets the cell comment and cell color filled in from here
                            ws3[_cell_pos].comment = _notes_dict[_item]
                            ws3[_cell_pos].fill = _reuse_fill   
                            
                    _x += 1
            _c += 1    
             
       
        if self._Chng_NR_type == "withECN":
            _start_row = 3

        elif self._Chng_NR_type == "withECR":
            _start_row = 4
            ws3[_ecr_cell] = self._Chng_NR
           
        
        _start = ws3.cell(row = _start_row, column = _start_col)
        _end = ws3.cell(row = _start_row, column = _col_max )
        ws3.merge_cells(start_row=_start_row, start_column=_start_col, end_row=_start_row, end_column=_col_max -1)
        ws3.cell(row = _start_row, column = _start_col).fill = _newno_fill
        
        # HW sheet update
        for _set in self._set_info_list:
            _sap_sheet = "SAP Partnumbers"
            _sapsheet_cont_letter = chr(_sapsheet_cont_col + 64)
            _cont_cell_ref = "='" + _sap_sheet + "'!" + _sapsheet_cont_letter + str(_cont_row)
            _pn_cell_ref = "='" + _sap_sheet + "'!" + _sapsheet_cont_letter + str(_pn_row)
            _ufs_bno_cell_ref = "='" + _sap_sheet + "'!" + _sapsheet_cont_letter + str(_UFS_blank_row)
            _aurix_bno_cell_ref = "='" + _sap_sheet + "'!" + _sapsheet_cont_letter + str(_aurix_blank_row)
            _ublox_bno_cell_ref = "='" + _sap_sheet + "'!" + _sapsheet_cont_letter + str(_ublox_blank_row)
            
            _cont_col_letter = get_column_letter(_cont_name_col)
            _pn_col_letter = get_column_letter(_pn_col)
            _bno_col_letter = get_column_letter(_blank_dev_pn_col )
            
            _hwsheet_cont_name = _cont_col_letter + str(_elec_pn_row)
            _hwsheet_pn = _pn_col_letter + str(_elec_pn_row)            
            _hwsheet_ublox_bno = _bno_col_letter + str(_elec_pn_row )
            _hwsheet_ufs_bno = _bno_col_letter + str(_elec_pn_row + 1)
            _hwsheet_aurix_bno = _bno_col_letter + str(_elec_pn_row + 2 )
			
            ws4[_hwsheet_cont_name] = _cont_cell_ref
            ws4[_hwsheet_pn] = _pn_cell_ref
            ws4[_hwsheet_ufs_bno] = _ufs_bno_cell_ref
            ws4[_hwsheet_aurix_bno] = _aurix_bno_cell_ref
            ws4[_hwsheet_ublox_bno] = _ublox_bno_cell_ref
            _sapsheet_cont_col += 1
            _elec_pn_row += 5

        wb1.save(self._sap_request_sheet)
        wb1.close()  
       


def main():
         
    _spl_template = "CCS2_SPL_TMPL_V01.00.xlsx"
    _sap_template = "SAP_Request_template.xlsx"
    if not os.path.isfile(_spl_template):
        raise FileNotFoundError ( "File", _spl_template, " not found ")
        
    if not os.path.isfile(_sap_template):
        raise FileNotFoundError ( "File", _sap_template, " not found ")
    _curr_Dir = os.getcwd() 
    parser =  argparse.ArgumentParser(description="Process either with master xml file or SetDefinition xls file ")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('-x', '--xml', type=str, help='master xml file')
    group.add_argument('-t', '--template', type=str, help='SetDefinition file')
    args = parser.parse_args()
    if args.xml:
        input_masterxml = args.xml
        _xml_Filename = input_masterxml
        _start_str = _xml_Filename.startswith("P2_")
        _end_str = _xml_Filename.endswith(".xml")
        #print("\n _start_str: ", _start_str) 
        if _start_str and _end_str:
            print ("\n Master xml ", _xml_Filename," will be used ")
           
            #_xls_File = os.path.join(_curr_Dir,_xls_Filename)   
            _xml_File = os.path.join(_curr_Dir,_xml_Filename) 
    
            ob1 = _SAP_request()
            ob1.process_xml(_xml_File)
            ob1.parse_master_xml()
            ob1.get_blank_dev_numbers(_spl_template)
            ob1.check_del_sheet()
            ob1.prepare_sap_sheet_template(_sap_template)
    
        else:
            print ("\n please provide the master xml that starts with P2_<SetDefNo>_<BU/EU>_SW_MasterInfo_CCS2.xml")
            print ("\n exiting .. ")
            sys.exit(0)
    
    elif args.template:
        #print ("\n here .. ")
        input_xls = args.template
        _xls_Filename = input_xls
        _start_str = _xls_Filename.startswith("SetDefinition_CCS2_P2_")
        _end_str = _xls_Filename.endswith(".xlsx")
        #print("\n _start_str: ", _start_str) 
        if _start_str and _end_str:
            print ("\n SetDefinition template ", _xls_Filename," will be used ")
            _xls_File = os.path.join(_curr_Dir,_xls_Filename) 
            ob2 = _SAP_request()
            ob2.process_xls(_xls_File)
            ob2.read_SetDefinition_sheet()
            ob2.get_blank_dev_numbers(_spl_template)
            ob2.check_del_sheet()
            ob2.prepare_sap_sheet_template(_sap_template)
    
            
        else:
            print ("\n please provide the SetDefinition template that starts with SetDefinition_CCS2_P2_<SetDefNo>_<BU/EU>.xlsx")
            print ("\n exiting .. ")
            sys.exit(0)
            
   
   
    print ("\n exiting .. ")
    sys.exit(0)
    
    
    
        
if __name__ == '__main__':
    main()  
    
    
    
    
    
    
# if __name__ == '__main__':
    # parser = optparse.OptionParser()
    # parser.add_option('-x', '--xml', dest='xml', default=None,
                      # help='Master xml')
    # # parser.add_option('-t', '--template', dest='template', default=None,
                      # # help='SAP Request sheet template')
                     
    # (options, args) = parser.parse_args()
    # if not options.xml:
        # #input_masterxml = input("\nPlease enter masterxml filename\n")
        # print ("\n Please enter the master xml name as input ")
        # print ("\n please run the script in the form : ")
        # print ("\t create_sap_sheet_ccs2.py -x <xml_name> -t <sap_sheet_template> ")
        # sys.exit(0) 
    # else:
        # input_masterxml = options.xml
        # _xml_Filename = input_masterxml
        # _start_str = _xml_Filename.startswith("P2_")
        # print("\n _start_str: ", _start_str) 
        # if _start_str:
            # print ("\n Master xml ", _xml_Filename," will be used ")
        # else:
            # print ("\n please provide the master xml that starts with P2_<SetDefNo>_<BU/EU>_SW_MasterInfo_CCS2.xml"
            # print ("\n exiting .. ")
            # sys.exit(0)
	
    # _spl_template = "CCS2_SPL_TMPL_V01.00.xlsx"
    # _sap_template = "SAP_Request_template.xlsx"
    # if not os.path.isfile(_spl_template):
        # raise FileNotFoundError ( "File", _spl_template, " not found ")
        
    # if not os.path.isfile(_sap_template):
        # raise FileNotFoundError ( "File", _sap_template, " not found ")
        
    
    
    
    # _curr_Dir = os.getcwd()
    # #_xls_File = os.path.join(_curr_Dir,_xls_Filename)   
    # _xml_File = os.path.join(_curr_Dir,_xml_Filename) 
    
    # ob1 = _SAP_request(_xml_File)
    # ob1.parse_master_xml()
    
    # #parse_master_xml(_xml_File)
    
    
    # # if os.path.exists(_sap_template):
        # # print ("\n sap template exists .. ")
        # # #_sap_sheet  = _set_def_no +  "_SAP_Request.xlsx"
        # # #copy( _sap_template, _sap_sheet)
        
        # # ob1.get_blank_dev_numbers(_spl_template)
        # # ob1.check_del_sheet()
        # # ob1.prepare_sap_sheet_template(_sap_template)
        # #ob1.fill_sap_sheet()
    # #update_sap_xml("sap_info", _new_xml)
    # ob1.get_blank_dev_numbers(_spl_template)
    # ob1.check_del_sheet()
    # ob1.prepare_sap_sheet_template(_sap_template)
    
    
        

        
        