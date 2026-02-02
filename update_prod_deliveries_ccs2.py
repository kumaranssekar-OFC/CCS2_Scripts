###############################################################################################################################################################################
#
# FILE:         update_prod_deliveries_ccs2.py
# DESCRIPTION:  This script can be used to create the SAP request sheet for CCS2 project
# USAGE:        see help_text
# PREREQUISITE: The file ProdDeliveries_Overview_CCS2.xlsx has to be available in the _PCM_internal\Reference_sheets\ path in the production server
# HISTORY:
# Date         | Author          		| Modification
# 29.08.2024   | Nisharani C  			| Initial version
# 18.09.2024   | Nisharani C  			| help section update
# 26.06.2025   | Nisharani C            | Adaptation for Image file name changes - SXM, DTV, Lontium - filenames to be updated in deliveries sheet for reuse check
# 08.10.2025   | Nisharani C			| Include app_id column in prod deliveries sheet
# 25.11.1981   | Nisharani C            | Include Aurix_Sub_Path and version number sheet in prod deliveries sheet 
###############################################################################################################################################################################

import os
import re
import hashlib
import subprocess
import optparse
import sys
import openpyxl
import xml.etree.ElementTree as ET
import win32com.client as win32

# _startlines = ""

global _rel_version
global _sw_ver
global _Chng_NR


class _update_overview_file:

    def __init__(self, _infodict):
        self._maildict = _infodict
        #self._xls_File = r'D:\Nisha\Tasks\CCS2_Tooling\set_def\Read_Artifactory\test_runs\ProdDeliveries_Overview_CCS2.xlsx'     
        self._xls_File = r'\\bosch.com\dfsrb\DfsDE\DIV\CM\AI\SW_Production\Nissan\0060_CCS2_7515752366\00_SW\_PCM_internal\Reference_sheets\ProdDeliveries_Overview_CCS2.xlsx'
    def update_sheet(self):
        print ("\n at update_sheet ..")
        wb =  openpyxl.load_workbook(self._xls_File, data_only=True)     #load the work book , data_only flag helps to get the value instead of the formaula from a cell in the excel
        ws =  wb['Sheet1']               
        
        _row_start = 3
        _col_start = 2
        _rel_name_col = 2
        _ecn_ecr_col = 3
        _pn_col = 4
        _ufs_BNo_col = 5
        _cts_ver_col = 6
        _cts_DNo_col = 7
        _cts_file_col = 8
        _app_id_col = 9
        _app_ver_col = 10
        _app_DNo_col = 11
        _app_file_col = 12
        _aurix_BNo_col = 13
        _aurix_PNo_col = 14
        _aurix_ver_col = 15
        _aurix_sub_path = 16
        _aurix_srec_DNo_col = 17
        _aurix_srec_file_col = 18
        _aurix_dnl_DNo_col = 19
        _aurix_dnl_file_col = 20
        _ublox_BNo_col =  21
        _ublox_PNo_col = 22
        _ublox_ver_col = 23
        _ublox_DNo_col = 24
        _ublox_file_col = 25
        
        _sxm_ver_col = 26
        _sxm_fwver_col = 27
        _sxm_DNo_col = 28
        _sxm_file_col = 29
        _dtv_ver_col = 30
        _dtv_DNo_col = 31
        _dtv_file_col = 32
        _lont_ver_col =33
        _lont_DNo_col = 34
        _lont_file_col = 35

        _set_defNo_col = 36
        
        _cell_list = []
        _r_start = ws.max_row + 1
        _col_list = [_rel_name_col, _ecn_ecr_col, _pn_col, _ufs_BNo_col, _cts_ver_col, _cts_DNo_col, _cts_file_col, _app_id_col, _app_ver_col, _app_DNo_col, _app_file_col, _aurix_BNo_col, _aurix_PNo_col, _aurix_ver_col, _aurix_sub_path, _aurix_srec_DNo_col, _aurix_srec_file_col, _aurix_dnl_DNo_col, _aurix_dnl_file_col, _ublox_BNo_col, _ublox_PNo_col, _ublox_ver_col, _ublox_DNo_col, _ublox_file_col, _sxm_ver_col, _sxm_fwver_col, _sxm_DNo_col, _sxm_file_col, _dtv_ver_col, _dtv_DNo_col, _dtv_file_col, _lont_ver_col, _lont_DNo_col, _lont_file_col, _set_defNo_col]
        
        
        # print ("\n length of _col_list : ", len( _col_list ) )
        # print ("maximum rows : ", ws.max_row)
        # print ("\n maximun cols :", ws.max_column ) 
        #print ("\n set_dict :", _set_dict )
        
        
        
        for x in _set_dict:
            print ("\n update for PN : ", _set_dict[x]["part_number"] )
            
            if _set_dict[x]["device_type"] == "yes":
                _app_ver = _maildict["_app_gas_ver"]
                _app_id = _maildict["_SWID_gas"]
                _app_filename = _set_dict[x]["_app_gfilename"]
                _app_ino = _set_dict[x]["app_image_number"]
            elif _set_dict[x]["device_type"] == "non":
                _app_ver = _maildict["_app_nongas_ver"]
                _app_id = _maildict["_SWID_nongas"]
                _app_filename = _set_dict[x]["_app_ngfilename"]
                _app_ino = _set_dict[x]["app_image_number"]
            else:
                _app_filename = "NA"
                _app_ino = "NA"
                _app_ver = "NA"
                _app_id = "NA"

            if _maildict["_rel_type"] == "BU":
                _aurix_pno = _set_dict[x]["scc_programmed_device_number"]
                _aurix_srec_ino = _set_dict[x]["aurix_srec_image_number"]
                _aurix_srec_filename = _set_dict[x]["_aurix_srec_filename"]
                _ublox_pno = _set_dict[x]["ublox_programmed_device_number"]
            elif _maildict["_rel_type"] == "EU":
                _aurix_pno = "NA"
                _aurix_srec_ino = "NA"
                _aurix_srec_filename = "NA"
                _ublox_pno = "NA"
                
            if _set_dict[x]["sxm"] == "yes":
                _sxm_ino = _set_dict[x]["sxm_image_number"]
                _sxm_filename = _set_dict[x]["_sxm_filename"]
                _sxm_ver = _maildict["_sxm_ver"] 
                _sxm_fwver = _maildict["_sxm_fwver"] 
            elif _set_dict[x]["sxm"] == "none":
                _sxm_ino = "NA"
                _sxm_filename = "NA"
                _sxm_ver = "NA"
                _sxm_fwver = "NA"
                
            if _set_dict[x]["dtv"] == "yes":
                _dtv_sw_ino = _set_dict[x]["dtv_image_number"]
                _dtv_filename = _set_dict[x]["_dtv_swfilename"]
                _dtv_sw_ver = _maildict["_dtv_sw_ver"]
                
                _lont_sw_ino = _set_dict[x]["lont_image_number"]
                _lont_filename = _set_dict[x]["_lont_filename"]
                _lont_ver = _maildict["_lont_ver"] 
                
            elif _set_dict[x]["dtv"] == "none":
                _dtv_sw_ino = "NA"
                _dtv_filename = "NA"
                _dtv_sw_ver = "NA"                
                _lont_sw_ino = "NA"
                _lont_filename = "NA"
                _lont_ver = "NA"
               
            _col_val_list = [ _maildict["rel_version"], _maildict["ecn_ecr"], _set_dict[x]["part_number"], _set_dict[x]["ufs_blank_device_number"], _maildict["_cts_ver"], _set_dict[x]["cts_image_number"], _set_dict[x]["_cts_filename"], _app_id, _app_ver, _app_ino, _app_filename, _set_dict[x]["scc_blank_device_number"], _aurix_pno, _maildict["_aurix_ver"], _maildict["_aurix_sub_path"], _aurix_srec_ino, _aurix_srec_filename, _set_dict[x]["aurix_dnl_image_number"], _set_dict[x]["_aurix_dnl_filename"], _set_dict[x]["ublox_blank_device_number"], _ublox_pno, _maildict["_ublox_ver"],_set_dict[x]["ublox_image_number"], _set_dict[x]["_ublox_filename"], _sxm_ver, _sxm_fwver, _sxm_ino, _sxm_filename, _dtv_sw_ver, _dtv_sw_ino, _dtv_filename, _lont_ver, _lont_sw_ino, _lont_filename, _maildict["_set_def"]]
            # print("\n col value list :", _col_val_list )
            # print ("\n length of col value list : ", len(_col_val_list) )
            # print ("\n data type of row start  :", type(_r_start))
            # print ("\n data type of   :", type(_rel_name_col))
            _index = 0
            #ws.autofit()
            for _val in _col_val_list:
                #_col_list[_index ].value 
                _c = _col_list[_index]
                #print ("\n data type of row start  :", type(_r_start))
                ws.cell(row= _r_start, column= _c ).value  = _val
                #ws.cell(row=_r_start, column=_c).value = "testentry"
                _index  +=  1
            _r_start += 1    
        
        for sheets in wb.sheetnames:   
            if("V" in sheets):
                VerSheet = wb[sheets]
                tmp = re.sub('V','',sheets)
                New_ver =float(tmp)+0.01
                New_ver = round(New_ver,2)
                VerSheet.title = 'V'+str(New_ver)
        
        
        wb.save(self._xls_File)
                

def parse_sap_xml(_sap_xml):
    print ("\n at parse sap xml .. ")
    global _set_dict
    global _maildict
    _maildict = {}
    tree = ET.ElementTree()
    tree.parse(_sap_xml)
    root = tree.getroot()
    
                  
    _doc_sec = root.find("./DocInfo")
    _doc_info = _doc_sec.attrib
    _set_def = _doc_sec.attrib["Doc_SetDef"]
    _maildict["_set_def"] = _set_def
    
    _dtv_sw = root.find("./SW_Versions[@key='DTV_SW']")
    _dtv_sw_info = _dtv_sw.attrib
    _dtv_sw_ver = _dtv_sw.attrib["version"]
    _maildict["_dtv_sw_ver"] = _dtv_sw_ver
    
    _lont = root.find("./SW_Versions[@key='Lontium_SW']")
    _lont_info = _lont.attrib
    _lont_ver = _lont.attrib["version"]
    _maildict["_lont_ver"] = _lont_ver
    
    _sxm = root.find("./SW_Versions[@key='SXM']")
    _sxm_info = _sxm.attrib
    _sxm_fwver = _sxm.attrib["fw_version"]
    _sxm_ver = _sxm.attrib["version"]
    _maildict["_sxm_fwver"] = _sxm_fwver
    _maildict["_sxm_ver"] = _sxm_ver
    
    _cts = root.find("./SW_Versions[@key='cts']")
    _cts_info = _cts.attrib
    _cts_ver = _cts.attrib["version"]
    _cts_version = _cts_ver.lower()  
    _cts_version = (_cts_version.split("cts_")[1]).upper()
    _maildict["_cts_ver"] = _cts_version
    
    _app = root.find("./SW_Versions[@key='android_app_sw']")
    _app_info = _app.attrib
    _app_gas_ver = _app.attrib["app_sw_gas_version"]
    _app_nongas_ver = _app.attrib["app_sw_nongas_version"]
    _maildict["_app_gas_ver"] = _app_gas_ver
    _maildict["_app_nongas_ver"] = _app_nongas_ver
    
    _aurix = root.find("./SW_Versions[@key='aurix_autosar_app_sw']")
    _aurix_info = _aurix.attrib
    _aurix_ver = _aurix.attrib["version"]
    _aurix_sub_path = _aurix.attrib["_sub_path"]
    _maildict["_aurix_ver"] = _aurix_ver
    _maildict["_aurix_sub_path"] = _aurix_sub_path
    
    _ublox = root.find("./SW_Versions[@key='ublox']")
    _ublox_info = _ublox.attrib
    _ublox_ver = _ublox.attrib["version"]
    _maildict["_ublox_ver"] = _ublox_ver
    
    _ecn =  root.find("./Overall_Infos[@Col1='ecn']")
    _ecn_no = _ecn.attrib["Col2"]
    
    _ecr =  root.find("./Overall_Infos[@Col1='ecr']")
    _ecr_no = _ecr.attrib["Col2"]

    if _ecr_no != "" and _ecr_no !="None":
        _Chng_NR = _ecr_no
    elif _ecn_no !="" and _ecn_no != "None":
        _Chng_NR = _ecn_no
    _Chng_NR = str(_Chng_NR) 

    _maildict["ecn_ecr"] = _Chng_NR
    
    
    _cfs = root.find("./SW_Versions[@key='Prod_Tooling']")
    _PT_cfs_version = _cfs.attrib["version"]
    _maildict["_PT_cfs_version"] = _PT_cfs_version
    
    _pd_file = root.find("./Overall_Infos[@Col1='kds_pd_file']")
    _PD_Config_zip = _pd_file.attrib["Col2"]
    
    _sw_id_gas = root.find("./Overall_Infos[@Col1='swid_gas']")
    _SWID_gas = _sw_id_gas.attrib["Col2"]
    _maildict["_SWID_gas"] = _SWID_gas
    
    _sw_id_nongas = root.find("./Overall_Infos[@Col1='swid_nongas']")
    _SWID_nongas = _sw_id_nongas.attrib["Col2"]
    _maildict["_SWID_nongas"] = _SWID_nongas
    
    _release_type = root.find("./Overall_Infos[@Col1='purpose']")
    _rel_name = _release_type.attrib["Col2"]
    _maildict["_rel_name"] = _rel_name    

    _sw_ver = root.find("./Overall_Infos[@Col1='sw_full_ver']")
    _rel_version = _sw_ver.attrib["Col2"]
    _maildict["rel_version"] = _rel_version
    _sw_rel_type = _rel_version.split("_")[-1]
    
    _template_dir = root.find("./Overall_Infos[@Col1='template_dir']")
    _tmpl_dir = _template_dir.attrib["Col2"]
    
    
    if "Endunit" in _tmpl_dir and "EU" in _sw_rel_type:
        _maildict["_rel_type"] = "EU" 
    
    elif "Baseunit" in _tmpl_dir and "BU" in _sw_rel_type:
        _maildict["_rel_type"] = "BU"
    
   
    _tar_dir = root.find("./Overall_Infos[@Col1='target_dir']")
    _target = _tar_dir.attrib["Col2"]
    
    _maildict["target"] = _target
    
    _set_dict ={}
    i=0    
    for SET in root.iter("SET_Infos"):    
        _SET_Infos_all = SET.attrib
        _set_dict[i] = _SET_Infos_all
        _PN =_set_dict[i]["part_number"]   
        i+=1
       
         
            
if __name__ == "__main__":
    parser = optparse.OptionParser()
    parser.add_option('-x', '--xml', dest='xml', default=None,
                      help='Master xml')
   
    (options, args) = parser.parse_args()
    # print ("\n options : ", options)
    # print("\n args :", args)
    if not options.xml:
        #input_masterxml = input("\nPlease enter masterxml filename\n")
        print ("\n Please enter the master xml name as input ")
        print ("\n please run the script in the form : ")
        print ("\t update_prod_deliveries_ccs2.py -x <sap_xml_name> ")
        sys.exit(0) 
    else:
        #input_masterxml = options.xml
        input_masterxml = sys.argv[2]
        _sap_xml = input_masterxml
        print ("\n Master xml ", _sap_xml )
        if _sap_xml.startswith("SAP_") == True and _sap_xml.endswith(".xml") == True :
            print ("\n Master xml ", _sap_xml," will be used ")
        else:
            print ("please enter the proper sap master xml file ")
            sys.exit(0)
            
    _curr_Dir = os.getcwd()   
    _xml_File = os.path.join(_curr_Dir,_sap_xml) 
    # print ("\n call parse xml  ..")
    parse_sap_xml(_sap_xml)
    # print("\n set dict : ", _set_dict )
   
    # ob1 = _mail_content(_maildict)
    # ob1._compose_mail()
   
    ob1 = _update_overview_file(_maildict)
    ob1.update_sheet()
    