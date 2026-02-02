# 29.08.2024   | Nisharani C  | Version 16
# 26.02.2025   | Nisharani C  | Version 17 - Change to read Flw4Linux section from Jira
# 28.05.2025   | Nisharani C  | Version 18 - Changes to fill app sw section in xml, if gas/non-gas sw section in jira is left blank
# 24.06.2025   | Abinaya M	  | Version 19 - Sync with latest changes in Set definition tmpl (to add Supplier feed column and change from col name ublox to gnss)
# 25.06.2025   | Nisharani C  | Version 20 - Image name changes for SXM, DTV, Lontium
# 01.07.2025   | Nisharani C  | Version 21 - Including the SW verfication part - to crosscheck the versions from appsw
# 14.07.2025   | Nisharani C  | Version 22 - Merging 
# 21.08.2025   | Nisharani C  | Version 23 - Change to have the SPL include plant name based on info from VM
# 16.09.2025   | Abinaya M    | Version 24 - Change for new column BU in set definition
# 14.10.2025   | Nisharani C  | Version 25 - Set supplier feed as "NA" for baseunit release
# 24.10.2025   | Nisharani C  | Version 26 - Bugfixes with respect to python latest 
# 10.11.2025   | Nisharani    | Version 27 - update release name from Inventory into Set Definition and update aurix subpath into master xml
# 4.12.2025    | Nisharani C  | Version 28 - Changes to include new DBKey ASS_App_Android_Version into the SOS
# 6.01.2026    | Nisharani C  | Version 29 - Bug fix - set Android key as "n.a." for baseunit releases
# 29.01.2026    | Nisharani C  | Version 29 - Bug fix - set SXM as "n.a." for no entry from Jira

import os
import sys
import re
import optparse
import xml.etree.ElementTree as ET
import xml.etree.ElementTree as ET1
import pandas as pd # for dataframes
import openpyxl 
import subprocess
import shutil
#import datetime
from jira import JIRA
from datetime import date 
from shutil import copy
from bs4 import BeautifulSoup
from optparse import OptionParser
from datetime import datetime as dt


global _set 
global _cts_tag
global _xsl_content
global _curr_Dir

#set definition cell values
global _set_def_cell
global _rel_type_cell
global _rel_version_cell
global _ecn_ecr_cell
global _PD_conf_cell
global _Prod_tool_cell
global _CTS_conf_cell 
global _CTS_ver_cell
global _AppSw_ver_cell
global _SWID_gas_cell
global _SWID_nongas_cell
global _GAS_ver_cell
global _NonGAS_ver_cell
global _Aurix_ver_cell
global _Aurix_sub_path_cell
global _TestManager_cell
global _GNSS_cell
global _jira_ID_cell
global _del_resp_cell
global _pm_cell
global _sw_pm_cell
global _hw_pm_cell
global _del_resp_cell

global _rel_OEM
global _rel_TaskName



_set_def_cell = "B1"
_rel_type_cell = "B2"
_rel_version_cell = "B3"
_ecn_ecr_cell = "B4"
_ecr_cell = "C4"
_PD_conf_cell = "B5"
_Prod_tool_cell = "B6"
_CTS_conf_cell = "B7"
_CTS_ver_cell = "B8"
_AppSw_ver_cell = "B9"
_SWID_gas_cell = "C10"
_SWID_nongas_cell = "C11"
_GAS_ver_cell = "B10"
_NonGAS_ver_cell = "B11"
_Aurix_ver_cell = "B12"
_Aurix_sub_path_cell = "C12"
_TestManager_cell = "B13"
_GNSS_cell = "B14"
_SXM_cell = "B15"
_SXM_fw_cell = "C15"
_DTV_cell = "B16"
_Lontium_cell = "B17"
_Inventory_cell = "C3"
_jira_ID_cell = "D3"
_pm_cell = "B19"
_sw_pm_cell = "B20"
_hw_pm_cell = "B21"
_del_resp_cell = "B22"

#_set_info_start_row = 23

_set = "P2_G_HD_SA8155_S128_4"
_cts_ver = ""
_cts_binary = ""
_cts_link = ""


def read_jira_info (_desc_File,_count,_rel_OEM):
    print("\nat read_jira_info..")
    global _cts_ver
    global _cts_binary
    global _cts_link
    global _cts_config_ver
    global _cts_config_zip
    global _cts_config_link
    global _gas_appsw_ver
    global _gas_appsw
    global _SWID_gas
    global _gas_appsw_zip
    global _gas_appsw_link
    global _nongas_appsw_ver
    global _nongas_appsw
    global _SWID_nongas
    global _nongas_appsw_zip
    global _nongas_appsw_link
    global _aurixsw_ver
    global _aurix_dnl
    global _aurix_support_dnl
    global _aurix_srec
    global _aurix_link
    global _aurix_sub_path
    global _test_manager_ver
    global _test_manager_zip
    global _test_manager_link
    global _ublox_ver
    global _ublox_bu_bin
    global _ublox_eu_bin
    global _PD_Config_ver
    global _PD_Config_zip
    global _PD_Config_link
    global _prod_tooling_ver
    global _prod_tooling_zip
    global _prod_tooling_tar
    global _prod_tooling_link
    global _appsw_ver
    global _sxm_ver
    global _dtv_ver
    global _dtv_fw_zip
    global _dtv_link
    global _sxm_fw_ver
    global _sxm_fw_zip
    global _sxm_link
    global _cryptoBosch		
    global _cryptoCust		
    global _cryptosignaurix
    global _Flw4Linux_ver
    global _dtv_BL_ver
    global _dtv_HW_ver
    global _lt_ver          
    global _lt_fw_hex 
    global _lt_fw_zip 
    global _inventory_versions_dict

    _inventory_versions_dict = {}
    _Inv_File = _desc_File
    _line_count = _count
    #print ("\n Inventory file, count :", _Inv_File,_line_count )
    fj = open(_jira_File, "w")
    f1 = open(_Inv_File, "r")
    _all_lines = f1.readlines()
    f1.close()
    _line_no = 0
    #print("****************************************************************************************************************************************")
    while _line_no < _line_count:
        #print ("_line_count:",_line_count )
        _line =  _all_lines[_line_no]
        
        #checklist
        if ( len( re.findall(r'\|\|\sFor\sI.No.\s\|\|\sChecklist' , _line )) > 0):
            _line_no = _line_no + 4
            _line =  _all_lines[_line_no]
            _cryptoBosch_line = _all_lines[_line_no]
            if "(update) dev/prd " in _cryptoBosch_line: 
                _cryptoBosch = (_cryptoBosch_line .split("|")[-2]).strip()
                #print("\n _cryptoBosch : ", _cryptoBosch)
                _cryptoBosch = (_cryptoBosch.lower()).strip()
                _cryptoBosch = _cryptoBosch.strip()
                #print("\n _cryptoBosch : ", _cryptoBosch)
                
                
                if _cryptoBosch == "dev" or _cryptoBosch == "prd":
                    _cryptoBosch_entry = "_cryptoBosch :" + _cryptoBosch
                    fj.write(_cryptoBosch_entry)
                else:
                    print("\n _cryptoBosch value not dev  or prd, please check the jira and update ")
                    fj.write("_cryptoBosch :not available " )
                    
                   
                #print ("_cryptoBosch_line", _cryptoBosch)
            else:
                print('\n Crypto Bosch Environment string "(update) dev/prd" not found , please check the jira and update ')
            
            _line_no = _line_no + 1
            _line =  _all_lines[_line_no]
            #print (_all_lines[_line_no])
            #_cryptoCustomer
            _cryptoCust_line = _all_lines[_line_no]
            if "Crypto Environment (vNext)" in _cryptoCust_line: 
                _cryptoCust = _cryptoCust_line .split("|")[-2].strip()
                
                if _cryptoCust == "sit" or _cryptoCust == "stg" or _cryptoCust == "prd":
                    _cryptoCust_entry = "\n" + "_cryptoCust :" + _cryptoCust
                    fj.write(_cryptoCust_entry)
                else:
                    print("\n _cryptoCust value not sit or stg or prd, please check the jira and update ")
                    fj.write("_cryptoCust :not available " )
                    
                    
                #print ("_cryptoCust_line", _cryptoCust)
            else:
                print('\n Crypto Customer Environment string "Crypto Environment (vNext)" not found , please check the jira and update ')
            
            _line_no = _line_no + 5
            _line =  _all_lines[_line_no]
            #print (_all_lines[_line_no])
            #_crypto environment aurix
            _cryptosignaurix_line = _all_lines[_line_no]
            if "Crypto Environment dev/kms" in _cryptosignaurix_line: 
                _cryptosignaurix = _cryptosignaurix_line .split("|")[-2].strip()
                
                if _cryptosignaurix == "dev" or _cryptosignaurix == "kms":
                    _cryptosignaurix_entry = "\n" + "_cryptosignaurix :" + _cryptosignaurix
                    fj.write(_cryptosignaurix_entry)
                else:
                    print("\n _cryptoCust value not sit or stg or prd, please check the jira and update ")
                    fj.write("_cryptosignaurix :not available " )
                    
                    
                #print ("_cryptosignaurix_line", _cryptosignaurix)
            else:
                print('\n Crypto Environment string "Crypto Environment dev/kms " not found , please check the jira and update ')
         
        elif ( len( re.findall(r'"vm_version":' , _line )) > 0):
            # _line_no = _line_no + 4
            # _line =  _all_lines[_line_no]
            _vm_line = _all_lines[_line_no]
            _vm_str = '"vm_version":'
            if _vm_str in _vm_line:
                #print ("\n _vm_line :", _vm_line) 
                #print("****************************************************************************************************************************************")
                _vm_ver = re.search(r'\"\S+\s\S+\d\S', _line).group()
                _vm_ver =(_vm_ver.split(": ")[-1]).replace('"','')
                _vm_entry = "\n" + "_vm_ver :" + _vm_ver  + "\n"
                #print ("\n _vm_entry :", _vm_entry)
                fj.write(_vm_entry)
                _inventory_versions_dict["_vm_ver"] = _vm_ver

                
        elif ( len( re.findall(r'^Reference ' , _line )) > 0):  
            _line = _line.replace('\\r','') 
            _ref_entry = _line
            fj.write(_ref_entry)
			
    
        #|| *#1a*                      || *CTS SW* *(QC SoC)* 
        #_cts_sec = re.findall(r'\WCTS\sSW\W' , _line )
        elif ( len( re.findall(r'\W\*#1a\*\s+\|\|\s\*CTS\sSW\*' , _line )) > 0):
            _line_no = _line_no + 2
            _line =  _all_lines[_line_no]
            _cts_ver = re.search(r'(?i)\*Version\sID1\*\s+\W.prj_ccs2\S+', _line)
            if _cts_ver != None:        #at here we confirm it is cts section
                _cts_ver = re.search(r'(?i)prj_ccs2\S+', _line).group()
                _cts_ver = _cts_ver.strip()
                #print("****************************************************************************************************************************************")

                
                _cts_ver_entry = "\n" + "_cts_ver :" + _cts_ver
                fj.write(_cts_ver_entry)
                _line_no = _line_no + 5
                _line =  _all_lines[_line_no]
                #_cts_binary = re.search(r'\*Version\sID2\*\s+\W.ccs2\S+', _line)
                _cts_binary = re.search(r'\*Name\sof\sArtifact\*\s+\W.ccs2\S+', _line)
                #print ("content:",_line )
                if _cts_binary != None:
                    _cts_binary = re.search(r'ccs2_\S+', _line).group()
                    _cts_binary = _cts_binary.strip()
                    _cts_binary_entry = "\n" + "_cts_binary :" + _cts_binary
                    fj.write(_cts_binary_entry)
                else:
                    print("\n could not find cts binary ")
                    fj.write("_cts_binary :not available ")
                _line_no = _line_no + 1
                _line =  _all_lines[_line_no]
                _cts_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                #print ("content:",_line )
                if _cts_link != None:
                    _cts_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                    _cts_link = _cts_link.strip()
                    if not _cts_link.endswith('/'):
                        _cts_link = _cts_link + "/"
                    _cts_link_entry = "\n" + "_cts_link :" + _cts_link
                    fj.write(_cts_link_entry)
                else:
                    print ("cts link not found ")
                    fj.write("_cts_link :not available " )
                #print("****************************************************************************************************************************************")
            else:
                print("\n cts version not found ")
                fj.write("_cts_version :not available " )


        #CTS Config
        elif ( len( re.findall(r'\W#1b\*\s+\|\|' , _line )) > 0):
                _line_no = _line_no + 3
                _line =  _all_lines[_line_no]
                _cts_config_ver = re.search(r'\*Version\sID2\*\s+\W\s[v|V]?\d.\d.\d+', _line)
                #print ("_cts_config_ver:" ,_cts_config_ver)
                if _cts_config_ver != None:
                    _cts_config_ver = re.search(r'[v|V]?\d.\d.\d+', _line).group()
                    _cts_config_ver = _cts_config_ver.upper()
                    _cts_config_ver_entry = "\n" + "_cts_config_ver :" + _cts_config_ver
                    fj.write(_cts_config_ver_entry)

                    _line_no = _line_no + 4    
                    _line =  _all_lines[_line_no]
                    _cts_config_zip = re.search(r'\*Name\sof\sArtifact\*\s+\W\s.\S+\.', _line)
                    #print ("_cts_config_zip:",_cts_config_zip )
                    if _cts_config_zip != None:
                        _cts_config_zip = re.search(r'variantInfo\S*\.zip', _line).group()
                        _cts_config_zip = _cts_config_zip.strip()
                        _cts_config_zip_entry = "\n" + "_cts_config_zip :" + _cts_config_zip
                        fj.write(_cts_config_zip_entry)
                    else:
                        print("\n could not find _cts_config_zip ")
                        fj.write(" _cts_config_zip :not available " )
                    _line_no = _line_no + 1   
                    _line =  _all_lines[_line_no]

                    _cts_config_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                    if _cts_config_link != None:
                        _cts_config_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _cts_config_link = _cts_config_link.strip()
                        if not _cts_config_link.endswith('/'):
                            _cts_config_link = _cts_config_link + "/"
                        #print (" _cts_config_link :",_cts_config_link )
                        _cts_config_link_entry = "\n" + "_cts_config_link :" + _cts_config_link
                        fj.write(_cts_config_link_entry)
                    else:
                        print ("_cts_config_link not found ")
                        fj.write(" _cts_config_link :",_cts_config_link )
                    #print("****************************************************************************************************************************************")

                else:
                    print ("cts config version not found, further checks excluded ")
                    fj.write("_cts_config_ver: not available")

        #App SW GAS
        #elif( len( re.findall(r'\W##2a\*\s+\|\|' , _line )) > 0):
        elif (len(re.findall(r'\W#2a\*\s+\|\|\s\*APP\sSW*' , _line )) > 0 ):
            _gas_appsw_sec = re.findall(r'\W#2a\*\s+\|\|\s\*APP\sSW\*\s\*\WQC\sSoC\W\*\s\*GAS\*' , _line )
            #print ("\n _gas_appsw_sec ",_gas_appsw_sec )
            if (len(_gas_appsw_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                if _rel_OEM == "MMC":
                    #print ("MMC Release")
                    _gas_appsw_ver = re.search(r'\*Version\sID1\*\s+\W.NGO\S+', _line)
                    if _gas_appsw_ver != None:        
                        _gas_appsw_ver = re.search(r'NGO\S+', _line).group()
                elif _rel_OEM == "Nissan":
                    #print ("Nissan Release")
                    _gas_appsw_ver = re.search(r'\*Version\sID1\*\s+\W.ISH\S+', _line)
                    if _gas_appsw_ver != None:        
                        _gas_appsw_ver = re.search(r'ISH\S+', _line).group()
                #print ("_gas_appsw_ver :",_gas_appsw_ver )
                if _gas_appsw_ver != None:        
                    _gas_appsw_ver = _gas_appsw_ver.strip()
                    _gas_appsw_ver_entry = "\n" + "_gas_appsw_ver :" + _gas_appsw_ver
                    _inventory_versions_dict["_gas_appsw_ver"] = _gas_appsw_ver
                    fj.write(_gas_appsw_ver_entry)
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    if _rel_OEM == "MMC":
                        _gas_appsw = re.search(r'\*Version\sID2\*\s+\W.ccs2mmc_\S+', _line)
                        if _gas_appsw != None: 
                            _gas_appsw = re.search(r'ccs2mmc_\S+', _line).group()
                    elif _rel_OEM == "Nissan":
                        _gas_appsw = re.search(r'\*Version\sID2\*\s+\W.ccs2_\S+', _line)
                        if _gas_appsw != None: 
                            _gas_appsw = re.search(r'ccs2_\S+', _line).group()
                    if _gas_appsw != None:		
                        _gas_appsw = _gas_appsw.strip()  
                        _gas_appsw_entry = "\n" + "_gas_appsw :" + _gas_appsw
                        fj.write(_gas_appsw_entry)
                        #print (" App SW version mentioned in GAS section:", _gas_appsw )
                    else:
                        _gas_appsw = "not available"
                        fj.write("_gas_appsw :", _gas_appsw)
                        print ( "App SW version not found in GAS App SW section")
                    #inclusion of SW-ID 
                    #exit() 			
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _SWID_g = re.search(r'\*Version\sID3\*', _line)
                    if _SWID_g != None: 
                        _SWID_gas = re.search(r'\d\w+', _line).group()
                        _SWID_gas =  _SWID_gas.strip()
                        #print ("GAS SW ID: ", _SWID_gas )
                        _SWID_gas_entry = "\n" + "_SWID_gas :" + _SWID_gas
                        fj.write(_SWID_gas_entry)
                        _line_no = _line_no + 5
                    
                    else:
                        _SWID_gas = "not available"
                        fj.write("\n GAS SW ID: not available" )
                        print ("\n SW ID from GAS section could not be found ")
                        _line_no = _line_no + 4
                    
                    
                    #_line_no = _line_no + 5          #only this line is needed when the SWID row inclusion is followed in all jira
                    
                    _line =  _all_lines[_line_no]
                    _gas_appsw_zip = re.search(r'\*Name\sof\sArtifact\*\s+\W\s.\S+\.', _line)
                    #print ("_gas_appsw_zip:",_gas_appsw_zip )
                    if _gas_appsw_zip != None:
                        if _rel_OEM == "MMC":
                            _gas_appsw_zip = re.search(r'NGO\S*\.zip', _line).group()
                        elif _rel_OEM == "Nissan":
                            _gas_appsw_zip = re.search(r'ISH\S*\.zip', _line).group()
                        
                        _gas_appsw_zip = _gas_appsw_zip.strip()
                        _gas_appsw_zip_entry = "\n" + "_gas_appsw_zip :" + _gas_appsw_zip
                        fj.write(_gas_appsw_zip_entry)
                    else:
                        _gas_appsw_zip = "not available"
                        fj.write("\n _gas_appsw_zip :not available" )
                        print("\n could not find _gas_appsw_zip ")

                    _line_no = _line_no + 1   
                    _line =  _all_lines[_line_no]
                    _gas_appsw_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)    
                    if _gas_appsw_link != None:
                        _gas_appsw_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _gas_appsw_link = _gas_appsw_link.strip()
                        if not _gas_appsw_link.endswith('/'):
                            _gas_appsw_link = _gas_appsw_link + "/"
                        _gas_appsw_link_entry = "\n" + "_gas_appsw_link :" + _gas_appsw_link
                        fj.write(_gas_appsw_link_entry)
                        #print ("_gas_appsw_link :",_gas_appsw_link )
                    else:
                        _gas_appsw_link = "not available"
                        print ("_gas_appsw_link not found ")
                        fj.write("\n _gas_appsw_link :not available" )
                    #print("****************************************************************************************************************************************")

                else:
                    print("GAS App SW version not found ")
                    _gas_appsw_ver = "not available"
                    _gas_appsw = "not available"
                    _SWID_gas = "not available"
                    _gas_appsw_zip = "not available"
                    _gas_appsw_link = "not available"
                    fj.write("\n _gas_appsw_ver :not available")                    
                    fj.write("\n _gas_appsw :not available " )
                    fj.write("\n _SWID_gas :not available " )
                    fj.write("\n _gas_appsw_zip :not available " )
                    fj.write("\n _gas_appsw_link :not available " )  
                   

        #App SW NON GAS
        elif (len(re.findall(r'\W#2b\*\s+\|\|\s\*APP\sSW*' , _line )) > 0 ):
            _nongas_appsw_sec = re.findall(r'\W#2b\*\s+\|\|\s\*APP\sSW\*\s\*\WQC\sSoC\W\*\s\*\s?NON\sGAS\*', _line )
            if (len(_nongas_appsw_sec) > 0):
                # print ("\n _nongas_appsw_sec ",_nongas_appsw_sec )
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                if _rel_OEM == "MMC":
                    #print ("MMC Release")
                    _nongas_appsw_ver = re.search(r'\*Version\sID1\*\s+\W.HND\S+', _line)
                    #print ("_nongas_appsw_ver :",_nongas_appsw_ver )
                    if _nongas_appsw_ver != None:        
                        _nongas_appsw_ver = re.search(r'HND\S+', _line).group()
                elif _rel_OEM == "Nissan":
                    #print ("Nissan Release")
                    _nongas_appsw_ver = re.search(r'\*Version\sID1\*\s+\W.KUM\S+', _line)
                    if _nongas_appsw_ver != None:        
                        _nongas_appsw_ver = re.search(r'KUM\S+', _line).group()
                #_nongas_appsw_ver = re.search(r'\*Version\sID1\*\s+\W.HND\S+', _line)
                if _nongas_appsw_ver != None:       
                    #_nongas_appsw_ver = re.search(r'HND\S+', _line).group()
                    _nongas_appsw_ver = _nongas_appsw_ver.strip()
                    _nongas_appsw_ver_entry = "\n" + "_nongas_appsw_ver :" + _nongas_appsw_ver
                    fj.write(_nongas_appsw_ver_entry)
                    _inventory_versions_dict["_nongas_appsw_ver"] = _nongas_appsw_ver
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    if _rel_OEM == "MMC":
                        _nongas_appsw = re.search(r'\*Version\sID2\*\s+\W.ccs2mmc_\S+', _line)
                        if _nongas_appsw != None: 
                            _nongas_appsw = re.search(r'ccs2mmc_\S+', _line).group()
                    elif _rel_OEM == "Nissan":
                        _nongas_appsw = re.search(r'\*Version\sID2\*\s+\W.ccs2_\S+', _line)
                        if _nongas_appsw != None: 
                            _nongas_appsw = re.search(r'ccs2_\S+', _line).group()
                    #_nongas_appsw = re.search(r'\*Version\sID2\*\s+\W.ccs2mmc_\S+', _line)
                    if _nongas_appsw != None: 
                        _nongas_appsw = _nongas_appsw.strip()
                        _nongas_appsw_entry = "\n" + "_nongas_appsw :" + _nongas_appsw
                        fj.write(_nongas_appsw_entry)                        
                        #print ("App SW version mentioned in Non GAS section:", _nongas_appsw )
                    else:
                        _nongas_appsw = "not available"
                        fj.write("\n _nongas_appsw:", _nongas_appsw)
                        print ("App SW version not found in Non GAS App SW section")
                        
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _SWID_ng = re.search(r'\*Version\sID3\*', _line)
                    if _SWID_ng != None: 
                        _SWID_nongas = re.search(r'\d\w+', _line).group()
                        _SWID_nongas =  _SWID_nongas.strip()
                        #print ("\n NON GAS SW ID: ", _SWID_nongas )
                        _SWID_nongas_entry = "\n" + "_SWID_nongas :" + _SWID_nongas
                        fj.write(_SWID_nongas_entry)
                        _line_no = _line_no + 5
                    
                    else:
                        _SWID_nongas = "not available"
                        fj.write("\n NON GAS SW ID: not available" )
                        print ("\n SW ID from NON GAS section could not be found ")
                        _line_no = _line_no + 4
                    
                    
                    #_line_no = _line_no + 5
                    _line =  _all_lines[_line_no]
                    _nongas_appsw_zip = re.search(r'\*Name\sof\sArtifact\*\s+\W\s.\S+\.', _line)
                    #print ("_nongas_appsw_zip:",_nongas_appsw_zip )
                    if _nongas_appsw_zip != None:
                        if _rel_OEM == "MMC":
                            _nongas_appsw_zip = re.search(r'HND\S*\.zip', _line).group()
                        elif _rel_OEM == "Nissan":
                            _nongas_appsw_zip = re.search(r'KUM\S*\.zip', _line).group()
                        #_nongas_appsw_zip = re.search(r'HND\S*\.zip', _line).group()
                        _nongas_appsw_zip = _nongas_appsw_zip.strip()
                        _nongas_appsw_zip_entry = "\n" + "_nongas_appsw_zip :" + _nongas_appsw_zip
                        fj.write(_nongas_appsw_zip_entry)
                    else:
                        _nongas_appsw_zip= "not available"
                        fj.write("\n _nongas_appsw_zip :not available " )
                        print("\n could not find _nongas_appsw_zip ")

                    _line_no = _line_no + 1   
                    _line =  _all_lines[_line_no]
                    _nongas_appsw_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                    if _nongas_appsw_link != None:
                        _nongas_appsw_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _nongas_appsw_link = _nongas_appsw_link.strip()
                        if not _nongas_appsw_link.endswith('/'):
                            _nongas_appsw_link = _nongas_appsw_link + "/"
                        _nongas_appsw_link_entry = "\n" + "_nongas_appsw_link :" + _nongas_appsw_link
                        fj.write(_nongas_appsw_link_entry)    
                        #print ("_nongas_appsw_link :",_nongas_appsw_link )
                    else:
                        _nongas_appsw_link = "not available"
                        print ("\n _nongas_appsw_link: not available" )
                        fj.write("\n _nongas_appsw_link : not available")
                    #print("****************************************************************************************************************************************")
                else:
                    print("Non GAS App SW version not found ")
                    _nongas_appsw_ver = "not available" 
                    _nongas_appsw = "not available"
                    _SWID_nongas = "not available"
                    _nongas_appsw_zip= "not available"
                    _nongas_appsw_link = "not available" 
                    fj.write("\n _nongas_appsw_ver :not available " )                     
                    fj.write("\n _nongas_appsw :not available " )
                    fj.write("\n _SWID_nongas :not available " )
                    fj.write("\n _nongas_appsw_zip :not available " )
                    fj.write("\n _nongas_appsw_link :not available " )                    


            else:
                print(" NON GAS App SW section not found ")


        #Aurix
        elif (len(re.findall(r'\W#3\*\s+\|\|\s\*AUTOSAR\sSW*' , _line )) > 0 ):
            _aurixsw_sec = re.findall(r'\W#3\*\s+\|\|\s\**AUTOSAR\sSW\*\s\*\WAurix\W\*' , _line )
            #print ("\n _aurixsw_sec ",_aurixsw_sec )
            if (len(_aurixsw_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                _aurixsw_ver = re.search(r'\*Version\sID1\*\s+\W.\d\d\.\d*\S+', _line)
                if _aurixsw_ver != None:        #at here we confirm it is cts section
                    _aurixsw_ver = re.search(r'\d\d\S+', _line).group()
                    _aurixsw_ver = _aurixsw_ver.strip()
                    _aurixsw_ver_entry = "\n" + "_aurixsw_ver :" + _aurixsw_ver
                    fj.write(_aurixsw_ver_entry) 
                    _inventory_versions_dict["_aurixsw_ver"] = _aurixsw_ver                  
                    #print ("_aurixsw_ver :",_aurixsw_ver )

                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]          
                    _aurix_dnl = re.search(r'\*Name\sof\sArtifact\sIa\*\s+\W\s.\S+\:*', _line)
                    if _aurix_dnl != None:
                        _aurix_dnl = re.search(r'\w+.dnl', _line).group()
                        _aurix_dnl = _aurix_dnl.strip()
                        _aurix_dnl_entry = "\n" + "_aurix_dnl :" + _aurix_dnl
                        fj.write(_aurix_dnl_entry)    
                        #print ("_aurix_dnl :",_aurix_dnl )
                    else:
                        print(" aurix dnl file not found ")
                        fj.write("\n _aurix_dnl :not available " )

                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _aurix_support_dnl = re.search(r'\*Name\sof\sArtifact\sIb\*\s+\W\s.\S+\:*', _line)
                    if _aurix_support_dnl != None:
                        _aurix_support_dnl = re.search(r'\w+_DisableHsmOnly.dnl', _line).group()
                        _aurix_support_dnl = _aurix_support_dnl.strip()
                        _aurix_support_dnl_entry = "\n" + "_aurix_support_dnl :" + _aurix_support_dnl
                        fj.write(_aurix_support_dnl_entry)    
                        #print ("_aurix_support_dnl :",_aurix_support_dnl )
                    else:
                        print("\n could not find _aurix_support_dnl ")
                        fj.write("\n _aurix_support_dnl :not available " )

                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _aurix_srec = re.search(r'\*Name\sof\sArtifact\sII\*\s+\W\s.\S+\:*', _line)
                    if _aurix_srec != None:
                        _aurix_srec = re.search(r'\w+.SREC', _line).group()
                        _aurix_srec = _aurix_srec.strip()
                        _aurix_srec_entry = "\n" + "_aurix_srec :" + _aurix_srec
                        fj.write(_aurix_srec_entry)    
                        #print (" aurix_srec :",_aurix_srec )
                    else: 
                        print ("\n could not find aurix srec file ")
                        fj.write("\n aurix_srec :not available " )

                    _line_no = _line_no + 1   
                    _line =  _all_lines[_line_no]
                    _aurix_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                    if _aurix_link != None:
                        _aurix_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _aurix_link = _aurix_link.strip()
                        if not _aurix_link.endswith('/'):
                            _aurix_link = _aurix_link + "/"
                        _aurix_sub_path = _aurix_link.split("Output/vector/")[1]
                        _aurix_link_entry = "\n" + "_aurix_link :" + _aurix_link
                        _aurix_sub_path_entry = "\n" + "_aurix_sub_path :" + _aurix_sub_path
                        fj.write(_aurix_link_entry)    
                        #print ("_aurix_link :",_aurix_link )
                        #print ("_aurix_sub_path :",_aurix_sub_path )
                    else:
                        print ("_aurix_link not found ")
                        fj.write("\n _aurix_link :not available " )
                        fj.write("\n _aurix_sub_path :not available " )
                    #print("****************************************************************************************************************************************")
                else:
                    print(" Aurix SW version not found ")
                    fj.write("\n _aurixsw_ver :not available " )


        #Test Manager || *#5*                       || *Testmanager* *(Aurix)*  
        elif (len(re.findall(r'\W#5a\*\s+\|\|\s\*Testmanager\*\s\*\S+' , _line )) > 0 ):
            _test_manager_sec = re.findall(r'\W#5a\*\s+\|\|\s\*Testmanager\*\s\*\S+' , _line )

            if (len(_test_manager_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                #|  *Version ID1*               | TM_AURIX_83.0V21_TM_INT 
                _test_manager_ver = re.search(r'\*Version\sID1*', _line)
                if _test_manager_ver != None:        #at here we confirm it is cts section
                    _test_manager_ver = re.search(r'\d\d\.\d[v|V]\d\d', _line).group()
                    _test_manager_ver = _test_manager_ver.strip()
                    _test_manager_ver_entry = "\n" + "_test_manager_ver :" + _test_manager_ver
                    fj.write(_test_manager_ver_entry)    
                    #print ("_test_manager_ver :",_test_manager_ver )

                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]   
                    _test_manager_zip = re.search(r'\*Name\sof\sArtifact\*\s+\W\s.\S+\.', _line)
                    #print ("_test_manager_zip:",_test_manager_zip )
                    if _test_manager_zip != None:
                        _test_manager_zip = re.search(r'TM_AURIX_\S*\.zip', _line).group()
                        _test_manager_zip = _test_manager_zip.strip()
                        _test_manager_zip_entry = "\n" + "_test_manager_zip :" + _test_manager_zip
                        fj.write(_test_manager_zip_entry)    
                    else:
                        print("\n could not find _test_manager_zip ")
                        fj.write("\n _test_manager_zip :not available " )

                    _line_no = _line_no + 1   
                    _line =  _all_lines[_line_no]
                    _test_manager_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                    if _test_manager_link != None:
                        _test_manager_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _test_manager_link = _test_manager_link.strip()
                        _test_manager_link_entry = "\n" + "_test_manager_link :" + _test_manager_link
                        fj.write(_test_manager_link_entry)    
                        #print ("_test_manager_link :",_test_manager_link )
                    else:
                        print ("_test_manager_link not found ")
                        fj.write("\n _test_manager_link :",_test_manager_link )
                    #print("****************************************************************************************************************************************")

                else:
                    print(" _test_manager_ver not found ")
                    fj.write("\n _test_manager_ver :not available " )

        #Flw4Linux inclusion
        elif (len(re.findall(r'\W#5b\*\s+\|\|\s\*Flw4Linux\*' , _line )) > 0 ):
            _Flw4Linux_sec = re.findall(r'\W#5b\*\s+\|\|\s\*Flw4Linux\*' , _line )
            #print ("_Flw4Linux_sec ",_Flw4Linux_sec )
            
            if (len(_Flw4Linux_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                _Flw4Linux_ver = re.search(r'\*Version\sID1*', _line)
                if _Flw4Linux_ver != None:        
                    _Flw4Linux_ver = re.search(r'\d\W\w+\s\S+', _line).group()
                    _Flw4Linux_ver = _Flw4Linux_ver.strip()
                    _Flw4Linux_ver_entry = "\n" + "_Flw4Linux_ver :" + _Flw4Linux_ver
                    fj.write(_Flw4Linux_ver_entry)    
                    #print ("_Flw4Linux_ver :",_Flw4Linux_ver )
                    #print("****************************************************************************************************************************************")
                else:
                    print(" _Flw4Linux_ver not found ")
                    fj.write("\n _Flw4Linux_ver :not available " )

        #Ublox || *#6*                       || *GNSS SW* *(UBlox)*   
        elif (len(re.findall(r'\W#6\*\s+\|\|\s\*GNSS\sSW\*\s\*\S+' , _line )) > 0 ):
            _ublox_sec = re.findall(r'\W#6\*\s+\|\|\s\*GNSS\sSW\*\s\*\WUBlox*' , _line )
            if (len(_ublox_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]

                #|  *Version ID1*               | TM_AURIX_83.0V21_TM_INT 
                _ublox_ver = re.search(r'\*Version\sID1*', _line)
                if _ublox_ver != None:       
                    _ublox_ver = re.search(r'\d.\d+', _line).group()
                    _ublox_ver = _ublox_ver.strip()
                    _ublox_ver_entry = "\n" + "_ublox_ver :" + _ublox_ver
                    fj.write(_ublox_ver_entry)   
                    _inventory_versions_dict["_ublox_ver"] = _ublox_ver
                    #print (" _ublox_ver :",_ublox_ver )   

                    #  *Name of Artifact I*        | Re-flash  image: JU_EXT_404.907640e53126add7de4708d754b575ef.bin                                                                          |\r
                    #  *Name of Artifact II*       | Pre-flash Image: JU_EXT_404.80ff71bfcf34c8929b04aa9d16d9ede0_shifted_0x48.bin 
                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]
                    _ublox_eu_bin = re.search(r'\*Name\sof\sArtifact\sI\*\s+\|\sRe-flash*', _line)
                    if _ublox_eu_bin != None:
                        _ublox_eu_bin = re.search(r'\w+\d+\.\w+\.bin', _line).group()
                        _ublox_eu_bin = _ublox_eu_bin.strip()
                        _ublox_eu_bin_entry = "\n" + "_ublox_eu_bin :" + _ublox_eu_bin
                        fj.write(_ublox_eu_bin_entry)    
                        #print (" _ublox_eu_bin :",_ublox_eu_bin )
                    else:
                        print("\n could not find _ublox_eu_bin ")
                        fj.write("\n _ublox_eu_bin :not available " )

                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _ublox_bu_bin = re.search(r'\*Name\sof\sArtifact\sII\*\s+\|\sPre-flash*', _line)
                    #print ("_ublox_bu_bin:",_ublox_bu_bin )
                    if _ublox_bu_bin != None:
                        _ublox_bu_bin = re.search(r'\w+\d+\.\w+\.bin', _line).group()
                        _ublox_bu_bin = _ublox_bu_bin.strip()
                        _ublox_bu_bin_entry = "\n" + "_ublox_bu_bin :" + _ublox_bu_bin
                        fj.write(_ublox_bu_bin_entry)    
                    else:
                        print("\n could not find _ublox_bu_bin ") 
                        fj.write("\n _ublox_bu_bin :not available " )
                else:
                    print(" _ublox_ver not found ") 
                    fj.write("\n _ublox_ver: not available")
                #print("****************************************************************************************************************************************")

        #PD Config     || *#7*                       || *Config* *(PD)*   
        elif (len(re.findall(r'\W#7\*\s+\|\|\s\*Config\*\s\*\S+' , _line )) > 0 ):
            _PD_Config_sec = re.findall(r'\W#7\*\s+\|\|\s\*Config\*\s\*\WPD\W', _line )
            #print ("\n _PD_Config_sec ",_PD_Config_sec )
            if (len(_PD_Config_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                #|  *Version ID1*               | V56 
                _PD_Config_ver = re.search(r'\*Version\sID1*', _line)
                if _PD_Config_ver != None:       
                    _PD_Config_ver = re.search(r'[v|V].\d+', _line).group()
                    _PD_Config_ver = _PD_Config_ver.strip()
                    #print (" _PD_Config_ver :",_PD_Config_ver )
                    _PD_Config_ver = _PD_Config_ver.upper()
                    _PD_Config_ver_entry = "\n" + "_PD_Config_ver :" + _PD_Config_ver
                    fj.write(_PD_Config_ver_entry)    


                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]
                    _PD_Config_zip = re.search(r'\*Name\sof\sArtifact\sI\*\s+\|\scalibContainer_*', _line)
                    #print ("_PD_Config_zip:",_PD_Config_zip )
                    if _PD_Config_zip != None:
                        _PD_Config_zip = re.search(r'\S+.zip', _line).group()
                        _PD_Config_zip = _PD_Config_zip.strip()
                        _PD_Config_zip_entry = "\n" + "_PD_Config_zip :" + _PD_Config_zip
                        fj.write(_PD_Config_zip_entry)    
                    else:
                        print("\n could not find _PD_Config_zip ")
                        fj.write("\n _PD_Config_zip :not available " )

                    _line_no = _line_no + 2   
                    _line =  _all_lines[_line_no]
                    _PD_Config_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://sites.inside-share3.bosch.com\S+', _line)
                    if _PD_Config_link != None:
                        _PD_Config_link = re.search(r'https?://sites.inside-share3.bosch.com\S+', _line).group()
                        _PD_Config_link = _PD_Config_link.strip()
                        _PD_Config_link_entry = "\n" + "_PD_Config_link :" + _PD_Config_link
                        fj.write(_PD_Config_link_entry)    
                        #print (" _PD_Config_link :",_PD_Config_link )
                    else:
                        print ("_PD_Config_link not found ")
                        fj.write("\n _PD_Config_link :not available " )
                    #print("****************************************************************************************************************************************")
                else:
                    print(" _PD_Config_ver not found ") 
                    fj.write("\n _PD_Config_ver :not available " )

        #PD Config     || *#8*                       || *Production Tooling*      
        elif (len(re.findall(r'\W#8\*\s+\|\|\s\*Production\sTooling\s\S+' , _line )) > 0 ):
            _prod_tooling_sec = re.findall(r'\W#8\*\s+\|\|\s\*Production\sTooling\s\WCFS\W', _line )
            #print ("\n _prod_tooling_sec ",_prod_tooling_sec )
            if (len(_prod_tooling_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                #|  *Version ID1*               | v2.9  
                _prod_tooling_ver = re.search(r'\*Version\sID1*', _line)
                if _prod_tooling_ver != None:        
                    _prod_tooling_ver = re.search(r'[v|V]\d\S+', _line).group()
                    _prod_tooling_ver = _prod_tooling_ver.strip()
                    _prod_tooling_ver = _prod_tooling_ver.upper()
                    _prod_tooling_ver_entry = "\n" + "_prod_tooling_ver :" + _prod_tooling_ver
                    fj.write(_prod_tooling_ver_entry)    
                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]
                    _prod_tooling_zip = re.search(r'\*Name\sof\sArtifact\sI\*\s+\|\s\S+', _line)
                    #print ("_prod_tooling_zip:",_prod_tooling_zip )
                    if _prod_tooling_zip != None:
                        _prod_tooling_zip = re.search(r'\S+.zip', _line)
                        
                        if _prod_tooling_zip is not None:
                            _prod_tooling_zip = re.search(r'\S+.zip', _line).group()
                            _prod_tooling_zip = _prod_tooling_zip.strip()
                            _prod_tooling_zip_entry = "\n" + "_prod_tooling_zip :" + _prod_tooling_zip                            
                        
                        fj.write(_prod_tooling_zip_entry)    
                        #print (" _prod_tooling_zip :",_prod_tooling_zip )
                    else:
                        print("\n could not find _prod_tooling_zip ")
                        fj.write("\n _prod_tooling_zip :not available" )
  
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _prod_tooling_tar = re.search(r'\*Name\sof\sArtifact\sII\*\s+\|\s\S+', _line)
                    #print ("_prod_tooling_tar:",_prod_tooling_tar )
                    if _prod_tooling_tar != None:
                        _prod_tooling_tar = re.search(r'\S+.tar.bz2', _line)
                        
                        if _prod_tooling_tar is not None:
                            _prod_tooling_tar = re.search(r'\S+.tar.bz2', _line).group()
                            _prod_tooling_tar = _prod_tooling_tar.strip()
                            _prod_tooling_tar_entry = "\n" + "_prod_tooling_tar :" + _prod_tooling_tar
 
                        fj.write(_prod_tooling_tar_entry)    
                    else:
                        print("\n could not find _prod_tooling_tar ")
                        fj.write("\n _prod_tooling_tar :not available" )

                    _line_no = _line_no + 1   
                    _line =  _all_lines[_line_no]
                    _prod_tooling_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://sites.inside-share3.bosch.com\S+', _line)
                    if _prod_tooling_link != None:
                        _prod_tooling_link = re.search(r'https?://sites.inside-share3.bosch.com\S+', _line).group()
                        _prod_tooling_link = _prod_tooling_link.strip()
                        _prod_tooling_link_entry = "\n" + "_prod_tooling_link :" + _prod_tooling_link
                        fj.write(_prod_tooling_link_entry)    
                        #print (" _prod_tooling_link :",_prod_tooling_link )
                    else:
                        print ("_prod_tooling_link not found ")
                        fj.write("\n _prod_tooling_link :not available" )
                    #print("****************************************************************************************************************************************")
                else:
                    print(" _prod_tooling_ver not found ") 
                    fj.write("\n _prod_tooling_ver :not available " )

            else:
                print("could not find Prod Tooling section ")
        
        #SXM Section
        elif (len(re.findall(r'\W#9\*\s+\|\|\s\*SXM\sFW\S+' , _line )) > 0 ):          
            _sxm_sec = re.findall(r'\W#9\*\s+\|\|\s\*SXM\sFW\*', _line )
            #print ("\n _sxm_sec ",_sxm_sec )
            if (len(_sxm_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]
                #*Version ID1*               | 06.38.00_221 
                #_sxm_ver = re.search(r'\*Version\sID1*', _line)
                _sxm_ver = re.search(r'\*Version\sID1\*\s+\W.\w+', _line)
                print (" _sxm_ver :",_sxm_ver )
                if _sxm_ver != None: 
                    _sxm_ver = re.search(r'\d+.\d+.\w+', _line).group()
                    _sxm_ver = _sxm_ver.strip()
                    #print (" _sxm_ver :",_sxm_ver )                    
                    _sxm_ver_entry = "\n" + "_sxm_ver :" + _sxm_ver
                    fj.write(_sxm_ver_entry)    
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    #*Version ID2*               | V08.44.00 
                    _sxm_fw_ver = re.search(r'\*Version\sID2*', _line)
                    if _sxm_fw_ver != None:
                        _sxm_fw_ver = re.search(r'[v|V]\d+.\d+.\d+', _line)
                        #print (" _sxm_fw_ver :",_sxm_fw_ver )
                        if _sxm_fw_ver != None:
                            _sxm_fw_ver = re.search(r'[v|V]\d+.\d+.\d+', _line).group()
                            _sxm_fw_ver = _sxm_fw_ver.strip()                           
                            _sxm_fw_ver_entry = "\n" + "_sxm_fw_ver :" + _sxm_fw_ver
                            fj.write(_sxm_fw_ver_entry)    
                            _inventory_versions_dict["_sxm_fw_ver"] = _sxm_fw_ver
                        else:
                            print(" _sxm_fw_ver entry not found in Version ID2 section") 
                            fj.write("\n _sxm_fw_ver :not available " )

                    else:
                        print(" _sxm_fw_ver not found ") 
                        _sxm_fw_ver = "not available"
                        fj.write("\n _sxm_fw_ver :not available " ) 
                    #add Name of Artifact II section 
                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]
                    _sxm_fw_zip = re.search(r'\*Name\sof\sArtifact\sII\*\s+\|\s\S+', _line)
                    #print ("_sxm_fw_zip:",_sxm_fw_zip )
                    if _sxm_fw_zip != None:
                        _sxm_fw_zip = re.search(r'\S+.zip', _line)    
                        if _sxm_fw_zip is not None:
                            _sxm_fw_zip = re.search(r'\S+.zip', _line).group()
                            _sxm_fw_zip = _sxm_fw_zip.strip()
                            _sxm_fw_zip_entry = "\n" + "_sxm_fw_zip :" + _sxm_fw_zip
                                
                            fj.write(_sxm_fw_zip_entry)    
                    else:
                        print("\n could not find _sxm_fw_zip ")
                        _sxm_fw_zip = "not available"
                        fj.write("\n _sxm_fw_zip :not available" )
                        
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _sxm_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                    if _sxm_link != None:
                        _sxm_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _sxm_link = _sxm_link.strip()
                        #print ("_sxm_link :",_sxm_link )
                        _sxm_link_entry = "\n" + "_sxm_link :" + _sxm_link
                        fj.write(_sxm_link_entry)
                    else:
                        print ("SXM link not found ")
                        _sxm_link = "not available"
                        fj.write("SXM link :not available " )
                    #print (" *****************************************************************************************")
                
                else:
                    print(" _sxm_ver not found ") 
                    _sxm_ver = "not available"
                    _sxm_fw_ver = "not available"
                    _sxm_fw_zip = "not available"
                    _sxm_link = "not available"
                    fj.write("\n _sxm_ver :not available " )
                    fj.write("\n _sxm_fw_ver :not available " )
                    fj.write("\n _sxm_fw_zip :not available " )
                    fj.write("\n _sxm_link :not available " )
            else:
                print("could not find SXM section ")
                
                
                
        
        #DTV Section
        elif (len(re.findall(r'\W#10\*\s+\|\|\s\*DTV\sFW\S+' , _line )) > 0 ):           
            _dtv_sec = re.findall(r'\W#10\*\s+\|\|\s\*DTV\sFW\*', _line )
            if (len(_dtv_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]                   
                _dtv_ver = re.search(r'\*Version\sID1\*\s+\W.\w+', _line)
                #print (" _dtv_ver :",_dtv_ver )
                if _dtv_ver != None: 
                    _dtv_ver = re.search(r'TH\w+', _line).group()                 
                    _dtv_ver_entry = "\n" + "_dtv_ver :" + _dtv_ver
                    fj.write(_dtv_ver_entry)               
                    if re.search("BL:", _line):
                        #print(repr(_line))     #displays line with hidden chars
                        _pattern = r'BL:\s\S+' #_pattern = 'BL:\s\S+'
                        _dtv_BL_ver = re.search(_pattern, _line).group()
                        _dtv_BL_ver = _dtv_BL_ver.split(":")[-1].strip()
                        _dtv_BL_ver = _dtv_BL_ver[:4]
                        #print (" _dtv_BL_ver :",_dtv_BL_ver )
                        fj.write("\n _dtv_BL_ver :not available " )
                    else:
                        _dtv_BL_ver = "not available"
                        fj.write("\n _dtv_BL_ver :not available " )
                        
                    if re.search("HW:", _line):
                        #print(repr(_line))
                        _pattern = r'HW:\s\S+'   #_pattern = 'HW:\s\S+'
                        _dtv_HW_ver = re.search(_pattern, _line).group()
                        _dtv_HW_ver = "00" + _dtv_HW_ver.split(":")[-1].strip()
                        #print (" _dtv_HW_ver :",_dtv_HW_ver )
                        fj.write("\n _dtv_HW_ver :not available " )
                    else:
                        _dtv_HW_ver = "not available"
                        fj.write("\n _dtv_HW_ver :not available " )
                    
                    _line_no = _line_no + 6
                    _line =  _all_lines[_line_no]
                    _dtv_fw_zip = re.search(r'\*Name\sof\sArtifact\sII\*\s+\|\s\S+', _line)
                    #print ("_dtv_fw_zip:",_dtv_fw_zip )
                    if _dtv_fw_zip != None:
                        _dtv_fw_zip = re.search(r'\S+.zip', _line)    
                        if _dtv_fw_zip is not None:
                            _dtv_fw_zip = re.search(r'\S+.zip', _line).group()
                            _dtv_fw_zip = _dtv_fw_zip.strip()
                            _dtv_fw_zip_entry = "\n" + "_dtv_fw_zip :" + _dtv_fw_zip
                                
                            fj.write(_dtv_fw_zip_entry)    
                    else:
                        print("\n could not find _dtv_fw_zip ")
                        _dtv_fw_zip = "not available"
                        fj.write("\n _dtv_fw_zip :not available" )
                        
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _dtv_link = re.search(r'\*Link\sto\sArtifact\*\s+\W.https?://rb-cmbinex\S+', _line)
                    if _dtv_link != None:
                        _dtv_link = re.search(r'https?://rb-cmbinex\S+', _line).group()
                        _dtv_link = _dtv_link.strip()
                        #print ("_dtv_link :",_dtv_link )
                        _dtv_link_entry = "\n" + "_dtv_link :" + _dtv_link
                        fj.write(_dtv_link_entry)
                    else:
                        print ("DTV link not found ")
                        _dtv_link = "not available"
                        fj.write("\n _dtv_link  :not available " )
                    #print (" *****************************************************************************************")
                    ##############################
                else:
                    print(" _dtv_ver not found ") 
                    _dtv_ver = "not available"
                    _dtv_BL_ver = "not available"
                    _dtv_HW_ver = "not available"
                    _dtv_fw_zip = "not available"
                    _dtv_link = "not available"
                    
                    fj.write("\n _dtv_ver :not available " )
                    fj.write("\n _dtv_BL_ver :not available " )
                    fj.write("\n _dtv_HW_ver :not available " )
                    fj.write("\n _dtv_fw_zip :not available" )
                    fj.write("\n _dtv_link  :not available ") 
                    

            else:
                print("could not find DTV section ")
                
                
        #LT FW
        elif (len(re.findall(r'\W#11\*\s+\|\|\s\*LT\sFW\S+' , _line )) > 0 ):    
            _lt_sec = re.findall(r'\W#11\*\s+\|\|\s\*LT\sFW\*', _line ) 
            if (len(_lt_sec) > 0):
                _line_no = _line_no + 2
                _line =  _all_lines[_line_no]                   
                _lt_ver = re.search(r'\*Version\sID1\*\s+\W.\w+', _line)
                #print (" _lt_ver :",_lt_ver )
                if _lt_ver != None: 
                    _lt_ver = re.search(r'\|\s\S+', _line).group()
                    _lt_ver = _lt_ver.split("|")[-1].strip()
                    _lt_ver = _lt_ver[1:]
                    #print (" _lt_ver :",_lt_ver )     
                    _lt_ver_entry = "\n" + "_lt_ver :" + _lt_ver
                    fj.write(_lt_ver_entry)               
                    
                    _line_no = _line_no + 5
                    _line =  _all_lines[_line_no]
                    _lt_fw_hex = re.search(r'\*Name\sof\sArtifact\sI\*\s+\|\s\S+', _line)
                    if _lt_fw_hex != None:
                        _lt_fw_hex = re.search(r'\S+.hex', _line)    
                        if _lt_fw_hex is not None:
                            _lt_fw_hex = re.search(r'\S+.hex', _line).group()
                            _lt_fw_hex = _lt_fw_hex.strip()
                            _lt_fw_hex_entry = "\n" + "_lt_fw_hex :" + _lt_fw_hex 
                            fj.write(_lt_fw_hex_entry)    
                            #print (" _lt_fw_hex :",_lt_fw_hex )
                    else:
                        print("\n could not find _lt_fw_hex ")
                        _lt_fw_hex = "not available"
                        fj.write("\n _lt_fw_hex :not available" )
                        
                    _line_no = _line_no + 1
                    _line =  _all_lines[_line_no]
                    _lt_fw_zip = re.search(r'\*Name\sof\sArtifact\sII\*\s+\|\s\S+', _line)
                    #print ("_lt_fw_zip:",_lt_fw_zip )
                    if _lt_fw_zip != None:
                        _lt_fw_zip = re.search(r'\S+.zip', _line)    
                        if _lt_fw_zip is not None:
                            _lt_fw_zip = re.search(r'\S+.zip', _line).group()
                            _lt_fw_zip = _lt_fw_zip.strip()
                            _lt_fw_zip_entry = "\n" + "_lt_fw_zip :" + _lt_fw_zip    
                            fj.write(_lt_fw_zip_entry)    
                    else:
                        print("\n could not find _lt_fw_zip ")
                        _lt_fw_zip = "not available"
                        fj.write("\n _lt_fw_zip :not available" )
                    
                else:
                    print(" _lt_ver not found ") 
                    _lt_ver = "not available"                   
                    _lt_fw_hex = "not available"
                    _lt_fw_zip = "not available"
                    
                    fj.write("\n _lt_ver :not available " )
                    fj.write("\n _lt_fw_hex :not available " )
                    fj.write("\n _lt_fw_zip :not available " )
                    exit
            else:
                print("could not find LT section ")
            

        else:
            pass
 




        _line_no += 1
    fj.close()
    print ("\nInventory Jira details are logged in :", _jira_File)
    print("\n")
    if "EU" in _rel_name:
        if _gas_appsw != "not available" and _nongas_appsw != "not available":      #None cases will be set as "not available"
            if _gas_appsw != _nongas_appsw:
                print ( "******Error: PLS CHECK APP SW version in GAS section does not match with the version in NON-GAS section****** ")
                exit()
        elif _gas_appsw == "not available" and _nongas_appsw == "not available":
            print ("******App SW version is not updated in GAS and NON-GAS section******")
            exit()
        elif _gas_appsw == "not available":
            print ("******Please note App SW in GAS section is not updated****** ")
        elif _nongas_appsw == "not available":
            print ("******Please note App SW in NON-GAS section is not updated****** ")
    else:
        print ("not an end unit release, App SW will not be delivered")



def update_master_xml(_new_xml):
    global _PN_len
    global _PN_dict
    global _branch
    
    _PN_len = ""
    print ("\n start to update the master xml :    ", _new_xml )
    tree = ET.ElementTree()
    tree.parse(_new_xml)
    root = tree.getroot()
    ET.indent(tree, space='  ', level=0)  
    
    _ecn =  root.find("./Overall_Infos[@Col1='ecn']")
    _ecn.attrib["Col2"] = _ecn_no
    
    _ecr =  root.find("./Overall_Infos[@Col1='ecr']")
    _ecr.attrib["Col2"] = _ecr_no

    _pd_dir = root.find("./Overall_Infos[@Col1='kds_pd_dir']")
    _pd_dir.attrib["Col2"] = _PD_Config_link

    _pd_sw_ver = root.find("./Overall_Infos[@Col1='kds_pd_ver']")
    _pd_sw_ver.attrib["Col2"] = _PD_Config_ver

    _pd_file = root.find("./Overall_Infos[@Col1='kds_pd_file']")
    _pd_file.attrib["Col2"] = _PD_Config_zip
    
    _sw_id_gas = root.find("./Overall_Infos[@Col1='swid_gas']")
    _sw_id_gas.attrib["Col2"] = _SWID_gas
    
    _sw_id_nongas = root.find("./Overall_Infos[@Col1='swid_nongas']")
    _sw_id_nongas.attrib["Col2"] = _SWID_nongas
    
    _pcm = root.find("./Overall_Infos[@Col1='del_resp']")
    _pcm.attrib["Col2"] = _del_resp
    
    _proj_man = root.find("./Overall_Infos[@Col1='PM']")
    _proj_man.attrib["Col2"] = _pm
    
    
    _sw_man = root.find("./Overall_Infos[@Col1='SW_PM']")
    _sw_man.attrib["Col2"] = _sw_pm
    
    _hw_man = root.find("./Overall_Infos[@Col1='HW_PM']")
    _hw_man.attrib["Col2"] = _hw_pm
    
    _release_type = root.find("./Overall_Infos[@Col1='purpose']")
    _release_type.attrib["Col2"] = _rel_type.split(":")[-1]
   

    #Production folder name
    _sw_ver = root.find("./Overall_Infos[@Col1='sw_full_ver']")
    _sw_ver.attrib["Col2"] = _rel_version
    
    _tmpl_sec = root.find("./Overall_Infos[@Col1='template_dir']")
    _templ_dir = _tmpl_sec.attrib["Col2"]
    _templ_dir = _templ_dir.split("UnitType")[0]
    if "EU" in _rel_name: 
        _templ_dir = _templ_dir + "Endunit"
        _tmpl_sec.attrib["Col2"] = _templ_dir
    elif "BU" in _rel_name:
        _templ_dir = _templ_dir + "Baseunit"
        _tmpl_sec.attrib["Col2"] = _templ_dir       
   
    _today = dt.now()
    _date = _today.strftime("%a, %d.%m.%Y,%X")      

    _dtv_sw_sec = root.find("./SW_Versions[@key='DTV_SW']") 
    _dtv_sw_sec.attrib["version"] = _dtv_ver
    _dtv_sw_sec.attrib["src_file"] = _dtv_link
    _dtv_sw_sec.attrib["file"] = _dtv_fw_zip
    _dtv_sw_sec.attrib["label"] = _dtv_sw_sec.attrib["label"].replace("DTV_SW_VER", _dtv_ver )
     
    _dtv_boot_sec = root.find("./SW_Versions[@key='DTV_Boot_SW']")    
    _dtv_boot_sec.attrib["version"] = _dtv_BL_ver
    _dtv_boot_sec.attrib["label"] = _dtv_boot_sec.attrib["label"].replace("DTV_BOOT_VER", _dtv_BL_ver)
    _dtv_hw_sec = root.find("./SW_Versions[@key='DTV_HW']")    
    _dtv_hw_sec.attrib["version"] = _dtv_HW_ver
    _dtv_hw_sec.attrib["label"] = _dtv_hw_sec.attrib["label"].replace("DTV_HW_VER", _dtv_HW_ver)
   
    _sxm_sec = root.find("./SW_Versions[@key='SXM']") 
    _sxm_sec.attrib["version"] = _sxm_ver
    _sxm_sec.attrib["fw_version"] = _sxm_fw_ver
    _sxm_sec.attrib["file"] = _sxm_fw_zip
    _sxm_sec.attrib["src_file"] = _sxm_link
     
    _lt_fw_sec = root.find("./SW_Versions[@key='Lontium_SW']")
    _lt_fw_sec.attrib["version"] = _lt_ver
    _lt_fw_sec.attrib["hex_file"] = _lt_fw_hex
    _lt_fw_sec.attrib["zip_file"] = _lt_fw_zip
    
       
     #CTS section
    _cts_sec = root.find("./SW_Versions[@key='cts']")
    _cts_sec.attrib["version"] = _cts_ver
    _cts_sec.attrib["CTS_file"] = _cts_binary
    _cts_sec.attrib["src_file"] = _cts_link
    _cts_sec.attrib["CTS_config_version"] = _cts_config_ver
    _cts_sec.attrib["CTS_config_src_file"] = _cts_config_zip
    _cts_sec.attrib["CTS_config_src_file"] = _cts_config_link

    #APP SW section  -- to be added - errorhandling if gas app sw version does not match with non-gas app sw version - done
    _app_sec = root.find("./SW_Versions[@key='android_app_sw']")
    if _gas_appsw != "not available" :
        _app_sec.attrib["version"] = _gas_appsw.strip()
    elif _nongas_appsw != "not available" :   
        _app_sec.attrib["version"] = _nongas_appsw.strip()
    _app_sec.attrib["app_sw_gas_version"] = _gas_appsw_ver
    _app_sec.attrib["app_sw_nongas_version"] = _nongas_appsw_ver
    _app_sec.attrib["gas_file"] = _gas_appsw_zip
    _app_sec.attrib["nongas_file"] = _nongas_appsw_zip
    _app_sec.attrib["gas_src_file"] = _gas_appsw_link
    _app_sec.attrib["nongas_src_file"] = _nongas_appsw_link
    _branch =  _app_sec.attrib["version"]
    if "EU" in _rel_name:
        if _gas_appsw_ver != "not available":
            _android_gas_key = _gas_appsw_ver.split(".")[-1][0]
            _android_key = _android_gas_key
        else:
            _android_gas_key = ""
            
        if _nongas_appsw_ver != "not available":
            _android_nongas_key = _nongas_appsw_ver.split(".")[-1][0]
            _android_key = _android_nongas_key
        else:
            _android_nongas_key = ""
        
        if _android_gas_key == "" and _android_nongas_key == "":
            print("ERROR: The GAS and NON GAS App versions are not available, please check")
            print("GAS App SW Version  :", _gas_appsw_ver)
            print("NON GAS App SW Version  :", _nongas_appsw_ver)
            exit()

        elif _android_gas_key == "" or _android_nongas_key == "": 
            print("\n Android key :", _android_key)
            pass
        
        elif _android_gas_key != _android_nongas_key:
            print("****Anrdoid version does not match****")
            print("Gas Anroid Version key :", _android_gas_key)
            print("NonGas Anroid Version key :", _android_nongas_key)
            exit()
        
        if _android_key.upper() == "S":
            _android_key = "A12"
        elif _android_key.upper() == "U":
            _android_key = "A14"
        else:
            print("Unknown Android version key :", _android_key)
            exit()
    elif "BU" in _rel_name:
        _android_key = "n.a."
    #Aurix section
    _aurix_sec = root.find("./SW_Versions[@key='aurix_autosar_app_sw']")
    _aurix_sec.attrib["version"] = _aurixsw_ver
    _aurix_sec.attrib["srec_file"] = _aurix_srec
    _aurix_sec.attrib["dnl_file"] = _aurix_dnl
    _aurix_sec.attrib["support_file"] = _aurix_support_dnl
    _aurix_sec.attrib["src_file"] = _aurix_link
    _aurix_sec.attrib["crypto_env"] = _cryptosignaurix
    _aurix_sec.attrib["_sub_path"] = _aurix_sub_path
    #_aurix_sec.attrib["_sub_path"] = _aurix_link.split("Output/vector/")[1]
    #_aurix_subpath = _aurix_sec.attrib["_sub_path"]

    #Ublox section
    _ublox_sec = root.find("./SW_Versions[@key='ublox']")
    _ublox_sec.attrib["version"] = _ublox_ver
    _ublox_sec.attrib["bu_file"] = _ublox_bu_bin
    _ublox_sec.attrib["eu_file"] = _ublox_eu_bin
    
    #_ublox_sec.attrib["src_file"] = _aurixsw_verlink #TBD

    #Test Manager section
    _tm_sec = root.find("./SW_Versions[@key='aurix_tm']")
    _tm_sec.attrib["version"] = _test_manager_ver
    _tm_sec.attrib["file"] = _test_manager_zip
    _tm_sec.attrib["src_file"] = _test_manager_link
    
    #Flw4Linux section
    _flw4_sec = root.find("./SW_Versions[@key='Flash4Linux']")
    _flw4_sec.attrib["version"] = _Flw4Linux_ver
    
    #Production Tooling section
    _ptooling_sec = root.find("./SW_Versions[@key='Prod_Tooling']")
    _ptooling_sec.attrib["version"] = _prod_tooling_ver
    _ptooling_sec.attrib["file"] = _prod_tooling_zip
    _ptooling_sec.attrib["tar_file"] = _prod_tooling_tar
    _SWVer_dict = _ptooling_sec.attrib
    _ptooling_dir = _SWVer_dict["src_file"]
    _new_ptooling_dir = _ptooling_dir.replace("version",_prod_tooling_ver)
    # _SWVer_dict["src_file"] = _new_ptooling_dir
    # _SWVer_dict["SP_link"] = _prod_tooling_link
    _SWVer_dict["src_file"] = _prod_tooling_link
    
    
    #PD Delivery state
    _pd_sec = root.find("./SW_Versions[@key='PD_Delivery_State']")
    _pd_sec.attrib["version"] = _PD_Config_ver
    _pd_sec.attrib["file"] = _PD_Config_zip
    _pd_sec.attrib["src_file"] = _PD_Config_link
   
    _target = root.find("./Overall_Infos[@Col1='target_dir']")
    _OverallInfo_dict = _target.attrib
    _target_dir = _OverallInfo_dict["Col2"]
    _new_tar_dir = _target_dir.replace("Release_Version",_rel_version)
    _OverallInfo_dict["Col2"] = _new_tar_dir
    
    _row_count = len(df_list)
    _col_count = len(df.columns)
    tree.write(_new_xml, encoding="utf-8") 
    
    _PN_dict = {}
    _setInfoList = ["part_number", "product", "scope", "ufs_info", "ufs_spl_id", "aurix_info", "aurix_spl_id", "gnss_prod", "gnss_spl_id", "fpga_prod", "sxm", "dtv", "adr_type", "device_type", "supplier_feed", "plant", "BU", "variant" ]
    
    for i in range(len(df_list)):
        _del_req = df_list[i][0]
        if _del_req in ["X","x"]:
            #k = 0
            _PN = str(df_list[i][1]).strip()
            _PN_dict[_PN] = df_list[i][14]          #mapping PN with device type, gas and non gas
            #print ("\n _del_required for PN ",df_list[i][1] )
            if len(_PN) == 10:
                _PN_len = "sop"
            elif len(_PN) == 13:
                _PN_len = "tsb"
            parent=ET.Element(root.tag)
            child = ET.SubElement(parent, 'SET_Infos')
            k = 0
            for j in range(1,_col_count):
                _item_name = _setInfoList[k] 
                #child.attrib[_item_name] = df_list[i][j ]
                child.set(_item_name,str(df_list[i][j ]))
                #child.tail = "\n\t"
                
                k = k + 1  
            #root[-1].tail = "\n\t"  
            root.append(child)
            ET.indent(tree,' ')            
            tree.write(_new_xml, encoding="utf-8")            
        else:
           #print("\n PN will be excluded ")
           pass
        
    for _element in root:
        if _element.tag == "DocInfo":
            _element.attrib["Doc_Version"] = "V01.00"
            _element.attrib["Doc_Date"] = _date
            _element.attrib["Doc_SetDef"] = _set_Definiton
            
        if _element.tag == "ProjectInfo":
            _element.attrib["SW_Version"] = _rel_version
            if _SWID_gas == "not available":
                _element.attrib["SW_ID_GAS"] = ""
            else:            
                _element.attrib["SW_ID_GAS"] = _SWID_gas
            
            if _SWID_nongas == "not available":
                _element.attrib["SW_ID_NONGAS"] = ""
            else:            
                _element.attrib["SW_ID_NONGAS"] = _SWID_nongas
            
            _element.attrib["SW_ID"] = ""
            _element.attrib["SW_PM"] = _sw_pm
            #_element.attrib["SW_PM"] => HARDCODED VALUE IN XML 
            #_element.attrib["CHNG_NR"] => will be filled in sap xml during create_prod run in set_info section
    
    _supplier_list = ["VNEXT", "AERIS"]    
    _suplier_feed_set = "no"
    _plant_set = "no"
    #print ("setattrib:", SET.attrib )
    for SET in root.iter("SET_Infos"):
        _supplierfeed = SET.attrib["supplier_feed"] 
        if _suplier_feed_set == "no":
            _supplier = _supplierfeed
            _suplier_feed_set = "yes"
        elif _suplier_feed_set == "yes":
            if _supplier != _supplierfeed:
                print("****************************************************************")
                print("ERROR: Release cannot have PNs with different Supplier_Feed ... exiting ")
                sys.exit(0)
        
        _plant_setInfo = SET.attrib["plant"] 
        if _plant_set == "no":
            _plant = _plant_setInfo
            _plant_set = "yes"
        elif _plant_set == "yes":
            if _plant != _plant_setInfo:
                print("****************************************************************")
                print("ERROR: Release cannot have PNs with different Plant ... exiting ")
                sys.exit(0)
    
    if "EU" in _rel_name: 
        if _supplier.upper() not in _supplier_list:
            print("ERROR: Release cannot have PNs with Supplier_Feed other than AERIS or VNEXT ... exiting ")
            sys.exit(0)
        if _supplier.upper() == "VNEXT":
            _supplier_feed = "NISSAN"
        elif _supplier.upper() == "AERIS":
            _supplier_feed = "MMC"       
    elif "BU" in _rel_name:
        _supplier_feed = "NA"     
    
    print ("\n _supplier_feed :", _supplier_feed)
    
    _CTS_parameter_sec = root.find("./SW_Versions[@key='CTS_Parameter']")
    _App_parameter_sec = root.find("./SW_Versions[@key='Application_Parameter']")     
    _SecBoot_parameter_sec = root.find("./SW_Versions[@key='Secure_Boot_Parameter']") #no change required
    _UCB_Lock_sec = root.find("./SW_Versions[@key='UCB_Lock_Activate']")     
    _Sec_Lock_sec = root.find("./SW_Versions[@key='Sec_Lock_Activate']")
    _crypto_bosch_sec = root.find("./SW_Versions[@key='cryptoenv_bosch']") 
    _crypto_cust_sec = root.find("./SW_Versions[@key='cryptoenv_cust']") 
    _crypto_bosch_sec.attrib["version"] = _cryptoBosch
    _crypto_cust_sec.attrib["version"] = _cryptoCust
    _oemcrypto_sec = root.find("./SW_Versions[@key='oemcrypto']")
    _oemcrypto_sec.attrib["version"] = _supplier_feed
    _android_version_sec = root.find("./SW_Versions[@key='android_version']")
    _android_version_sec.attrib["version"] = _android_key
    
    
    _VIP_parameter_sec = root.find("./SW_Versions[@key='VIP_DNL_Parameter']")
    if "EU" in _rel_version.upper() and "END" in _rel_type.upper():
        _CTS_parameter_sec.attrib["cts_param_rxl"] = "rawprogram_ccs2_cts_C2_reflash.xml"
        _App_parameter_sec.attrib["app_param_rxl"] = "rawprogram_ccs2.0_aivi2-ccs20-D1.xml,rawprogram_ccs2.0_persist.xml"
        _UCB_Lock_sec.attrib["status"] = "YES"
        _Sec_Lock_sec.attrib["status"] = "YES"
        
        if _PN_len == "tsb":
            _VIP_parameter_sec.attrib["vip_param_op"]= "n.a."
            _VIP_parameter_sec.attrib["vip_param_rxl"]= "n.a."
            _VIP_parameter_sec.attrib["vip_param_va"]= "n.a."
            _VIP_parameter_sec.attrib["vip_param_pxl"]= "n.a."
            _UCB_Lock_sec.attrib["status"] = "NO"
            _Sec_Lock_sec.attrib["status"] = "NO"
            _App_parameter_sec.attrib["app_param_rxl"] = "rawprogram_ccs2.0_aivi2-ccs20-C2.xml,rawprogram_ccs2.0_persist.xml"
   

    else:
        _CTS_parameter_sec.attrib["cts_param_rxl"] = "rawprogram_ccs2_cts_C2.xml" 
    
    tree.write(_new_xml, encoding="utf-8")     
        
    #print("\n \n")
    #print("****************************************************************************************************************************************")
    print ("\n\nInventory Jira details are logged in :", _jira_File)
    print ("\nMaster xml: ",_new_xml," is created " )
    print("****************************************************************************************************************************************")


def update_setDefinition(_xls_File ):
    _xls_File = _xls_File
    wb =  openpyxl.load_workbook(_xls_File)     #load the work book
    ws =  wb['Set Definition']                  #activate the sheet set definition
    #print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    print ("\nat update set definition .. ")
    global _rel_version
    global _rel_type
    global df_list
    global _PN_col
    global df
    global _ecn_no
    global _ecr_no
    global _pm
    global _sw_pm
    global _hw_pm
    global _del_resp
    
        
    ws[_PD_conf_cell] = ""
    ws[_Prod_tool_cell] = ""
    ws[_CTS_conf_cell] = ""
    ws[_CTS_ver_cell] = ""
    ws[_AppSw_ver_cell] = ""
    ws[_SWID_gas_cell] = ""
    ws[_SWID_nongas_cell] = ""
    ws[_GAS_ver_cell] = ""
    ws[_NonGAS_ver_cell] = ""
    ws[_Aurix_ver_cell] = ""
    ws[_Aurix_sub_path_cell] = ""
    ws[_TestManager_cell] = ""
    ws[_GNSS_cell ] = ""
    ws[_SXM_cell ] = ""
    ws[_SXM_fw_cell] = ""
    ws[_DTV_cell ] = ""
    ws[_Lontium_cell ] = ""
   
    
    _set_info_start_row = 23 
    _ecn_no = ""
    _ecr_no = ""
    ws[_rel_type_cell] = "P2: " + _rel_TaskName
    _rel_type = ws[_rel_type_cell].value 
    _ecn_ecr_selection = ws[_ecn_ecr_cell].value
    _rel_version = ws[_rel_version_cell].value
    _rel_version = _rel_version.strip()
    _rel_type = _rel_type.strip()
    ws[_Inventory_cell] = _issue_Summary
    ws[_jira_ID_cell ] = _issue_ID
    _del_resp = ws[_del_resp_cell].value
    _pm = ws[_pm_cell].value
    _sw_pm = ws[_sw_pm_cell].value
    _hw_pm = ws[_hw_pm_cell].value

    try:    
        if _ecn_ecr_selection is not None:
            if _ecn_ecr_selection == "ECR":
                _ecr_no = ws[_ecr_cell].value
            elif _ecn_ecr_selection == "ECN":
                _ecn_no = "read_from_sap_sheet"
        else:
            _ecn_no = "None"
            _ecr_no = "None"
            print ("ecn value: ", _ecn_no )
            print ("ecr value: ", _ecr_no )
            print("please choose ECR / ECR ")  
    
        if _PD_Config_ver is not None:
            ws[_PD_conf_cell] = _PD_Config_ver 
        else:
            print("**** PD Config version is None **** ")
            raise TypeError 

        if _prod_tooling_ver is not None:
            ws[_Prod_tool_cell] = _prod_tooling_ver 
        else:
            print("**** Production tooling version is None **** ")
            raise TypeError

        if _cts_config_ver is not None:
            ws[_CTS_conf_cell] = _cts_config_ver 
        else:
            print("**** CTS Config version is None **** ")
            raise TypeError

        if _cts_ver is not None:
            ws[_CTS_ver_cell] = _cts_ver 
        else:
            print("**** CTS version is None **** ")
            raise TypeError

        if _aurixsw_ver is not None:
            ws[_Aurix_ver_cell ] = _aurixsw_ver 
            ws[_Aurix_sub_path_cell ] = _aurix_sub_path
        else:
            print("**** Aurix version is None **** ")
            raise TypeError

        if _test_manager_ver is not None:
            ws[_TestManager_cell ] = _test_manager_ver 
        else:
            print("**** Test Manager version is None **** ")
            raise TypeError

        if _ublox_ver is not None:
            ws[_GNSS_cell  ] = _ublox_ver 
        else:
            print("**** Ublox version is None **** ")
            raise TypeError
            
        if _sxm_ver is not None:
            ws[_SXM_cell  ] = _sxm_ver 
        else:
            print("**** SXM version is None **** ")
            raise TypeError
            
        if _sxm_fw_ver is not None:
            ws[_SXM_fw_cell  ] = _sxm_fw_ver 
        else:
            print("**** SXM firmware version is None **** ")
            raise TypeError
            
        if _dtv_ver is not None:
            ws[_DTV_cell  ] = _dtv_ver 
        else:
            print("**** DTV version is None **** ")
            raise TypeError
            
        if _lt_ver is not None:
            ws[_Lontium_cell  ] = _lt_ver 
        else:
            print("**** Lontium version is None **** ")
            raise TypeError
            
            
        _rel = re.search(r'end\sunit', _rel_type, re.IGNORECASE)    
        #_rel = re.search('end\sunit', _rel_type, re.IGNORECASE)
        
        if _gas_appsw != "not available" :
            ws[_AppSw_ver_cell ] = _gas_appsw
        elif _nongas_appsw != "not available" :
            ws[_AppSw_ver_cell ] = _nongas_appsw
        else:
            print ("**** App SW version is not available **** ")
            
        if _SWID_gas != "not available" :
            ws[_SWID_gas_cell ] = _SWID_gas
        else:
            print ("**** GAS SW ID is not available **** ")
        
        if _gas_appsw_ver != "not available" :
            #print (" data type of gas appsw ver:",type(_gas_appsw_ver))
            ws[_GAS_ver_cell ] = _gas_appsw_ver 
        else:
            print ("**** GAS App SW version is not available **** ")
        
        if _nongas_appsw_ver != "not available" :
            ws[_NonGAS_ver_cell]  = _nongas_appsw_ver 
        else:
           print ("**** NON GAS App SW version is not available **** ")

        if _SWID_nongas != "not available" :
            ws[_SWID_nongas_cell ] = _SWID_nongas
        else:
            print ("**** NON GAS SW ID is not available **** ")    
        
        if _rel is not None:
            pass

        
        _PN_count = 0
        _PN_details = []  
        columns = ['Delivery_Required','Part_Number','Set_Name','Scope','Manufacturer','UFS_type','Aurix-Variant','Aurix-SPL-ID','gnss-Variant','gnss-SPL-ID','FPGA_Version','SXM_Version','DTV','ADR','Device_Type','Supplier Feed','Plant','BU','Variant']
        df = pd.read_excel(_xls_File, header=None, names = columns, sheet_name='Set Definition', skiprows=22)
        df = df.fillna('')  #replace the "NAN"values from columns ( empty string )
        # Convert DataFrame to a list using the 'values' attribute
        df_list = df.values.tolist()
        
    except TypeError:
            print("****************************************************************************************************************************************")
            print ("Type error indicated, please check ")
            print("****************************************************************************************************************************************")

    wb.save(_xls_File)
    
    


def start_copy ( _source, _dest ):
    print ("at start copy .. ")
    # print ("source:", _source) 
    # print ("_dest:", _dest)

    if not os.path.isfile(_dest):
        try:
            shutil.copy(_source, _dest )
        except PermissionError:
            print("Permission denied.")
        except FileNotFoundError:
            print(f"Error: Source file '{_source}' not found.")
        except shutil.SameFileError:
            print("Error: Source and destination are the same file.")    
        # except:
            # print("Error occurred while copying file.") 
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
    else:
        #print ("file ", _dest, "exists")
        pass
            
#####################################################################################
def compare_json_jira (_json_File, _jira_File):
    _json_txt = _json_File
    _jira_txt = _jira_File
    _match_flag = ""
    print("\n at compare_json_jira ... ")
    # print ("json file :", _json_File )
    # print("jira_file:", _jira_File )
    
    with open(_jira_txt) as f1:
        _lines_list_f1 = f1.readlines()             #provide a list of lines
    
    #print("\nReading json file.. ")
    with open (_json_txt) as f2:
        #print(f1.readlines())
        _line_list_f2 = f2.readlines()
        for _line in _line_list_f2:
            #print(_lines )
            if _line.startswith("Reference"):
                #print(_line )
                #_line = "Reference not in document"
                if _line in _lines_list_f1:
                    #print("Match found for line :\n",_line)
                    pass
                else:
                    print("Error: No match found for line:\n",_line)
                    _match_flag = "no"
                    #break
    if _match_flag != "no":
        print("\nAll lines in ",_json_File," matches with the entries in the jira")
    else:
        print("\nERROR:MISMATCH FOUND, please check.. ")
#####################################################################################       
def compare_versions_from_appsw_to_inventory(_repo_version_file):
    #checks if the versions in the _inventory_versions_dict is available in the _jira_File (inventory_No_JiraId.txt file )
    #_inventory_versions_dict are formed during read jira.

    _v_file = _repo_version_file
    _dict1 = _inventory_versions_dict
    _appsw_str_list = ["ISH", "KUM", "NGO", "HND"] 
    print("\nat compare_versions_from_appsw_to_inventory ... ")
    #print ("\nInventory verions dict: ",_dict1 )
    with open(_v_file) as f3:
        _lines_list_f3 = f3.readlines()
    
    for _item in _dict1:
        _match_flag = ""
        _item_LC = _dict1[_item].lower()
        #print ("value : ", _item_LC)
        if _item == "_gas_appsw_ver" or _item == "_nongas_appsw_ver" :
            _appVer = _item_LC
            _len = len(_appVer)
            _sliceStart = 4
            #_sliceEnd = _len - 4
            #_item_LC = _appVer[_sliceStart:_sliceEnd]
            _item_LC = _appVer[_sliceStart:]
            #print ("value : ", _item_LC)
        
        for _line in _lines_list_f3:
            _line_LC = _line.lower()
            if _item_LC in _line_LC:
                _match_flag = "yes"
                
        if _match_flag != "yes":
            print( "version ", _item, " cannot find a match in ", _v_file)
    
if __name__ == '__main__':
    parser = optparse.OptionParser()
    parser.add_option('-x', '--xml', dest='xml', default=None,
                      help='Master xml')
    parser.add_option('-t', '--template', dest='template', default=None,
                      help='Set Definition sheet')
    parser.add_option('-j', '--jira', type="string",dest='jira_id', default=None,
                      help='JIRA ID like AIVI-XXXXXX')                   
    
    (options, args) = parser.parse_args()
    if not options.xml:
        #input_masterxml = input("\nPlease enter base masterxml filename\n")
        print ("\n Please enter the master xml name as input ")
        print ("\n please run the script in the form : ")
        print ("\t update_master.py -x <xml_name> -t <set_definition_name> -j <jira_id> ")
        sys.exit(0) 
    else:
        input_masterxml = options.xml
        _xml_name = input_masterxml
        print ("\n Master xml ", _xml_name," will be used ")
        

    if not options.template:
        #input_template = input("\nPlease enter set definition template filename\n")
        print ("\n Please enter the Set definition name as input ")
        print ("\n please run the script in the form : ")
        print ("\t update_master.py -x <xml_name> -t <set_definition_name> -j <jira_id> ")
        sys.exit(0)
        
    else:
        input_template = options.template
        #print ("\n Set definition :", input_template )
        _xls_FileName = input_template
        print ("\n Set Definition ", _xls_FileName ," will be used ")
        
        
    if not options.jira_id:
        #input_masterxml = input("\nPlease enter base masterxml filename\n")
        print ("\n Please enter the JIRA ID name as input ")
        print ("\n please run the script in the form : ")
        print ("\t update_master.py -x <xml_name> -t <set_definition_name> -j <jira_id> ")
        sys.exit(0) 
    else:
        input_jira = options.jira_id
        _issue_ID = input_jira
        print ("\n JIRA ID ", _issue_ID," will be used ")
        
    _curr_Dir = os.getcwd()
    # print ("\n _curr_Dir : ",_curr_Dir )
    # print(" ********************** ")
    
    # print(os.getlogin())
    global _set_Definiton
    _user = os.getlogin()
    #print ("\n _user : ",_user )
    _tracker_key_file = "Tracker_token.txt"
    _tracker_key_file_path = r'C:\Users'
    _tracker_key_file = os.path.join( _tracker_key_file_path, _user, _tracker_key_file )
    
       
    ft = open(_tracker_key_file, "r")
    _token = ft.readline()
    _token = _token.strip()

    JiraUrl= "https://rb-tracker.bosch.com/tracker05"
    Jira_Access_Token = _token
    #BoschJira = JIRA(options={'server': 'https://hi-cmts.apps.intranet.bosch.com:8443'}, basic_auth=("nsp1cob", "password"))
    #BoschJira = JIRA(options={'server': 'https://rb-tracker.bosch.com/tracker05', 'verify': False}, basic_auth=("nsp1cob", "password"))
    headers = JIRA.DEFAULT_OPTIONS["headers"].copy()
    headers["Authorization"] = f"Bearer {Jira_Access_Token}"
    BoschJira = JIRA(server=JiraUrl, options={"headers": headers})           
            
    _issue_Name = BoschJira.issue(_issue_ID )
    _issue_Summary = _issue_Name.fields.summary
    _rel_name = _issue_Summary
    _rel_fullname = _issue_Summary
    #print("\n issue _rel_name: ", _rel_name)
    _BU_name_list = ["BU", "bu", "Base Units", "Baseunits", "base units", "baseunits", "BASE UNITS", "BASEUNITS"]
    _EU_name_list = ["EU", "eu", "End Units", "Endunits", "end units", "endunits", "END UNITS", "END UNITS"]
    _name_list = [ _BU_name_list, _EU_name_list ] 
    for _list in _name_list:
        if _list == _BU_name_list:
            for _str in _list:
                if _str in _rel_name:
                   _rel_name = "BU"
                   break       
                   
        elif _list == _EU_name_list:
            for _str in _list:
                if _str in _rel_name:
                   _rel_name = "EU"  
                   break
    
    _issue_Summary = _issue_Summary.split("#")[-1]
    _rel_TaskName = (_rel_fullname.split("Inventory")[-1]).split("#")[0]
    _rel_OEM = _rel_TaskName.split("CCS2.0")[0].strip()

    _desc_Filename = "Inventory_" + _issue_Summary + ".txt"
    _desc_File = os.path.join(_curr_Dir,_desc_Filename)

    _jira_Filename = "Inventory_" + _issue_Summary + "-" + _issue_ID + ".txt"
    _jira_File = os.path.join(_curr_Dir,_jira_Filename)

    _desc = repr(_issue_Name.fields.description )
    f1 = open(_desc_File, "w")
    formatted_file =  _desc.replace('\\n', '\n') 
    f1.write(formatted_file)
    f1.close()

    _set_Definiton =  _xls_FileName.split(".xlsx")[0]
    _set_Definiton = _set_Definiton.split("CCS2_")[-1]
    _master_xml = os.path.join(_curr_Dir,_xml_name)
    _xml_basename = os.path.splitext(os.path.basename(_master_xml))[0]
    _new_xml_basename =  _set_Definiton + "_" + _xml_basename
    _new_xml_name = _set_Definiton + "_" + _xml_basename + ".xml"
    _new_xml = os.path.join(_curr_Dir,_new_xml_name)
    _xls_File = os.path.join(_curr_Dir,_xls_FileName) 

   
    if os.path.exists(_new_xml):
        #_suffix = datetime.now().strftime("%Y%m%d_%H%M%S") #creates suffix with current date and exact time
        _suffix = dt.now().strftime("%Y%m%d_%H%M%S") #creates suffix with current date and exact time
        _file_rename = _new_xml_basename + "." + _suffix + ".xml"
        _xml_rename = os.path.join(_curr_Dir,_file_rename)
        os.rename(_new_xml, _xml_rename) 
            
    copy(_master_xml, _new_xml) 
  
  
    with open(_desc_File, 'r') as fp:
        _count = len(fp.readlines())

       
    read_jira_info(_desc_File,_count,_rel_OEM)  
    update_setDefinition(_xls_File)   
    update_master_xml(_new_xml)


    with open (_new_xml , 'r+') as fp:
        readcontent = fp.read()
        fp.seek(0,0)
        fp.write('<?xml version="1.0" encoding="ISO-8859-1"?> \n<?xml-stylesheet version="0.1" type="text/xsl" href="SW_InfoSheet.xsl"?>\n')
        #fp.write('<?xml-stylesheet version="0.1" type="text/xsl" href="SW_InfoSheet.xsl"?>\n')
        fp.write(readcontent)
    
    if _rel_name == "EU":
        print ("\n Continue to proceed with crosschecking the details in the jira with that of the details in the SW ")
        _userinput = input("\n Please confirm to proceed : Y/N \t").upper()
        
        print ("\n _gas_appsw:", _gas_appsw)
        print ("\n _nongas_appsw:", _nongas_appsw)
        
        if _userinput == "Y":
            print("****************************************************************************************************************************************")
            # if _gas_appsw.strip() != _nongas_appsw.strip(): 
                # if _gas_appsw.strip() == "not available":
                    # _branch = _nongas_appsw
                # elif _nongas_appsw.strip() == "not available":
                    # _branch = _gas_appsw
                # else:
                    # print("\n difference in branch vesrions betweeen GAS and NONGAS, please check .. ")
                    # print("\nGAS branch : ",_gas_appsw)
                    # print("\nNON-GAS branch :",_nongas_appsw)
                    # exit()      
            # else:                
                # _branch = _gas_appsw
            print("\nAppsw:", _branch)
            idents_json_file = r'\vendor\bosch\build\idents.json'
            _verion_file = "versions_" + _branch + ".txt"      
            

            _repo_base_VM = "/home/uac2hi/samba/views/repo_init_ccs2/"
            _repo_base = r'\\HI-V-0003R.hi.de.bosch.com\uac2hi_samba\views\repo_init_ccs2'
            #_repo_base = r'\\HI7-V-0000T.hi.de.bosch.com\uac2hi_samba\views\repo_init_ccs2'
            _init_script = "appSW_Verification.sh"
            _repo_init_script = _repo_base_VM + _init_script
            _repo_branch = _repo_base + "\\" + _branch
            _repo_json = _repo_branch + idents_json_file
            _repo_version_file = _repo_branch + "\\" + _verion_file
            _inventory_file = os.path.basename(_jira_File)
          
         
            #_cmd1 = "ssh " + "uac2hi@HI7-V-0000T.hi.de.bosch.com" + " sh " + _repo_init_script + " -b " + _branch
            #_cmd1 = "ssh uac2hi@HI7-V-0000T.hi.de.bosch.com" + " bash -i " + _repo_init_script + " -b " + _branch
            _cmd1 = "ssh uac2hi@HI-V-0003R.hi.de.bosch.com" + " bash -i " + _repo_init_script + " -b " + _branch
            #print ("\n cmd :", _cmd1)
            
            # set the value as xterm as to avoid the warning message  - tput- No value for $TERM - since this opens the shell as interactive mode and there is no real terminal
            # stderr  is set as subprocess.DEVNULL as to suppress the warning  - bash: no job control on this terminal
            _env = os.environ.copy()
            _env["TERM"] = "xterm"                     

            print ("\nconnecting to the VM for the repo download to the path ", _repo_base, " ...")
            #return_code = subprocess.call(_cmd1, shell = True)
            return_code = subprocess.call(_cmd1, env=_env, stderr=subprocess.DEVNULL)

            if return_code == 0:
                print("Command executed successfully.")
                start_copy(_repo_json, _curr_Dir)
                start_copy(_repo_version_file, _curr_Dir)
            else:
                print("Command failed with return code", return_code)
                exit()
            print("****************************************************************************************************************************************")   
                
            _gas_file_check = ""
            _non_gas_file_check = ""
            _PNos = ""
            _PN_count = len(_PN_dict)
            _GPNos = ""
            _NGPNos = ""
            
            
            for _PartNo, _dev_type in _PN_dict.items():
                if _dev_type.lower() == "yes":
                    _gSW = _gas_appsw_ver
                    _gSW_len = len(_gSW.split("."))
                    if _gSW_len == 5:
                        _gSW_list = _gSW.split(".")
                        _gSW_lastStr = "." + _gSW_list[-1]
                        _gSW = _gSW.replace(_gSW_lastStr, "").strip()

                    _gas_file = _gSW + ".txt"
                    
                    if _gas_file_check != "yes":
                        _gas_file_check = "yes"     #the first time
                        _GPNos = _PartNo
                        
                    else:
                        _GPNos = _GPNos + "," + _PartNo
                    
                elif _dev_type.lower() == "non":
                    _ngSW = _nongas_appsw_ver
                    _ngSW_len = len(_ngSW.split("."))
                    if _ngSW_len == 5:
                        _ngSW_list = _ngSW.split(".")
                        _ngSW_lastStr = "." + _ngSW_list[-1]
                        _ngSW = _ngSW.replace(_ngSW_lastStr, "").strip()
                    
                    non_gas_file = _ngSW + ".txt"    
                    
                    if _non_gas_file_check != "yes":
                        _non_gas_file_check = "yes"     #the first time
                        _NGPNos = _PartNo
                        
                    else:
                        _NGPNos = _NGPNos + "," + _PartNo
            
            
            if _GPNos != "":
                _PNos = _GPNos
                _cmd2 = ["perl", "checkIdentsJson_v2.pl", "-i", "idents.json", "-p", _PNos, "-s", _gSW ]
                print ("\n cmd :", _cmd2)
                subprocess.run(_cmd2)
            else:
                pass
                #print ("\nNo GAS Part number available ");    
                
            if _NGPNos != "":
                _PNos = _NGPNos
                _cmd2 = ["perl", "checkIdentsJson_v2.pl", "-i", "idents.json", "-p", _PNos, "-s", _ngSW ]
                print ("\n cmd :", _cmd2)
                subprocess.run(_cmd2)
                
            else:
                pass
                #print ("\nNo NON GAS Part number available ");
               
            
            if _gas_file_check != "" and os.path.exists(_gas_file):
                compare_json_jira(_gas_file, _jira_File)
                
            if _non_gas_file_check != "" and os.path.exists(non_gas_file):
                compare_json_jira(non_gas_file, _jira_File)
                
            compare_versions_from_appsw_to_inventory(_repo_version_file)
                
        else:
            print("\nexiting ..")
            sys.exit(0)
    
    else:
        sys.exit(0)
        
