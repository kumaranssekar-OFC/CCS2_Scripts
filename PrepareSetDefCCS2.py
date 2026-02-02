#################################################################################################################
#
# FILE:         PrepareSetDefCCS2.py
# DESCRIPTION:  This script can be used to prepare the set definition document for a list of partnumbers (for CCS2 project)
# USAGE:        see help_text
# PREREQUISITE: VariantMatrix has to be in the same directory as this script
# HISTORY:
# Date         | Author          		| Modification
# 03.09.2024   | Abinaya Muthukrishnan  | Initial version
# 12.09.2024   | Abinaya Muthukrishnan  | compatibility changes
# 23.10.2024   | Abinaya Muthukrishnan  | Add _10 in UFS for both EU/BU, fill VM ver 
# 24.06.2025   | Abinaya Muthukrishnan  | Changes for filling Supplier feed 
# 12.08.2025   | Abinaya Muthukrishnan  | Changes for filling Plant info
# 16.09.2025   | Abinaya Muthukrishnan  | Changes for filling BaseUnit
#                                         change for MMC- HW_PNS_All sheet
# 23.10.2025   | Abinaya Muthukrishnan  | Bugfixes
#################################################################################################################

import re
import os
from datetime import datetime
import sys
import argparse
import openpyxl
from datetime import date

openpyxl.reader.excel.warnings.simplefilter(action='ignore')

Username = os.environ.get("user")
cwd = os.getcwd()
VM = ""
TemplateDir = "\\\\bosch.com\\dfsrb\\DfsDE\\DIV\\CM\\AI\\SW_Production\\Nissan\\0060_CCS2_7515752366\\00_SW\\_PCM_internal\\Reference_sheets\\Templates_Scripts\\"
docushare_file = cwd + r"\DocuDownload.txt"
SetDefName = cwd + r"\SetDefinition_CCS2_P2_xx_[X]U.xlsx"
SetDefTmpl = cwd + r"\SetDefinition_CCS2_TMPL.xlsx"
HW_PNs_All = cwd + r"\HW_PartNumbersAll_CCS2.xlsx"
SPL_tmpl = cwd + r"\CCS2_SPL_TMPL_V01.00.xlsx"
Scope = "p2"
VM_data = {"Product":{"Title":"Unit BOSCH No"},
           "SDB":{"Title":"SDB"},
           "FPGA":{"Title":"FPGA on Sub PCB for DTV"},
           "DTV":{"Title":"TV + BEACON"},
           "SXM":{"Title":"SXM"},
           "ADR":{"Title":"ADR software"},
           "BOM":{"Title":"BOM-No. main board"},
           "Manufacturer \n(with Size)":{"Title":"UFS size \n[GB]"},
           "AURIX":{"Title":"AURIX"},
           "Type":{"Title":"UFS\nM = Micron"},
           "GNSS":{"Title":"GNSS"},
           "SoC":{"Title":"CPU type (SoC)"},
           "Device_type":{"Title":"GAS/NON GAS"},
           "SamplePhase":{"Title":"Sample phase"},
           "BU":{"Title":"Base Unit reference"},
           "Plant":{"Title":"Plant"},
           "Customer":{"Title":"Customer"},
           "Supplier Feed":{"Title":"Supplier Feed"}
}
Memory_data = {"Manufact":{"Title":"Feature Sheet"},
               "Size":{"Title":"Density"},
               "Sample":{"Title":"Used for"},
               "UFSType":{"Title":"UFS System Specification"}
}
DocuFilesToDownload = {'SetDefTmpl':{'ID':'1368235','version':'LATEST'},
                       'HWPNsAll':{'ID':'1340739','version':'LATEST'}
                       }
Sample = {"DSB":"B-Sample","TSB":"C-Sample","PP":"D-Sample","MP":"D-Sample","SOP":"D-Sample"}
SplTmpl_data = {}
SplInfo = ["Manufacturer","Blank device\nPart No.","Memory size","Position"]
PartnoInfo = {}
Customer = ""
#Read PNs as input from user
HelpText = sys.argv[0]+"[-h] -p Partnumbers(comma separated)\n"
arguments = argparse.ArgumentParser(usage=HelpText)
arguments.add_argument('-p',metavar = "",help="Partnumbers (comma separated)", required = True)
args = arguments.parse_args()

args.p = re.sub(r"\.","",args.p)
PNs = re.split(",",args.p)
digit = len(PNs[0])
for num in PNs:
    if (len(num) is not digit):
        print("Error: Cannot have 10 digit and 12 digit PNs in same release, check...")
        sys.exit()
print("PNs=",PNs)

FilesToBkup = os.listdir('.')
VM_ver = ''
for exl in FilesToBkup:
    if (("SetDefinition_CCS2_P2" in exl) or ("HW_PartNumbersAll_CCS2.xlsx" in exl)) and ("U_2" not in exl):
        bkup = re.split(r'\.',exl)[0] +"_"+ datetime.now().strftime("%y%m%d_%H%M%S")+".xlsx"
        cmd = "move "+exl+" "+bkup
        os.system(cmd)
    elif("Matrix_CCS2_" in exl):
        VM_ver_new = re.split('_',exl)[-1]
        VM_ver_new = re.split(r'\.',VM_ver_new)[0]
        VM_ver_new = re.sub('V','',VM_ver_new)
        if(VM_ver == '') or (int(VM_ver) < int(VM_ver_new)):
            VM_ver = VM_ver_new
            VM = cwd + "\\"+ exl
if(VM == ""):
    print("Error: VM is not available")
    sys.exit()
else:
    print("VM ver used: ",VM_ver)
#Copy templates and scripts from server
print("Copying the templates and scripts from",TemplateDir)
cmd = "xcopy " + TemplateDir + " " +cwd +" /Y /exclude:"+TemplateDir+"exclude_list.txt"+">> .//TmplCopy.log"
os.system(cmd)
print("Done")

#Get the PN details from Variant Matrix
wb = openpyxl.load_workbook(VM,data_only=True)
sheet = wb["Feature"]
sheet2 = wb["Memory_Component"]
LastRow=sheet.max_row
LastCol=sheet.max_column
LastRow_s2=sheet2.max_row
LastCol_s2=sheet2.max_column
row = 1
col = 1
while (row<=LastRow):
    Data = (sheet.cell(row,col)).value
    if (Data) and (re.search("Index",Data) is not None):
        HeadingRow = row
        break
    else:
        row += 1
for Key in VM_data.keys():
    col = 1
    while (col<=LastCol):
        data = (sheet.cell(HeadingRow,col)).value 
        SubHeading = (sheet.cell(HeadingRow+2,col)).value
        if (data is None):
            data = SubHeading
        elif (SubHeading is not None):
            data +=  SubHeading
        if (data) and (VM_data[Key]['Title'] in data):
            VM_data[Key]["ColNum"] = col
            break
        else:
            col += 1
#Read UFS type
row_s2 = 1
col_s2 = 1
while (row_s2<=LastRow):
    CellVal = (sheet2.cell(row_s2,col_s2)).value
    if (CellVal) and (re.search("Position",CellVal) is not None) and (re.search("D5871",(sheet2.cell(row_s2,col_s2+1)).value) is not None):
        HeadingRow_s2 = row_s2 + 1
        while (col_s2<=LastCol):
            rowHeading = (sheet2.cell(HeadingRow_s2,col_s2)).value
            for key2 in Memory_data.keys():
                if (rowHeading) and (Memory_data[key2]["Title"] in rowHeading):
                    Memory_data[key2]["ColNum"] = col_s2
            col_s2 += 1
        break
    else:
        row_s2 += 1

keyIndex_PN = -1
PartnoInfo["Additional Info - UFS type"]=[]
PartnoInfo["container"]=[]
PartnoInfo["Electrical"]=[]
for partno in PNs:
    keyIndex_PN += 1
    row = 1
    PN_Found = "no"
    while (row<=LastRow):
        PN_row =sheet.cell(row,VM_data['Product']['ColNum']).value
        if(PN_row is not None) and (str(PN_row) == str(partno)):
            for Key in VM_data.keys():
                if (Key not in PartnoInfo.keys()):
                    PartnoInfo[Key]=[]
                Value = sheet.cell(row,VM_data[Key]['ColNum']).value
                if (Key == "Device_type"):
                    if(Value):
                        if (Value.lower() == "gas"):
                            Value = "yes"
                        elif (Value.lower() == "non gas"):
                            Value = "non"
                    else:
                        Value = "n.a"
                elif ((Key == "AURIX") and (Value != "-")):
                    Value = "scc"
                elif ((Key == "ADR") and (Value.upper() == "FMAM")):
                    Value = "FM"
                else:
                    if (Value == "-" or Value =="" or Value == "Blue Line"):
                        Value = "no"
                    elif (Value == "x"):
                        if (Key == "GNSS"):
                            Value = re.split(' ',sheet.cell(row,VM_data[Key]['ColNum']+1).value)[0]
                        else:
                            Value = "yes"
                    elif (Value == 2 or Value ==4) and (Key == "DTV"):
                        Value ="yes"
                    elif ("SXM" in str(Value)) and (Key == "SXM"):
                        Value ="yes"
                PartnoInfo[Key].insert(keyIndex_PN,Value)
            SearchStart = HeadingRow_s2+1
            UFSValue = ""
            while (SearchStart<=LastRow_s2):
                typ = (sheet2.cell(SearchStart,Memory_data["Manufact"]["ColNum"])).value
                siz = (sheet2.cell(SearchStart,Memory_data["Size"]["ColNum"])).value
                samp= (sheet2.cell(SearchStart,Memory_data["Sample"]["ColNum"])).value
                Sample_PN = ""
                for SamPh in Sample.keys():
                    if(SamPh in PartnoInfo["SamplePhase"][keyIndex_PN]):
                        Sample_PN = Sample[SamPh]
                        break
                if (typ is PartnoInfo["Type"][keyIndex_PN]) and (siz is PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN]) and (samp is Sample_PN):
                    UFSValue= (sheet2.cell(SearchStart,Memory_data["UFSType"]["ColNum"])).value
                    break
                else:
                    SearchStart += 1
            if(UFSValue):
                UFSValue = re.sub(" ","",UFSValue)
                PartnoInfo["Additional Info - UFS type"].insert(keyIndex_PN,UFSValue)
            else:
                PartnoInfo["Additional Info - UFS type"].insert(keyIndex_PN,"UFS3.1")
            PN_Found = "yes"
            break
        row += 1
    if(PN_Found == "no"):
        print("\nError: ",partno," is not found in VM")
        sys.exit()
    if((len(PartnoInfo["Type"][keyIndex_PN]) > 1) and (PartnoInfo["Type"][keyIndex_PN][0].lower() == 's' or PartnoInfo["Type"][keyIndex_PN][0].lower() == 'm')):
        CTS = PartnoInfo["Type"][keyIndex_PN][0] + re.sub("GB","",PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN]) + PartnoInfo["Type"][keyIndex_PN][1:-1]
    else:
        CTS = PartnoInfo["Type"][keyIndex_PN] + re.sub("GB","",PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN])
    if(re.search("_UFS",CTS) is not None):
        CTS = re.sub("_UFS","",CTS)
    PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN] =CTS
    PartnoInfo["Electrical"].insert(keyIndex_PN,PartnoInfo["BOM"][keyIndex_PN])
    PartnoInfo["Electrical"][keyIndex_PN] = re.sub(r"\s",'',PartnoInfo["Electrical"][keyIndex_PN])
    PartnoInfo["Electrical"][keyIndex_PN] = re.sub("x",'4',PartnoInfo["Electrical"][keyIndex_PN])
    if(len(PartnoInfo["Electrical"][keyIndex_PN]) < 10):
        PartnoInfo["Electrical"][keyIndex_PN] = "8E38"+PartnoInfo["Electrical"][keyIndex_PN]
    PartnoInfo["Electrical"][keyIndex_PN] = PartnoInfo["Electrical"][keyIndex_PN][:4]+"."+PartnoInfo["Electrical"][keyIndex_PN][4:7]+"."+PartnoInfo["Electrical"][keyIndex_PN][7:]
    if(PartnoInfo["Electrical"][keyIndex_PN][1] != 'E'):
        tmplist = list(PartnoInfo["Electrical"][keyIndex_PN])
        tmplist[1] = 'E'
        PartnoInfo["Electrical"][keyIndex_PN]="".join(tmplist)
wb.close()


#Fill Pn details in the set definition
keyindex = 0
BU_EU = ""
for pn in PartnoInfo["Product"]:
    if(str(PartnoInfo["BU"][keyindex]) != str(pn)):
        if(PartnoInfo["Device_type"][keyindex] == "n.a"):
            print("\nWarning: Device_type of ",pn, "is not filled in VM")
        if(BU_EU == ""):
            BU_EU = "E"
        elif(BU_EU == "B"):
            print("Error: Cannot have EU and BU in the same release")
            sys.exit()
    elif(PartnoInfo["BU"][keyindex] is not None):
        if(BU_EU == ""):
            BU_EU = "B"
        elif(BU_EU == "E"):
            print("Error: Cannot have EU and BU in the same release")
            sys.exit()
    if(PartnoInfo["Customer"][keyindex] != "no"):
        if(Customer == ""):
            Customer = PartnoInfo["Customer"][keyindex]
        elif(Customer != PartnoInfo["Customer"][keyindex]):
            print("Error: Cannot have ",Customer,PartnoInfo["Customer"][keyindex],"PNs in the same release")
            sys.exit()
    keyindex += 1
SetDefName = re.sub(r"\[X\]",BU_EU,SetDefName)
cmd = "move " + SetDefTmpl + " " +SetDefName
os.system(cmd)

#Fill the details in set definition
book = openpyxl.load_workbook(SetDefName,read_only=False,data_only=True)
Valid = book["ValidLists"]
LastRow=Valid.max_row
LastCol=Valid.max_column
col = 1
Valid_HeadingRow = 1
ValidList = {}
while (col<=LastCol):
    data = (Valid.cell(Valid_HeadingRow,col)).value
    SubHeading = (Valid.cell(Valid_HeadingRow+1,col)).value

    if (data is None):
        if(col !=  1):
            data = (Valid.cell(Valid_HeadingRow,col-1)).value
        else:
            data = ""
    if (SubHeading is not None):
        data += "_" + SubHeading
    row = 3
    Val = []
    while (row<=LastRow and (Valid.cell(row,col)).value is not None):
        Val.append((Valid.cell(row,col)).value)
        row += 1
    ValidList[data] = Val
    col += 1

Setdef = book["Set Definition"]
LastRow=Setdef.max_row
LastCol=Setdef.max_column
row = 1
col = 1
PNrow = LastRow+1

while (row<=LastRow):
    Data = (Setdef.cell(row,col)).value
    if (Data):
        if (re.search("Purpose of delivery",Data) is not None):
            if (re.search("VM ver",(Setdef.cell(row,col+2)).value) is not None):
                Setdef.cell(row,col+2,"V"+VM_ver)
        elif (re.search("Definition Identifier",Data) is not None):
            Setdef.cell(row,col+1,"P2_xx_"+BU_EU+"U")
    if (Data) and (re.search("Part of delivery",Data) is not None):
        PNHeadingRow = row
        break
    else:
        row += 1

keyIndex_PN = -1
for partno in PNs:
    keyIndex_PN += 1
    col = 1
    while (col<=LastCol):
        MainTitle = Setdef.cell(PNHeadingRow,col).value
        SubTitle = Setdef.cell(PNHeadingRow+1,col).value
        itr=1
        while(MainTitle is None):
            # if(col is not 1):
            MainTitle = (Setdef.cell(PNHeadingRow,col-itr)).value
            itr += 1
        if("Part of delivery" in MainTitle):
            Setdef.cell(PNrow,col,"x")
        elif(MainTitle == "Set"):
            if(SubTitle == "Scope"):
                Setdef.cell(PNrow,col,Scope)
            elif(SubTitle == "Name"):
                Container = Scope + "_" + PartnoInfo["SDB"][keyIndex_PN] + "_" + PartnoInfo["ADR"][keyIndex_PN]  + "_" + PartnoInfo["SoC"][keyIndex_PN]  + "_" +PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN]
                Container = Container.upper()
                if(Container in PartnoInfo["container"]):
                    Container += "_" + str(keyIndex_PN)
                PartnoInfo["container"].insert(keyIndex_PN,Container)
                Setdef.cell(PNrow,col,Container)
            elif("Product" in SubTitle):
                Setdef.cell(PNrow,col,PartnoInfo["Product"][keyIndex_PN])
        else:
            Heading=""
            if(MainTitle in PartnoInfo.keys()):
                Heading = MainTitle
            elif(SubTitle in PartnoInfo.keys()):
                Heading = SubTitle
            if (SubTitle is not None):
                MainTitle += "_" + SubTitle
            ValueToFill =""
            if(Heading):
                if ("Manufacturer" in Heading):
                    PartnoInfo[Heading][keyIndex_PN] = PartnoInfo[Heading][keyIndex_PN] + "_10"
                if("SPL-ID" in MainTitle):
                    ValueToFill = ValidList[MainTitle][-1]
                elif("BU" in Heading):
                    ValueToFill = PartnoInfo['BU'][keyIndex_PN]
                else:
                    for Validdata in ValidList[MainTitle]:
                        if(str(PartnoInfo[Heading][keyIndex_PN]).lower() in str(Validdata).lower()):
                            ValueToFill = Validdata
                            break
                if (ValueToFill != ""):
                    Setdef.cell(PNrow,col,ValueToFill)
                else:
                    print("\nWarning:",Heading,PartnoInfo[Heading][keyIndex_PN], "not in Validlist")
            
        col += 1
    PNrow += 1
book.save(SetDefName)
book.close()
print("\nSet definition is generated -",SetDefName)
#Read SPL template data
SPLTmpl = openpyxl.load_workbook(SPL_tmpl,data_only=True)
SPLSheet = SPLTmpl['SPL']
LastRow_spl=SPLSheet.max_row
LastCol_spl=SPLSheet.max_column
row_spl = 1
col_spl = 1
Startrow_spl = 1
while(row_spl <= LastRow_spl):
    val = (SPLSheet.cell(row_spl,col_spl)).value
    if (val is not None) and ("Download file/Part No" in val):
        Startrow_spl = row_spl
        break
    else:
        row_spl += 1
row_spl = Startrow_spl+1
pnIndex = -1
for ufstyp in PartnoInfo["Manufacturer \n(with Size)"]:
    pnIndex += 1
    UFSid = PartnoInfo["Additional Info - UFS type"][pnIndex]+"_"+ufstyp
    SplTmpl_data[UFSid]={}
SplTmpl_data[ValidList["AURIX_SPL-ID"][-1]]={}
SplTmpl_data[ValidList["GNSS_SPL-ID"][-1]]={}

while(row_spl <= LastRow_spl):
    col_spl = 1
    SplId = SPLSheet.cell(row_spl,col_spl).value
    SplId = re.sub(" TMPL","",str(SplId))
    if(SplId in SplTmpl_data.keys()):
        while(col_spl <= LastCol_spl):
            for rowheading in [0,2]:
                if(SPLSheet.cell(row_spl+rowheading,col_spl).value in SplInfo):
                    heading = SPLSheet.cell(row_spl+rowheading,col_spl).value
                    heading = re.sub("\n"," ",heading)
                    heading = re.sub(r"\.","",heading)
                    if("size" in heading):
                        heading="Type"
                    SplTmpl_data[SplId][heading] = SPLSheet.cell(row_spl+rowheading+1,col_spl).value
            col_spl += 1    
    row_spl += 4
SPLTmpl.close()

for Ids in SplTmpl_data.keys():
    if( len(SplTmpl_data[Ids].keys()) == 0):
        print("\nError:",Ids,"is not available in SPL tmpl")
        sys.exit()
PNsWb = openpyxl.load_workbook(HW_PNs_All,data_only=True)
HmPnsAll_ver=""
if(digit == 10):
    sheetname = "Prod_" + BU_EU
else:
    sheetname = "TSB_" + BU_EU

for sht in PNsWb.sheetnames:
    if(sheetname in sht):
        if((Customer == "MMC") and (Customer in sht)):
            sheetname = sht
        elif((Customer != "MMC") and ("MMC" not in sht)):
            sheetname = sht
    elif("V0" in sht):
        VerSheet = PNsWb[sht]
        tmp = re.split(r"\.",sht)
        tmp[-1] = int(tmp[-1])+1
        HmPnsAll_ver =str(tmp[0]) + "."+ str(tmp[1])
        VerSheet.title = HmPnsAll_ver
print("\nUpdating details in HW_PartNumbersAll_CCS2:")
PNsSheet = PNsWb[sheetname]
LastRow_PnsAll=PNsSheet.max_row
LastCol_PnsAll=PNsSheet.max_column
row_PnsAll = 2
col_PnsAll = 1
headingrow_PnsAll = 1
AvailPNs = {}
while (row_PnsAll <= LastRow_PnsAll):
    Partnum = PNsSheet.cell(row_PnsAll,col_PnsAll).value
    if(str(Partnum)[0:2] == "75"):
        Partnum = re.split(r"\(",str(Partnum))
        Partnum = Partnum[0].strip()
        AvailPNs[Partnum]=row_PnsAll
    row_PnsAll += 1

row_PnsAll = 2
keyIndex_PN = -1
for partno in PNs:
    keyIndex_PN += 1
    col_PnsAll = 1
    if(partno not in AvailPNs.keys()):
        PNsSheet.cell(LastRow_PnsAll+1,col_PnsAll,PartnoInfo["SamplePhase"][0] + " " + BU_EU + "U")
        while (col_PnsAll <= LastCol_PnsAll):
            Title = (PNsSheet.cell(headingrow_PnsAll,col_PnsAll)).value
            if (Title is not None):
                if("description" in Title.lower()):
                    TITLEs=[]
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll,col_PnsAll,PartnoInfo['SDB'][keyIndex_PN]+","+PartnoInfo['ADR'][keyIndex_PN]+","+PartnoInfo['SoC'][keyIndex_PN]+","+PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN]+"  ("+PartnoInfo['container'][keyIndex_PN]+")")
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,col_PnsAll+1,"CW"+str(date.today().year)[2:]+str(date.today().isocalendar()[1]))
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+3,col_PnsAll+1,"P2_xx_"+ BU_EU+"U")
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,col_PnsAll,"PD: Link")
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,col_PnsAll).hyperlink="https://inside-share-hosted-apps.bosch.com/DMS/GetDocumentService/Document.svc/GetDocumentURL?documentID=P13S152810-1302019018-12998"
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,col_PnsAll).style="Hyperlink"
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll,col_PnsAll-1,"uBlox")
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,col_PnsAll-1,"UFS")
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+2,col_PnsAll-1,"SCC")
                else:
                    TITLEs = re.split(r"\/",Title)
                pos = 0
                for title in TITLEs:
                    for KEY in PartnoInfo.keys():
                        if(title.lower() in KEY.lower()) or (KEY.lower() in title.lower()):
                            PNsSheet.cell(LastRow_PnsAll+row_PnsAll+pos,col_PnsAll,PartnoInfo[KEY][keyIndex_PN])
                    pos += 1
                if(Title in SplTmpl_data[ValidList["GNSS_SPL-ID"][-1]].keys()):
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll,col_PnsAll,SplTmpl_data[ValidList["GNSS_SPL-ID"][-1]][Title])
                    UFS_id = PartnoInfo["Additional Info - UFS type"][keyIndex_PN]+"_"+PartnoInfo["Manufacturer \n(with Size)"][keyIndex_PN]
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,col_PnsAll,SplTmpl_data[UFS_id][Title])
                    PNsSheet.cell(LastRow_PnsAll+row_PnsAll+2,col_PnsAll,SplTmpl_data[ValidList['AURIX_SPL-ID'][-1]][Title])
            col_PnsAll += 1 
        if(BU_EU == "E"):
            PNsSheet.cell(LastRow_PnsAll+row_PnsAll+1,1,"(BU:"+str(PartnoInfo['BU'][keyIndex_PN])+")")
        PNsSheet.cell(LastRow_PnsAll+row_PnsAll+2,1,PartnoInfo['BOM'][keyIndex_PN])
        row_PnsAll += 4
    else:
        print(partno," is already available")
        Existingrow = AvailPNs[partno]
        NewDescCol = 14
        while(PNsSheet.cell(Existingrow+1,NewDescCol).value):
            NewDescCol += 1
        PNsSheet.cell(Existingrow+1,NewDescCol,"CW"+str(date.today().year)[2:]+str(date.today().isocalendar()[1]))
        PNsSheet.cell(Existingrow+3,NewDescCol,"P2_xx_"+ BU_EU+"U")
                    
PNsWb.save(HW_PNs_All)
PNsWb.close()
print("Done...")