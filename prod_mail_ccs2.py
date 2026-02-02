#################################################################################################################
#
# FILE:         prod_mail_ccs2.py
# DESCRIPTION:  This script can be used to create the SAP request sheet for CCS2 project
# USAGE:        see help_text
# PREREQUISITE: The file prod_imail_ccs2.emtf has to be in the same directory as this script
# HISTORY:
# Date         | Author          		| Modification
# 29.08.2024   | Nisharani C  			| Initial version
# 18.09.2024   | Nisharani C  			| help section update
# 17.06.2025   | Abinaya Muthukrishnan  | Extended for Deliveries sheet update
# 26.06.2025   | Nisharani C            | Adaptation for Image file name change - SXN, DTV, Lontium
# 16.09.2025   | Abinaya Muthukrishnan  | Extended for CCS2_PN_EU_BU_mapping sheet update
#################################################################################################################


import os
import hashlib
import subprocess
import optparse
import sys
import openpyxl
import re
from openpyxl.styles import Font
import xml.etree.ElementTree as ET
import win32com.client as win32
import update_prod_deliveries_ccs2 as ob2
from datetime import datetime
# _startlines = ""

global _rel_version
global _sw_ver
global _Chng_NR



class _mail_content:
    
    def __init__(self, _infodict):
        self._target_path = _maildict["target"]
        
        self._tools_path = self._target_path.split("\\00_SW")[0 ] + "\\01_Tools\\production_tooling\\"  
        #print ("_tools_path: ",self._tools_path)
        
        self.startlines = "<font size = '10px'>" + "Hello All," + "<br> <br> Please find the SW and the documents in the ""_Production""-Area <br> " + f"<a href={self._target_path} >" +  _maildict["target"]  + '</a>' + "<br>" + "Note:" + "<br> <p> 1)	No pre-flash image for UFS flash <br> <br>" + "<b> SW Overview Sheets (SOS):</b> <br>" + "<b> SOS pdf-File (all sets): </b> <br>"
        #self._rel_name = _rel_name
        self._maildict = _infodict
        self._mailbody = ""
        self._SOS_sec = ""
        self._SPL_sec = "<br><br> <b> SPL: </b>"
        self._CTS = "<br><br><br> <b> CTS: </b>"
        self._AppGAS = "<br> <b> App-SW GAS: </b>"
        self._PDConf_sec = "<br> <b> PD Configuration: </b>" 
        self._CTSConf_sec = "<br> <b> CTS Configuration: </b>"
        self._UCBConf_sec = "<br> <b> UCB Configuration: </b>"
        self._PT_sec = "<br> <b> Production tooling : </b>"
        self._AppNonGAS = "<br> <b> App-SW Non-GAS: </b>"
        self._AppGas_Status = "no"
        self._AppNonGas_Status = "no"
        self.P_SW = "<br> <b> Product SW </b>"
        self._aurix_srec_sec = "<br> Aurix SREC: "
        self._aurix_dnl_sec = "Aurix DNL: "
        self._aurix_sec = ""
        self._ublox_sec = "<br> Ublox : "
        self._sxm_sec = "<br> SXM : "
        self._dtv_sec = "<br> DTV : " 
        self._lont_sec = "<br> Lontium : " 
        self._product_sec = ""
        self._prod_dir = "Product_SW"
        #self._sub = _maildict["_rel_name"]
        self._sub = "[ CCS2 ] SW " + _maildict["rel_version"] + " for " + _maildict["_rel_name"] + "released to production for PNs: " 
        self._PNs = ""
        self._req_to_splteam = "<br> <br> @SPL team: please check SPL and update material master in SAP. tnx! <br> <br>"
        self._to_string = ""
        self._cc_string = ""
        #def _mail_frame():
        self.html_tag    = {
                    "HB" : "<HTML><BODY><FONT face=Arial monospaced for SAP size=2><STYLE type=\"text/css\">.cns {font-size: 9pt}</STYLE><DIV></FONT><FONT face='Arial' size=2 color=#000000>",
                    "FE" : "</FONT><FONT face='Arial' size=2 color=#000000></DIV></PRE><BR><BR></FONT></FONT></BODY></HTML>",
                    "AS" : "</FONT><FONT face='Arial' size=2 color=#000000>"           , # Arial Schwarz
                    "AR" : "</FONT><FONT face='Arial' size=2 color=#ff0000>"           , # Arial Red
                    "SN" : "</STRONG>"                                                 , # Style Normal
                    "SB" : "<STRONG>"                                                  , # Style Bold
                    "PB" : "<DIV>"                                                     , # Paragraph Begin
                    "PE" : "</DIV>"                                                    , # Paragraph End
                    "TB" : "<PRE>"                                                     , # White Space Preservation Begin
                    "TE" : "</PRE>"                                                    , # White Space Preservation End                   "IB" => "<FONT size=4>&#8226;</FONT>"                               , # Bullet
                    "BR" : "<BR>"                                                      , # Line
                    "IS" : "&nbsp;"                                                    , # Space
                    "IT" : "&#0009;"                                                   , # Tabulator (needs White Space Preservation)
                    "LB" : "<a href='"                                                 , # Link Begin
                    "LM" : "'>"                                                        , # Link Middle
                    "LE" : "</a>"                                                      , # Link End
                    "UB" : "<U>"                                                      , # Underline Begin
                    "UE" : "</U>"                                                     , # Underline End
                };
            #return html_tag
    
    #def _compose_mail(text, subject, recipient):
    def _compose_mail(self):
        recipient = "paramba.nisharanic@in.bosch.com"
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        #mail.To = recipient
        mail.To = self._to_string
        mail.CC = self._cc_string
        _PNs_list = []
        #mail.Subject = subject
        _sosall_path = _maildict["target"] + "\\" + _maildict["ecn_ecr"] + "_" + _maildict["rel_version"] + "_SOS-CCS2.pdf"
        _sosall_sec = f"<a href = {_sosall_path} >" + _sosall_path + "</a>"
        
        
        for x in _set_dict:
            _setKeys = _set_dict[x].keys()
            _SAP_PN = _set_dict[x]["document_number"]
            _Bosch_PN = _set_dict[x]["part_number"]
            _PN_formatted = _Bosch_PN[0:1] + "." + _Bosch_PN[1:4] + "." + _Bosch_PN[4:7] + "." +  _Bosch_PN[7:]
            #self._PNs = self._PNs + "," + _PN_formatted
            #print ("formatted :", _PN_formatted)
            
            _PNs_list.append(_PN_formatted)
            _dtv_sec = ""
            _lont_sec = ""
            _sxm_sec = ""
            
            
            _set_no = "SET" + _set_dict[x]["set_number"]
            _sos_filename = _SAP_PN + "_" + _maildict["rel_version"] + "_SOS-CCS2" + ".pdf" 
            _spl_filename = _SAP_PN + "_" + _maildict["rel_version"] + "_SPL-CCS2_" + _set_dict[x]["product"] + "_" + _set_no + ".pdf"
            _cont_path = _maildict["target"] + "\\" + "SW_Container_" + _set_dict[x]["product"]
            _sos_path = _maildict["target"] + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + "Data_to_plant" + "\\" + _sos_filename
            _spl_path = _maildict["target"] + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + "Data_to_plant" + "\\" + _spl_filename
            self._SOS_sec = self._SOS_sec + self.html_tag["BR"] + "SW_Container_" + _set_dict[x]["product"] + self.html_tag["BR"] + f"<a href = {_sos_path} >" + _sos_path + "</a>"
            self._SPL_sec = self._SPL_sec + self.html_tag["BR"] + "SW_Container_" + _set_dict[x]["product"] + self.html_tag["BR"] + f"<a href = {_spl_path} >" + _spl_path + "</a>"
            _aurix_dnl = _maildict["target"] + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + self._prod_dir + "\\" +_set_dict[x]["_aurix_dnl_filename"]
            _aurix_dnl_sec = self.html_tag["BR"] + "SW_Container_" + _set_dict[x]["product"] + "<br>" + self._aurix_dnl_sec + f"<a href = {_aurix_dnl} >" + _aurix_dnl + "</a>"
            _aurix_sec = _aurix_dnl_sec
            if "_aurix_srec_filename" in _setKeys:
                _aurix_srec = _maildict["target"]  + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + self._prod_dir + "\\" + _set_dict[x]["_aurix_srec_filename"]
                _aurix_srec_sec  = self._aurix_srec_sec + f"<a href = {_aurix_srec} >" + _aurix_srec + "</a>"
                _aurix_sec = _aurix_sec + _aurix_srec_sec
            #elif "_aurix_srec_support_filename" in _setKeys:
                # _aurix_srec = _maildict["target"] + "\\Support_Files\\" + _set_dict[x]["_aurix_srec_support_filename"]
            #self._aurix_sec = _aurix_sec
            if "_sxm_filename" in _setKeys:
                _sxm_file = _maildict["target"]  + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + self._prod_dir + "\\" + _set_dict[x]["_sxm_filename"]
                _sxm_sec = self._sxm_sec + f"<a href = {_sxm_file} >" + _sxm_file + "</a>"
            if "_dtv_swfilename" in _setKeys:
                _dtv_file = _maildict["target"]  + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + self._prod_dir + "\\" + _set_dict[x]["_dtv_swfilename"]
                _dtv_sec = self._dtv_sec + f"<a href = {_dtv_file} >" + _dtv_file + "</a>"

            if "_lont_filename" in _setKeys:
                _lont_file = _maildict["target"]  + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + self._prod_dir + "\\" + _set_dict[x]["_lont_filename"]
                _lont_sec = self._lont_sec + f"<a href = {_lont_file} >" + _lont_file + "</a>"
                
            _ublox_dnl = _maildict["target"] + "\\" + "SW_Container_" + _set_dict[x]["product"] + "\\" + self._prod_dir + "\\" + _set_dict[x]["_ublox_filename"]
            _ublox_sec = self._ublox_sec + f"<a href = {_ublox_dnl} >" + _ublox_dnl + "</a>"
            
            self._product_sec = self._product_sec + _aurix_sec + _ublox_sec + _sxm_sec + _dtv_sec + _lont_sec           
            
            
            if "_app_gfilename" in _setKeys and self._AppGas_Status == "no":
                _gasfile_path = _maildict["target"] + "\\QFIL_Download\\" + _set_dict[x]["_app_gfilename"]
                self._AppGAS =  self._AppGAS + self.html_tag["BR"] + f"<a href = {_gasfile_path} >" + _gasfile_path + "</a>"
                self._AppGas_Status = "yes"
                #print ("\n _gasfile_path : ", _gasfile_path)
            if "_app_ngfilename" in _setKeys and self._AppNonGas_Status == "no": 
                _ngasfile_path = _maildict["target"] + "\\QFIL_Download\\" + _set_dict[x]["_app_ngfilename"]
                self._AppNonGAS =  self._AppNonGAS + self.html_tag["BR"] + f"<a href = {_ngasfile_path} >" + _ngasfile_path
                self._AppNonGas_Status = "yes"
                #print ("\n _ngasfile_path : ", _ngasfile_path)
            
        _cts_path = _maildict["target"] + "\\QFIL_Download\\" + _set_dict[x]["_cts_filename"]    
        self._CTS = self._CTS + self.html_tag["BR"] + f"<a href = {_cts_path} >" + _cts_path + "</a>"
        self._mailbody = self.startlines + _sosall_sec + self._SOS_sec + self._SPL_sec + self._CTS 
        
        if self._AppGas_Status == "yes":
            self._mailbody = self._mailbody + self._AppGAS
        if self._AppNonGas_Status == "yes":
            self._mailbody = self._mailbody + self._AppNonGAS
        
        _PDConf_path = _maildict["target"] + "\\PD_Configuration\\" + _set_dict[x]["_PD_filename"]    
        self._PDConf_sec = self._PDConf_sec + self.html_tag["BR"] + f"<a href = {_PDConf_path} >" + _PDConf_path + "</a>"
        
        _CTSConf_path = _maildict["target"] + "\\CTS_Configuration\\" + _set_dict[x]["_cts_config_filename"]    
        self._CTSConf_sec = self._CTSConf_sec + self.html_tag["BR"] + f"<a href = {_CTSConf_path} >" + _CTSConf_path + "</a>"
        
        _UCBConf_path = _maildict["target"] + "\\UCB_Configuration\\" + _set_dict[x]["_ucb_config_filename"]    
        self._UCBConf_sec = self._UCBConf_sec + self.html_tag["BR"] + f"<a href = {_UCBConf_path} >" + _UCBConf_path + "</a>"
        
        _PTfile_path = self._tools_path + _maildict["_PT_cfs_version"] + "\\" + _set_dict[x]["_CFS_filename"]    
        self._PT_sec = self._PT_sec + self.html_tag["BR"] + f"<a href = {_PTfile_path} >" + _PTfile_path + "</a>"
        
        _PSW_sec = "<br> <br> <b>" + self.P_SW + " </b>"
        self._mailbody = self._mailbody + self._PDConf_sec + self._CTSConf_sec + self._UCBConf_sec +  self._PT_sec + _PSW_sec + self._product_sec + self._req_to_splteam
        
        #_PNs_list = ["7.503.751.812","7.513.752.496","7.513.752.762","7.513.752.772","7.513.752.763","7.503.751.813","7.513.752.774"]	
        #print ("_PNs_list :", _PNs_list)
        _pns_mail = ""
        _i = 0
        #_pns_mailSub = _PNs_list[_i]
        
        _PNs_list.sort()
        #print ("_PNs_list :", _PNs_list)
        _len = len(_PNs_list)
        _pns_mailSub = _PNs_list[_i]
        
        if len(_PNs_list) > 1:
            while _i < _len:
            #for _pn in _PNs_list:
                _curr_pn = _PNs_list[_i]
                _curr_pn_7 = _curr_pn[0:9]
                _next_index = _i + 1
                if _next_index >= _len :
                    break
                else:
                    _next_pn = _PNs_list[_i+1]
                    _next_pn_7 = _next_pn[0:9]
                if _curr_pn_7 == _next_pn_7 :
                    _next_pn_8to = _next_pn[8:].replace(".","")
                    _pns_mailSub = _pns_mailSub + "," + _next_pn_8to
                else:
                    _pns_mailSub = _pns_mailSub + "," + _next_pn
                #_curr_pn = _next_pn
                _i += 1
        
        else:
            _pns_mail = _pns_mailSub
        print("\n pns for mail",  _pns_mailSub )  
        
        
        _set_def_ref = " (" + _maildict["_set_def"] + ")"
        subject = self._sub + _pns_mailSub + _set_def_ref
        mail.Subject = subject
        mail.HtmlBody = self._mailbody
        mail.Display(True)


    def _get_recepient_list(self):
        _to_list = []
        _cc_list = []
        _to_string = ""
        _cc_string = ""
        _rtemplate = "prod_imail_ccs2.emtf"
        f1 = open(_rtemplate, "r")
        print("****************************************************************************************************************************************")
        print ("\n using the template file:", _rtemplate, "to form the recepient list , please make sure that the template is the latest") 
        _userinput = input("\n Please confirn to proceed : Y/N \t").upper()
    
        if _userinput == "Y":  
            for _line in f1:
                #print ("line :", _line)
                if "TO:" in _line:
                    _to_address = _line.split("TO:")[-1].strip()
                    #print ("_to_address :", _to_address)
                    #self._to_list .append(_to_address)
                    _to_list .append(_to_address)
                    
                if "CC:" in _line:
                    #_cc_address = _line.split("CC:")[-1].strip()
                    _cc_address = (_line.split("CC:")[-1]).split("#")[0].strip()
                    # _cc_address = (_line.split("CC:")[-1]).strip()
                    #print ("_cc_address :", _cc_address)
                    #self._cc_list.append(_cc_address)
                    _cc_list.append(_cc_address)
            
            # print ("\n _to_list :", _to_list ) 
            # print ("\n _cc_list :", _cc_list )   
            for _i in range(len(_to_list)):
                _to_string = _to_string + _to_list[_i] + ";"
            
            for _i in range(len(_cc_list)):
                _cc_string = _cc_string + _cc_list[_i] + ";"
            
            # print ("\n _to_string :", _to_string)
            # print ("\n _cc_string :", _cc_string)
            #self.recepient_list = [_to_string, _cc_string]
            self._to_string = _to_string
            self._cc_string = _cc_string
            #return ( self.recepient_list )
            
        else:
            sys.exit(0)
        f1.close()

def parse_sap_xml(_sap_xml):
    print ("\n at parse sap xml .. ")
    global _set_dict
    global _maildict
    _maildict = {}
    tree = ET.ElementTree()
    tree.parse(_sap_xml)
    root = tree.getroot()
    
                  
    _doc_sec = root.find("./DocInfo")
    _doc_info = _doc_sec.attrib
    _set_def = _doc_sec.attrib["Doc_SetDef"]
    _maildict["_set_def"] = _set_def
    
    _sxm_sw_ver = root.find("./SW_Versions[@key='SXM']")
    _sxm_info = _sxm_sw_ver.attrib
    
    _ecn =  root.find("./Overall_Infos[@Col1='ecn']")
    _ecn_no = _ecn.attrib["Col2"]
    
    _ecr =  root.find("./Overall_Infos[@Col1='ecr']")
    _ecr_no = _ecr.attrib["Col2"]

    if _ecr_no != "" and _ecr_no !="None":
        _Chng_NR = _ecr_no
    elif _ecn_no !="" and _ecn_no != "None":
        _Chng_NR = _ecn_no
    _Chng_NR = str(_Chng_NR) 

    _maildict["ecn_ecr"] = _Chng_NR
    
    
    _cfs = root.find("./SW_Versions[@key='Prod_Tooling']")
    _PT_cfs_version = _cfs.attrib["version"]
    _maildict["_PT_cfs_version"] = _PT_cfs_version
    
    _pd_file = root.find("./Overall_Infos[@Col1='kds_pd_file']")
    _PD_Config_zip = _pd_file.attrib["Col2"]
    
    _sw_id_gas = root.find("./Overall_Infos[@Col1='swid_gas']")
    _SWID_gas = _sw_id_gas.attrib["Col2"]
    
    _sw_id_nongas = root.find("./Overall_Infos[@Col1='swid_nongas']")
    _SWID_nongas = _sw_id_nongas.attrib["Col2"]
    
    _release_type = root.find("./Overall_Infos[@Col1='purpose']")
    _rel_name = _release_type.attrib["Col2"]
    _maildict["_rel_name"] = _rel_name    

    _sw_ver = root.find("./Overall_Infos[@Col1='sw_full_ver']")
    _rel_version = _sw_ver.attrib["Col2"]
    _maildict["rel_version"] = _rel_version
    
   
    _tar_dir = root.find("./Overall_Infos[@Col1='target_dir']")
    _target = _tar_dir.attrib["Col2"]
    _maildict["target"] = _target

    _tar_prod_dir = root.find("./Overall_Infos[@Col1='target_prod_dir']")
    _target_prod  = _tar_prod_dir.attrib["Col2"]
    _maildict["target_prod"] = _target_prod
    
    _set_dict ={}
    i=0    
    for SET in root.iter("SET_Infos"):    
        _SET_Infos_all = SET.attrib
        _set_dict[i] = _SET_Infos_all
        _PN =_set_dict[i]["part_number"]   
        i+=1
       
         
#Function to update the SW deliveries sheet         
def update_Deliveries_sheet():
    Del_sht = _maildict["target_prod"]+"\Reference_sheets\RN_CCS2_SW_Deliveries.xlsx";
    #backup copy to working directory
    cmd = "copy " + Del_sht + " "+ os.getcwd()
    os.system(cmd)
    Wb = openpyxl.load_workbook(Del_sht, data_only=True)
    Customer = "MMC_CCS2"
    if Customer not in _maildict["rel_version"]:
        Customer = "Nissan_CCS2"
    SWver = _maildict["rel_version"].split('_')
    SheetName = Customer + "_"
    Env = ""
    PN = _set_dict[0]["part_number"].replace(".","")
    if (len(PN) == 10) and ("-" not in PN):
        Env = "PreProd"
        SheetName = SheetName + Env + "_"
    SheetName = SheetName + SWver[-1]
    SelectedSheet = Wb[SheetName]
    StartRow = SelectedSheet.max_row + 1
    Col = 1
    PNsList = ""
    for Set in _set_dict:
        SelectedSheet.cell(StartRow, Col, _maildict["_rel_name"])
        SelectedSheet.cell(StartRow, Col+1, _set_dict[Set]["part_number"])
        SelectedSheet.cell(StartRow, Col+2, _maildict["rel_version"])
        SelectedSheet.cell(StartRow, Col+2).hyperlink=_maildict["target"]
        SelectedSheet.cell(StartRow, Col+2).style = "Hyperlink"
        SelectedSheet.cell(StartRow, Col+3, _set_dict[Set]["container_name"])
        ContainerLink = _maildict["target"] + "\SW_Container_" + _set_dict[Set]["container_name"]
        SelectedSheet.cell(StartRow, Col+3).hyperlink=ContainerLink
        SelectedSheet.cell(StartRow, Col+3).style = "Hyperlink"
        SelectedSheet.cell(StartRow, Col+4, _maildict["_set_def"])
        if (PNsList != ""):
            PNsList = PNsList + "," +(_set_dict[Set]["part_number"].replace(".",""))[6:]
        else:
            PNsList = _set_dict[Set]["part_number"][0] + "."+_set_dict[Set]["part_number"][1:4]+"."+_set_dict[Set]["part_number"][4:7]+"."+_set_dict[Set]["part_number"][7:]
        StartRow += 1
    HistSheet = Wb["History"]
    Row = 1
    RowToFill = Row
    Date = datetime.now().strftime("%m/%d/%y")
    while (Row < HistSheet.max_row):
        if (HistSheet.cell(Row, Col).value == "Comments"):
            RowToFill = Row+2
            HistSheet.insert_rows(Row+1)
        if (HistSheet.cell(Row, 3).value) and (HistSheet.cell(Row, 3).value != "Version"):
            LastUsedVer = HistSheet.cell(Row, 3).value
            break
        Row += 1
    HistSheet.cell(RowToFill, Col, "Added [CCS2] SW " + _maildict["rel_version"] + " for " + _maildict["_rel_name"] + " released to production for PNs: " +PNsList+"("+_maildict["_set_def"]+")")
    HistSheet.cell(RowToFill, Col).font = Font(name='Calibri',bold=False,size=11)
    HistSheet.cell(RowToFill, Col).alignment = openpyxl.styles.Alignment(wrapText=True)
    HistSheet.cell(RowToFill, Col+1, Date)
    HistSheet.cell(RowToFill, Col+1).font = Font(name='Calibri',bold=False,size=11)
    HistSheet.cell(RowToFill, Col+2, float(LastUsedVer)+0.01)
    HistSheet.cell(RowToFill, Col+2).font = Font(name='Calibri',bold=False,size=11)
    HistSheet.cell(RowToFill, Col+3, os.getlogin())
    HistSheet.cell(RowToFill, Col+3).font = Font(name='Calibri',bold=False,size=11)
    Wb.save(Del_sht)
    Wb.close()
    print("\nRN_CCS2_SW_Deliveries sheet is updated")

def update_EU_BU_Mapping_Sheet():
    global VM_ver_used
    global Inv_ID
    Mapping_sht = _maildict["target_prod"]+"\Reference_sheets\CCS2_PN_EU_BU_mapping.xlsx";
    #backup copy to working directory
    cmd = "copy " + Mapping_sht + " "+ os.getcwd()
    os.system(cmd)
    JIRA_ID = input("\n Please provide the main task ID (eg:AIVI-147997) : \t").upper()
    if(JIRA_ID):
        JIRA_link = "https://rb-tracker.bosch.com/tracker05/browse/"+JIRA_ID
    else:
        print("\n Task cannot be hyperlinked in the CCS2_PN_EU_BU_mapping sheet as task ID is not given");
        JIRA_link =""
    EU_BU_Wb = openpyxl.load_workbook(Mapping_sht,data_only=True)
    for sheets in EU_BU_Wb.sheetnames:   
        if("V" in sheets):
            VerSheet = EU_BU_Wb[sheets]
            tmp = re.sub('V','',sheets)
            New_ver =float(tmp)+0.01
            New_ver = round(New_ver,2)
            VerSheet.title = 'V'+str(New_ver)
    EU_BU_sheet = EU_BU_Wb["Sheet1"]
    LastRow=EU_BU_sheet.max_row
    LastCol=EU_BU_sheet.max_column
    Row = LastRow
    EU_Col = 1
    BU_Col = 6
    if ("_EU" in _maildict["rel_version"]):
        Col = EU_Col
    else:
        Col = BU_Col
    while(Row >1):
        if (EU_BU_sheet.cell(Row,Col).value is None):
            Row -= 1
        else:
            break
    Row += 1

    for Set in _set_dict:
        EU_BU_sheet.cell(Row,Col,_maildict["_rel_name"])
        if (JIRA_link):
            EU_BU_sheet.cell(Row,Col).hyperlink = JIRA_link
            EU_BU_sheet.cell(Row,Col).style = "Hyperlink"
        EU_BU_sheet.cell(Row,Col+1,_set_dict[Set]["part_number"])
        if ("_EU" in _maildict["rel_version"]):
            EU_BU_sheet.cell(Row,Col+2,_set_dict[Set]["BU"])
            if (_set_dict[Set]["device_type"] == "yes"):
                EU_BU_sheet.cell(Row,Col+3,"GAS")
            elif (_set_dict[Set]["device_type"] == "non"):
                EU_BU_sheet.cell(Row,Col+3,"NON GAS")
        Row +=1
    EU_BU_Wb.save(Mapping_sht)
    EU_BU_Wb.close()
    
    
if __name__ == "__main__":
    parser = optparse.OptionParser()
    parser.add_option('-x', '--xml', dest='xml', default=None,
                      help='Master xml')
   
    (options, args) = parser.parse_args()
    if not options.xml:
        #input_masterxml = input("\nPlease enter masterxml filename\n")
        print ("\n Please enter the master xml name as input ")
        print ("\n please run the script in the form : ")
        print ("\t prod_mail_ccs2.py -x <sap_xml_name> ")
        sys.exit(0) 
    else:
        input_masterxml = options.xml
        _sap_xml = input_masterxml
        print ("\n Master xml ", _sap_xml )
        if _sap_xml.startswith("SAP_") == True and _sap_xml.endswith(".xml") == True :
            print ("\n Master xml ", _sap_xml," will be used ")
        else:
            print ("please enter the proper sap master xml file ")
            sys.exit(0)
            
    _curr_Dir = os.getcwd()   
    _xml_File = os.path.join(_curr_Dir,_sap_xml) 
    
    parse_sap_xml(_sap_xml)
   
    ob1 = _mail_content(_maildict)
    #ob1._compose_mail()
    ob1._get_recepient_list()
    #_rec_list = ob1._get_recepient_list()
    ob1._compose_mail()
    print ("\n Mail is drafted, please check the links before you send : "  )
    
    print("****************************************************************************************************************************************")
    _userinput = input("\n Please confirm to update ProdDeliveries_Overview_CCS2.xlsx and RN_CCS2_SW_Deliveries.xlsx : Y/N \t").upper()
    if _userinput == "Y":  
        subprocess.run(['python', 'update_prod_deliveries_ccs2.py', "-x", _sap_xml])
        update_Deliveries_sheet()
        update_EU_BU_Mapping_Sheet()
    else:
        sys.exit(0)
    